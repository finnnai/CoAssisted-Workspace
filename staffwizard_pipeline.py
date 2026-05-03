# © 2026 CoAssisted Workspace. Licensed under MIT.
"""StaffWizard daily-operations pipeline — pure-Python core.

The MCP tools in `tools/staffwizard.py` and the terminal scripts in
`scripts/build_*` and `scripts/refresh_*` both import from this module.
Putting the logic here means we have one source of truth for the four
pipeline steps:

    1. ingest_latest_report()   — Gmail → .xls in reports_dir/
    2. build_master_xlsx()      — reports_dir/*.xls → master_<window>.xlsx
    3. push_master_to_sheets()  → reports_dir/*.xls → live Master + Archive Sheets
    4. build_dashboards()       — master xlsx → HTML/JSON + Drive folder

Plus a thin send_dashboards_email() helper that packages the dashboards
into a self-contained zip and emails it to a recipient list.

Each function returns a structured dict so the MCP wrappers can JSON-
serialize without thinking about it. Errors raise specific exceptions
the wrappers translate into MCP-friendly error responses.

Configuration lives in config.staffwizard. See config.py for defaults.
"""

from __future__ import annotations

import datetime as _dt
import io
import json
import logging
import pathlib
import re
import zipfile
from collections import defaultdict
from typing import Any, Optional

import config
import labor_ingest

_log = logging.getLogger(__name__)


# =============================================================================
# Errors
# =============================================================================

class StaffWizardError(Exception):
    """Base for pipeline errors that should be surfaced to the operator."""


class NoReportFoundError(StaffWizardError):
    """Gmail search came back empty for a window the operator expected."""


class MasterMissingError(StaffWizardError):
    """A pipeline step needs the master xlsx but it doesn't exist yet."""


# =============================================================================
# Config helpers
# =============================================================================

def _cfg() -> dict:
    block = config.get("staffwizard", {}) or {}
    return block if isinstance(block, dict) else {}


def reports_dir() -> pathlib.Path:
    """Filesystem dir where downloaded .xls reports land."""
    p = _cfg().get("reports_dir") or "~/Developer/google_workspace_mcp/staffwizard_overall_reports"
    return pathlib.Path(p).expanduser().resolve()


def master_xlsx_path() -> pathlib.Path:
    p = _cfg().get("master_xlsx_path") or str(reports_dir() / "master_april_2026.xlsx")
    return pathlib.Path(p).expanduser().resolve()


def dashboards_dir() -> pathlib.Path:
    p = _cfg().get("dashboards_dir") or "~/Developer/google_workspace_mcp/dashboards"
    return pathlib.Path(p).expanduser().resolve()


def master_sheet_id() -> str:
    return _cfg().get("master_sheet_id") or ""


def archive_sheet_id() -> str:
    return _cfg().get("archive_sheet_id") or ""


def dashboards_drive_folder() -> str:
    return _cfg().get("dashboards_drive_folder") or "Surefox Daily Operations Dashboards"


def window_days() -> int:
    return int(_cfg().get("window_days") or 90)


# =============================================================================
# Step 1 — ingest_latest_report
# =============================================================================

_OVERALL_RE = re.compile(r"Overall Report - (\d{2})/(\d{2})/(\d{4})")


def ingest_latest_report(
    *,
    work_date: Optional[_dt.date] = None,
    target_dir: Optional[pathlib.Path] = None,
) -> dict:
    """Find the latest Overall Report email + download its .xls attachment.

    If `work_date` is given, find the report for that specific date.
    Otherwise grab the most recently-received one.

    Returns:
        {status, work_date, message_id, attachment_path, attachment_name,
         size_bytes, already_existed}
    """
    target_dir = target_dir or reports_dir()
    target_dir.mkdir(parents=True, exist_ok=True)

    # Local Gmail imports — keep them lazy so the pipeline can be imported
    # from non-MCP contexts (tests, etc.) without dragging in the auth chain.
    import auth  # noqa: F401  (used implicitly by gservices)
    import gservices

    gmail = gservices.gmail_service()

    # Build the query.
    q = 'from:noreply@staffwizard.com subject:"Overall Report"'
    if work_date:
        # Filter by the date string in the subject for a specific day.
        q += f' subject:"{work_date.strftime("%m/%d/%Y")}"'

    resp = gmail.users().messages().list(
        userId="me", q=q, maxResults=10,
    ).execute()
    messages = resp.get("messages") or []
    if not messages:
        raise NoReportFoundError(
            f"No Overall Report found for query: {q!r}"
        )

    # Newest first (Gmail returns in reverse chronological order).
    msg_id = messages[0]["id"]
    msg = gmail.users().messages().get(
        userId="me", id=msg_id, format="full",
    ).execute()

    # Pull the subject + work date.
    subject = ""
    for h in (msg.get("payload") or {}).get("headers") or []:
        if h.get("name", "").lower() == "subject":
            subject = h.get("value", "")
            break
    m = _OVERALL_RE.search(subject)
    if not m:
        raise NoReportFoundError(
            f"Latest message subject {subject!r} doesn't match "
            "'Overall Report - MM/DD/YYYY' pattern."
        )
    mm, dd, yyyy = m.groups()
    actual_date = _dt.date(int(yyyy), int(mm), int(dd))

    # Find the .xls attachment.
    parts = (msg.get("payload") or {}).get("parts") or []
    target_part = None
    for part in parts:
        filename = part.get("filename") or ""
        if filename.lower().endswith(".xls"):
            target_part = part
            break
    if not target_part:
        raise StaffWizardError(
            f"Overall Report message {msg_id} has no .xls attachment."
        )

    attachment_id = (target_part.get("body") or {}).get("attachmentId")
    if not attachment_id:
        raise StaffWizardError(
            f"Attachment {target_part.get('filename')} has no attachmentId."
        )

    att = gmail.users().messages().attachments().get(
        userId="me", messageId=msg_id, id=attachment_id,
    ).execute()
    import base64
    payload_b64 = att.get("data") or ""
    # Gmail uses URL-safe base64.
    payload = base64.urlsafe_b64decode(payload_b64.encode("ascii"))

    out_name = f"overall_report_{actual_date.isoformat()}.xls"
    out_path = target_dir / out_name
    already_existed = out_path.exists()
    out_path.write_bytes(payload)

    return {
        "status": "downloaded" if not already_existed else "overwritten",
        "work_date": actual_date.isoformat(),
        "message_id": msg_id,
        "attachment_name": target_part.get("filename"),
        "attachment_path": str(out_path),
        "size_bytes": len(payload),
        "already_existed": already_existed,
    }


# =============================================================================
# Step 2 — build_master_xlsx
# =============================================================================

DAILY_DETAIL_COLS = [
    "Date", "Job Number", "Job Description",
    "Employee #", "Employee Name", "Post",
    "Shift Start", "Shift End",
    "Reg Hours", "OT Hours", "DT Hours", "Total Hours",
    "Reg Cost $", "Holiday Cost $", "OT Cost $", "DT Cost $", "Total Cost $",
    "Billable Hours", "Billable $", "Margin $",
]


def _row_to_detail_dict(r: "labor_ingest.LaborRow") -> dict:
    return {
        "Date": r.work_date.isoformat() if r.work_date else "",
        "Job Number": r.job_number,
        "Job Description": r.job_description,
        "Employee #": r.employee_number,
        "Employee Name": r.employee_name,
        "Post": r.post_description,
        "Shift Start": r.shift_start,
        "Shift End": r.shift_end,
        "Reg Hours": r.hours,
        "OT Hours": r.overtime_hours,
        "DT Hours": r.doubletime_hours,
        "Total Hours": round(r.total_hours, 2),
        "Reg Cost $": r.dollars,
        "Holiday Cost $": r.holiday_dollars,
        "OT Cost $": r.overtime_dollars,
        "DT Cost $": r.doubletime_dollars,
        "Total Cost $": round(r.total_cost, 2),
        "Billable Hours": r.billable_hours,
        "Billable $": r.billable_dollars,
        "Margin $": round(r.margin, 2),
    }


def _parse_all_reports(src_dir: pathlib.Path) -> list[dict]:
    """Parse every overall_report_*.xls in src_dir → list of detail rows."""
    rows = []
    for f in sorted(src_dir.glob("overall_report_*.xls")):
        try:
            parsed = labor_ingest.parse_overall_report(f)
        except Exception as e:
            _log.warning("skipping %s: %s", f.name, e)
            continue
        for r in parsed.rows:
            if r.work_date:
                rows.append(_row_to_detail_dict(r))
    return rows


def build_master_xlsx(
    *,
    src_dir: Optional[pathlib.Path] = None,
    out_path: Optional[pathlib.Path] = None,
) -> dict:
    """Combine every .xls in src_dir into a single master xlsx with 3 tabs.

    Tabs: Daily Detail / Project Rollup / Daily Totals.

    Returns: {status, file_count, detail_rows, days, projects, output_path}
    """
    import openpyxl

    src_dir = src_dir or reports_dir()
    out_path = out_path or master_xlsx_path()
    if not src_dir.exists():
        raise StaffWizardError(f"Reports dir not found: {src_dir}")

    files = sorted(src_dir.glob("overall_report_*.xls"))
    if not files:
        raise StaffWizardError(f"No overall_report_*.xls files in {src_dir}")

    detail_rows = _parse_all_reports(src_dir)

    # Aggregates: per-(project, date) and per-date.
    pd_agg: dict = defaultdict(
        lambda: {"shifts": 0, "hours": 0.0, "cost": 0.0,
                 "revenue": 0.0, "margin": 0.0, "job_description": ""}
    )
    daily_agg: dict = defaultdict(
        lambda: {"shifts": 0, "hours": 0.0, "cost": 0.0,
                 "revenue": 0.0, "margin": 0.0,
                 "projects": set(), "employees": set()}
    )
    for d in detail_rows:
        date = d["Date"]
        job = d["Job Number"]
        if not date:
            continue
        a = pd_agg[(job, date)]
        a["shifts"] += 1
        a["hours"] += d["Total Hours"]
        a["cost"] += d["Total Cost $"]
        a["revenue"] += d["Billable $"]
        a["margin"] += d["Margin $"]
        a["job_description"] = d["Job Description"]
        t = daily_agg[date]
        t["shifts"] += 1
        t["hours"] += d["Total Hours"]
        t["cost"] += d["Total Cost $"]
        t["revenue"] += d["Billable $"]
        t["margin"] += d["Margin $"]
        t["projects"].add(job)
        if d["Employee #"]:
            t["employees"].add(d["Employee #"])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Daily Detail.
    detail = wb.create_sheet("Daily Detail")
    detail.append(DAILY_DETAIL_COLS)
    for d in detail_rows:
        detail.append([d[c] for c in DAILY_DETAIL_COLS])
    detail.freeze_panes = "A2"

    # Project Rollup.
    rollup = wb.create_sheet("Project Rollup")
    rollup.append([
        "Date", "Job Number", "Job Description",
        "Shifts", "Hours", "Cost $", "Revenue $", "Margin $", "Margin %",
    ])
    for (job, date), a in sorted(pd_agg.items(), key=lambda x: (x[0][1], x[0][0])):
        margin_pct = (a["margin"] / a["revenue"] * 100) if a["revenue"] else 0
        rollup.append([
            date, job, a["job_description"], a["shifts"],
            round(a["hours"], 2), round(a["cost"], 2),
            round(a["revenue"], 2), round(a["margin"], 2),
            round(margin_pct, 1),
        ])
    rollup.freeze_panes = "A2"

    # Daily Totals.
    totals = wb.create_sheet("Daily Totals")
    totals.append([
        "Date", "Shifts", "Unique Projects", "Unique Employees",
        "Total Hours", "Total Cost $", "Total Revenue $",
        "Total Margin $", "Margin %",
    ])
    for date, t in sorted(daily_agg.items()):
        margin_pct = (t["margin"] / t["revenue"] * 100) if t["revenue"] else 0
        totals.append([
            date, t["shifts"], len(t["projects"]), len(t["employees"]),
            round(t["hours"], 2), round(t["cost"], 2),
            round(t["revenue"], 2), round(t["margin"], 2),
            round(margin_pct, 1),
        ])
    totals.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    return {
        "status": "ok",
        "file_count": len(files),
        "detail_rows": len(detail_rows),
        "days": len(daily_agg),
        "projects": len({d["Job Number"] for d in detail_rows}),
        "output_path": str(out_path),
    }


# =============================================================================
# Step 3 — push_master_to_sheets (rolling 90d window)
# =============================================================================

# Column orders for the live Sheets — matches refresh_staffwizard_master.py.
SHEET_PROJECT_ROLLUP_COLS = [
    "Date", "Job Number", "Job Description",
    "Shifts",
    "Reg Hours", "OT Hours", "DT Hours", "Total Hours",
    "Reg Cost $", "Holiday Cost $", "OT Cost $", "DT Cost $", "Total Cost $",
    "Revenue $", "Margin $", "Margin %",
]
SHEET_DAILY_TOTALS_COLS = [
    "Date", "Shifts", "Unique Projects", "Unique Employees",
    "Reg Hours", "OT Hours", "DT Hours", "Total Hours",
    "Cost $", "Revenue $", "Margin $", "Margin %",
]
SHEET_PROJECT_TOTALS_COLS = [
    "Project", "Days Active", "Shifts", "Unique Employees",
    "Reg Hours", "OT Hours", "DT Hours", "Total Hours",
    "Cost $", "Revenue $", "Margin $", "Margin %",
]

_TAB_FORBIDDEN = re.compile(r"[\[\]\*\?:\\/]")


def _sanitize_tab_name(name: str) -> str:
    safe = _TAB_FORBIDDEN.sub("-", name).strip().strip("'")
    return safe[:100] or "Untitled"


def _pct(numer: float, denom: float) -> float:
    return round((numer / denom * 100), 1) if denom else 0.0


def _build_sheet_payload(detail_rows: list[dict]) -> dict[str, list[list]]:
    """Assemble all tabs (main + per-project) for a Sheet — same shape
    as refresh_staffwizard_master.py emitted."""
    if not detail_rows:
        return {
            "Daily Detail": [DAILY_DETAIL_COLS],
            "Project Rollup": [SHEET_PROJECT_ROLLUP_COLS],
            "Daily Totals": [SHEET_DAILY_TOTALS_COLS],
            "Daily Totals by Project": [SHEET_PROJECT_TOTALS_COLS],
        }

    detail = [DAILY_DETAIL_COLS] + [
        [r[c] for c in DAILY_DETAIL_COLS] for r in detail_rows
    ]

    # Per-(project, date) aggregate with hour AND cost type breakdown.
    pd_agg: dict = defaultdict(
        lambda: {
            "shifts": 0, "reg_h": 0.0, "ot_h": 0.0, "dt_h": 0.0, "total_h": 0.0,
            "reg_cost": 0.0, "holiday_cost": 0.0, "ot_cost": 0.0,
            "dt_cost": 0.0, "total_cost": 0.0,
            "revenue": 0.0, "margin": 0.0, "job_description": "",
        }
    )
    daily_agg: dict = defaultdict(
        lambda: {
            "shifts": 0, "projects": set(), "employees": set(),
            "reg_h": 0.0, "ot_h": 0.0, "dt_h": 0.0, "total_h": 0.0,
            "cost": 0.0, "revenue": 0.0, "margin": 0.0,
        }
    )
    proj_agg: dict = defaultdict(
        lambda: {
            "shifts": 0, "dates": set(), "employees": set(),
            "reg_h": 0.0, "ot_h": 0.0, "dt_h": 0.0, "total_h": 0.0,
            "cost": 0.0, "revenue": 0.0, "margin": 0.0,
        }
    )
    for d in detail_rows:
        date = d["Date"]
        job = d["Job Number"]
        if not date:
            continue
        a = pd_agg[(job, date)]
        a["shifts"] += 1
        a["reg_h"] += d["Reg Hours"]
        a["ot_h"] += d["OT Hours"]
        a["dt_h"] += d["DT Hours"]
        a["total_h"] += d["Total Hours"]
        a["reg_cost"] += d["Reg Cost $"]
        a["holiday_cost"] += d["Holiday Cost $"]
        a["ot_cost"] += d["OT Cost $"]
        a["dt_cost"] += d["DT Cost $"]
        a["total_cost"] += d["Total Cost $"]
        a["revenue"] += d["Billable $"]
        a["margin"] += d["Margin $"]
        a["job_description"] = d["Job Description"]
        t = daily_agg[date]
        t["shifts"] += 1
        t["projects"].add(job)
        if d["Employee #"]:
            t["employees"].add(d["Employee #"])
        t["reg_h"] += d["Reg Hours"]
        t["ot_h"] += d["OT Hours"]
        t["dt_h"] += d["DT Hours"]
        t["total_h"] += d["Total Hours"]
        t["cost"] += d["Total Cost $"]
        t["revenue"] += d["Billable $"]
        t["margin"] += d["Margin $"]
        proj = d.get("Job Description") or job or "(unlabeled)"
        pa = proj_agg[proj]
        pa["shifts"] += 1
        pa["dates"].add(date)
        if d["Employee #"]:
            pa["employees"].add(d["Employee #"])
        pa["reg_h"] += d["Reg Hours"]
        pa["ot_h"] += d["OT Hours"]
        pa["dt_h"] += d["DT Hours"]
        pa["total_h"] += d["Total Hours"]
        pa["cost"] += d["Total Cost $"]
        pa["revenue"] += d["Billable $"]
        pa["margin"] += d["Margin $"]

    rollup = [SHEET_PROJECT_ROLLUP_COLS]
    for (job, date), a in sorted(pd_agg.items(), key=lambda x: (x[0][1], x[0][0])):
        rollup.append([
            date, job, a["job_description"], a["shifts"],
            round(a["reg_h"], 2), round(a["ot_h"], 2),
            round(a["dt_h"], 2), round(a["total_h"], 2),
            round(a["reg_cost"], 2), round(a["holiday_cost"], 2),
            round(a["ot_cost"], 2), round(a["dt_cost"], 2),
            round(a["total_cost"], 2),
            round(a["revenue"], 2), round(a["margin"], 2),
            _pct(a["margin"], a["revenue"]),
        ])

    totals = [SHEET_DAILY_TOTALS_COLS]
    for date, t in sorted(daily_agg.items(), reverse=True):
        totals.append([
            date, t["shifts"], len(t["projects"]), len(t["employees"]),
            round(t["reg_h"], 2), round(t["ot_h"], 2),
            round(t["dt_h"], 2), round(t["total_h"], 2),
            round(t["cost"], 2), round(t["revenue"], 2),
            round(t["margin"], 2), _pct(t["margin"], t["revenue"]),
        ])

    project_totals = [SHEET_PROJECT_TOTALS_COLS]
    for proj, a in sorted(proj_agg.items(), key=lambda x: x[1]["revenue"], reverse=True):
        project_totals.append([
            proj, len(a["dates"]), a["shifts"], len(a["employees"]),
            round(a["reg_h"], 2), round(a["ot_h"], 2),
            round(a["dt_h"], 2), round(a["total_h"], 2),
            round(a["cost"], 2), round(a["revenue"], 2),
            round(a["margin"], 2), _pct(a["margin"], a["revenue"]),
        ])

    # Per-project tabs (one per Job Description).
    by_desc: dict = defaultdict(list)
    for d in detail_rows:
        bucket = d.get("Job Description") or d.get("Job Number") or "(unlabeled)"
        by_desc[bucket].append(d)

    payload = {
        "Daily Detail": detail,
        "Project Rollup": rollup,
        "Daily Totals": totals,
        "Daily Totals by Project": project_totals,
    }
    for desc in sorted(by_desc.keys()):
        rows_sorted = sorted(by_desc[desc], key=lambda r: r["Date"], reverse=True)
        body = [DAILY_DETAIL_COLS] + [
            [r[c] for c in DAILY_DETAIL_COLS] for r in rows_sorted
        ]
        payload[_sanitize_tab_name(desc)] = body
    return payload


def _push_sheet(svc, sheet_id: str, payload: dict[str, list[list]]) -> int:
    """Write payload to a Google Sheet. Returns total cells written.

    Adds missing tabs, deletes stale ones, clears each before writing.
    Same logic as refresh_staffwizard_master.py.
    """
    meta = svc.spreadsheets().get(spreadsheetId=sheet_id).execute()
    existing = {s["properties"]["title"]: s["properties"]["sheetId"]
                for s in meta.get("sheets", [])}
    requests = []
    want_tabs = list(payload.keys())

    if "Sheet1" in existing and want_tabs and want_tabs[0] not in existing:
        requests.append({
            "updateSheetProperties": {
                "properties": {
                    "sheetId": existing["Sheet1"], "title": want_tabs[0],
                },
                "fields": "title",
            }
        })
        existing[want_tabs[0]] = existing.pop("Sheet1")
    for t in want_tabs:
        if t not in existing:
            requests.append({"addSheet": {"properties": {"title": t}}})
    if requests:
        svc.spreadsheets().batchUpdate(
            spreadsheetId=sheet_id, body={"requests": requests},
        ).execute()
        meta = svc.spreadsheets().get(spreadsheetId=sheet_id).execute()
        existing = {s["properties"]["title"]: s["properties"]["sheetId"]
                    for s in meta.get("sheets", [])}

    stale = [name for name in existing if name not in want_tabs]
    if stale and len(want_tabs) >= 1:
        svc.spreadsheets().batchUpdate(
            spreadsheetId=sheet_id,
            body={"requests": [{"deleteSheet": {"sheetId": existing[name]}}
                              for name in stale]},
        ).execute()

    # Delegate the bulk write to gservices.sheets_batch_write so we
    # share the quota-aware path with any other module that needs it
    # (Finnn 2026-05-03 — moved from inlined batchClear/batchUpdate).
    import gservices
    return gservices.sheets_batch_write(sheet_id, payload)


def push_master_to_sheets(
    *,
    src_dir: Optional[pathlib.Path] = None,
    master_id: Optional[str] = None,
    archive_id: Optional[str] = None,
    win_days: Optional[int] = None,
) -> dict:
    """Refresh both the live Master Sheet and the Historic Archive Sheet.

    Splits parsed rows by `today - win_days`. Recent rows go to Master,
    older rows go to Archive.

    Returns: {status, current_rows, archive_rows, master_cells, archive_cells, ...}
    """
    import auth  # noqa: F401
    from googleapiclient.discovery import build

    src_dir = src_dir or reports_dir()
    master_id = master_id or master_sheet_id()
    archive_id = archive_id or archive_sheet_id()
    win_days = win_days if win_days is not None else window_days()

    if not master_id or not archive_id:
        raise StaffWizardError(
            "config.staffwizard.master_sheet_id and archive_sheet_id "
            "must both be set."
        )

    today = _dt.date.today()
    cutoff = today - _dt.timedelta(days=win_days)

    detail_rows = _parse_all_reports(src_dir)
    current = []
    archive = []
    for d in detail_rows:
        try:
            row_date = _dt.date.fromisoformat(d["Date"])
        except (ValueError, TypeError, KeyError):
            continue
        (current if row_date >= cutoff else archive).append(d)

    creds = auth.get_credentials()
    svc = build("sheets", "v4", credentials=creds, cache_discovery=False)

    master_cells = _push_sheet(svc, master_id, _build_sheet_payload(current))
    archive_cells = _push_sheet(svc, archive_id, _build_sheet_payload(archive))

    return {
        "status": "ok",
        "window_days": win_days,
        "cutoff": cutoff.isoformat(),
        "current_rows": len(current),
        "archive_rows": len(archive),
        "master_cells": master_cells,
        "archive_cells": archive_cells,
        "master_url": f"https://docs.google.com/spreadsheets/d/{master_id}/edit",
        "archive_url": f"https://docs.google.com/spreadsheets/d/{archive_id}/edit",
    }


# =============================================================================
# Step 4 — build_dashboards
# =============================================================================

# We delegate the heavy HTML/JSON emission to the existing build script's
# logic. To keep this module the single entry point, we re-implement a
# thin wrapper that runs the same code — importing the script is brittle
# because it has its own argparse-driven main(). Instead we run it as a
# subprocess of the venv interpreter.

def build_dashboards(
    *,
    xlsx_path: Optional[pathlib.Path] = None,
    win_days: Optional[int] = None,
    upload: bool = True,
) -> dict:
    """Run scripts/build_project_dashboards.py against the master xlsx.

    Returns: {status, dashboards_dir, drive_folder_url?, projects, flagged}

    Subprocess-execs the existing script so we don't duplicate the
    ~800 LOC of HTML/CSS/JS in this module. Once you commit fully to
    the MCP path you can refactor to import the rendering functions
    directly.
    """
    import subprocess
    import sys

    project_root = pathlib.Path(__file__).resolve().parent
    script = project_root / "scripts" / "build_project_dashboards.py"
    xlsx_path = xlsx_path or master_xlsx_path()
    win_days = win_days if win_days is not None else window_days()

    if not xlsx_path.exists():
        raise MasterMissingError(
            f"Master xlsx not found at {xlsx_path}. "
            "Run build_master_xlsx() first."
        )

    cmd = [
        sys.executable, str(script),
        "--xlsx", str(xlsx_path),
        "--window-days", str(win_days),
    ]
    if not upload:
        cmd.append("--no-upload")

    proc = subprocess.run(
        cmd, capture_output=True, text=True, cwd=str(project_root),
    )
    if proc.returncode != 0:
        raise StaffWizardError(
            f"build_project_dashboards.py exited {proc.returncode}: "
            f"{proc.stderr or proc.stdout}"
        )

    # Parse the script's last few lines for the Drive folder URL.
    drive_url = None
    for line in proc.stdout.splitlines():
        if line.strip().startswith("Drive folder:"):
            drive_url = line.split(":", 1)[1].strip()
            break

    # Read data.json to get final counts.
    data_json = dashboards_dir() / "data.json"
    projects = flagged = 0
    if data_json.exists():
        try:
            data = json.loads(data_json.read_text())
            projects = data.get("totals", {}).get("projects", 0)
            flagged = sum(1 for p in data.get("projects", []) if p.get("flags"))
        except (json.JSONDecodeError, OSError):
            pass

    return {
        "status": "ok",
        "dashboards_dir": str(dashboards_dir()),
        "drive_folder_url": drive_url,
        "projects": projects,
        "flagged": flagged,
        "uploaded": upload,
        "stdout_tail": "\n".join(proc.stdout.splitlines()[-10:]),
    }


# =============================================================================
# Step 5 — send_dashboards_email
# =============================================================================

def _bundle_dashboards_zip(
    *,
    src_dir: Optional[pathlib.Path] = None,
    out_dir: Optional[pathlib.Path] = None,
) -> pathlib.Path:
    """Zip the dashboards/ folder with Chart.js bundled for offline use.

    If chart.umd.min.js isn't already in src_dir, fetch it once. Rewrite
    project HTMLs to use the local path.
    """
    import urllib.request
    src_dir = src_dir or dashboards_dir()
    out_dir = out_dir or pathlib.Path("/tmp")
    if not src_dir.exists():
        raise MasterMissingError(
            f"Dashboards dir not found at {src_dir}. Run build_dashboards() first."
        )

    chart_path = src_dir / "chart.umd.min.js"
    if not chart_path.exists():
        url = "https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"
        with urllib.request.urlopen(url, timeout=30) as r:
            chart_path.write_bytes(r.read())

    # Rewrite CDN URL → local path in project pages.
    cdn = "https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"
    for f in (src_dir / "projects").glob("*.html"):
        s = f.read_text()
        if cdn in s:
            f.write_text(s.replace(cdn, "../chart.umd.min.js"))
    # Also rewrite the index page's chart import.
    idx = src_dir / "index.html"
    if idx.exists():
        s = idx.read_text()
        if cdn in s:
            idx.write_text(s.replace(cdn, "chart.umd.min.js"))

    ts = _dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    out_path = out_dir / f"surefox-dashboards-{ts}.zip"
    out_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED, compresslevel=9) as zf:
        for p in src_dir.rglob("*"):
            if p.suffix in (".bak", ".tmp"):
                continue
            if p.is_file():
                zf.write(p, p.relative_to(src_dir.parent))
    return out_path


def send_dashboards_email(
    *,
    recipients: list[str],
    subject: Optional[str] = None,
    body: Optional[str] = None,
) -> dict:
    """Package the dashboards into a self-contained zip and email it.

    Returns: {status, zip_path, message_ids}
    """
    import auth  # noqa: F401
    import gservices

    if not recipients:
        raise StaffWizardError("recipients list is empty")

    zip_path = _bundle_dashboards_zip()

    gmail = gservices.gmail_service()

    subject = subject or "Surefox Daily Operations dashboards — open locally"
    body = body or (
        "Attached: zipped Surefox Daily Operations dashboards. Self-contained\n"
        "(Chart.js bundled). Save the zip, extract, double-click index.html.\n"
        "\n"
        "Tabs: Summary · All Projects Ranked by Revenue · Action Items.\n"
        "Click a project name in the ranked table to drill into its detail page.\n"
        "\n"
        "— Josh\n"
    )

    # Use the existing tools/gmail.py send path so attachments are
    # handled the same way the rest of the MCP does it.
    import tools.gmail as _gmail_tools  # noqa: F401

    # Build the MIME message manually so we don't reach into a tool's
    # internal Pydantic-wrapped function.
    import base64
    from email.message import EmailMessage
    from email.utils import formatdate

    msg_ids = []
    for to_addr in recipients:
        msg = EmailMessage()
        msg["To"] = to_addr
        msg["Subject"] = subject
        msg["Date"] = formatdate(localtime=True)
        msg.set_content(body)
        with zip_path.open("rb") as fp:
            msg.add_attachment(
                fp.read(),
                maintype="application", subtype="zip",
                filename=zip_path.name,
            )
        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode("ascii")
        sent = gmail.users().messages().send(
            userId="me", body={"raw": raw},
        ).execute()
        msg_ids.append(sent.get("id"))

    return {
        "status": "ok",
        "zip_path": str(zip_path),
        "size_bytes": zip_path.stat().st_size,
        "recipients": recipients,
        "message_ids": msg_ids,
    }


# =============================================================================
# Step 6 — refresh_all (orchestrator)
# =============================================================================

def refresh_all(
    *,
    fetch_latest: bool = True,
    work_date: Optional[_dt.date] = None,
    skip_dashboards: bool = False,
    upload_dashboards: bool = True,
) -> dict:
    """Run the whole pipeline: ingest → build master → push sheets → dashboards.

    Returns: {status, steps: {ingest, build_master, push_sheets, dashboards}}
    """
    out: dict[str, Any] = {"status": "ok", "steps": {}}

    if fetch_latest:
        out["steps"]["ingest"] = ingest_latest_report(work_date=work_date)

    out["steps"]["build_master"] = build_master_xlsx()
    out["steps"]["push_sheets"] = push_master_to_sheets()

    if not skip_dashboards:
        out["steps"]["dashboards"] = build_dashboards(upload=upload_dashboards)

    return out
