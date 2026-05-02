# © 2026 CoAssisted Workspace. Licensed under MIT.
"""Renderer + send_invoice coverage for ar_send.py.

Complement to tests/test_ar_collections_gate.py — that file covers
the Part F three-mode gate. This one covers the legacy renderers
(invoice HTML/text/xlsx, reminder HTML/text by tier) and the
send_invoice public entry point.

Together they bring ar_send.py from "only gate logic tested" to
broad behavioral coverage.
"""

from __future__ import annotations

import datetime as _dt
import io
import sys
import types

import pytest
from openpyxl import load_workbook

import ar_invoicing
import ar_send
import labor_ingest
import project_registry


# -----------------------------------------------------------------------------
# Fixtures
# -----------------------------------------------------------------------------

@pytest.fixture
def fresh_stores(tmp_path, monkeypatch):
    """Each test gets a fresh ar_invoices + project registry."""
    monkeypatch.setattr(ar_invoicing, "_INVOICES_PATH", tmp_path / "ar.json")
    monkeypatch.setattr(project_registry, "_REGISTRY_PATH", tmp_path / "p.json")
    yield tmp_path


def _seed_invoice(
    *,
    customer_email: str | None = "ap@google.com",
    customer_name: str = "Google, LLC",
    project_code: str = "GE1",
    paid_amount: float = 0.0,
) -> ar_invoicing.InvoiceRecord:
    """Drop a real invoice in the store + return the record.

    Mirrors the helper in test_ar_collections_gate.py but returns
    the record itself so renderer tests can introspect it.
    """
    project_registry.register(
        project_code,
        name=f"Test - {project_code}",
        client=customer_name,
        billing_terms="Net-15",
        customer_email=customer_email,
    )
    rows = [labor_ingest.LaborRow(
        job_number="", job_description="",
        work_date=_dt.date(2026, 4, 5),
        employee_name="Crystal Test", employee_number="E1",
        post_description="Day Shift",
        shift_start="", shift_end="",
        hours=8, overtime_hours=0, doubletime_hours=0,
        dollars=200, holiday_dollars=0, overtime_dollars=0,
        doubletime_dollars=0,
        billable_hours=8, billable_dollars=400,
    )]
    inv = ar_invoicing.generate_invoice_from_labor(
        project_code,
        period_start=_dt.date(2026, 4, 1),
        period_end=_dt.date(2026, 4, 30),
        labor_rows=rows,
        invoice_date=_dt.date(2026, 4, 1),
    )
    if paid_amount:
        inv.paid_amount = paid_amount
    ar_invoicing.persist(inv)
    return inv


# -----------------------------------------------------------------------------
# render_invoice_html
# -----------------------------------------------------------------------------

def test_render_invoice_html_contains_key_fields(fresh_stores):
    """HTML body contains invoice number, customer, project, total, due date."""
    inv = _seed_invoice()
    html = ar_send.render_invoice_html(inv)
    assert inv.invoice_number in html
    assert "Google, LLC" in html
    assert inv.project_code in html
    assert inv.due_date.isoformat() in html
    # Total appears formatted with thousands separator.
    assert f"${inv.total:,.2f}" in html


def test_render_invoice_html_lists_each_line_item(fresh_stores):
    """Every line item appears in the rendered table."""
    inv = _seed_invoice()
    html = ar_send.render_invoice_html(inv)
    for line in inv.lines:
        assert line.description in html
        # Quantity with 2-decimal formatting
        assert f"{line.quantity:.2f}" in html


def test_render_invoice_html_is_valid_html_skeleton(fresh_stores):
    """Body must contain html/body open + close tags."""
    inv = _seed_invoice()
    html = ar_send.render_invoice_html(inv)
    assert html.startswith("<html>") or "<html" in html[:50]
    assert "</html>" in html
    assert "<body" in html
    assert "</body>" in html


# -----------------------------------------------------------------------------
# render_invoice_text
# -----------------------------------------------------------------------------

def test_render_invoice_text_has_no_html_tags(fresh_stores):
    """Plain-text fallback should not leak HTML."""
    inv = _seed_invoice()
    text = ar_send.render_invoice_text(inv)
    assert "<" not in text
    assert ">" not in text or text.count(">") <= 1  # → arrow is allowed
    # The arrow character is U+2192, not <>.
    assert "<html" not in text
    assert "</body>" not in text


def test_render_invoice_text_contains_required_fields(fresh_stores):
    """Text body has invoice number, customer, total, due date."""
    inv = _seed_invoice()
    text = ar_send.render_invoice_text(inv)
    assert inv.invoice_number in text
    assert "Google, LLC" in text
    assert f"${inv.total:,.2f}" in text
    assert inv.due_date.isoformat() in text


# -----------------------------------------------------------------------------
# render_invoice_xlsx
# -----------------------------------------------------------------------------

def test_render_invoice_xlsx_returns_valid_workbook(fresh_stores):
    """Bytes parse as a valid xlsx and contain invoice metadata."""
    inv = _seed_invoice()
    xlsx_bytes = ar_send.render_invoice_xlsx(inv)
    assert isinstance(xlsx_bytes, bytes)
    assert len(xlsx_bytes) > 0
    # Round-trip parse — proves it's a real xlsx not just a string.
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    # Title cell carries the invoice number.
    assert inv.invoice_number in str(ws.cell(row=1, column=1).value)
    # Customer line.
    assert "Google, LLC" in str(ws.cell(row=2, column=1).value)


def test_render_invoice_xlsx_writes_each_line(fresh_stores):
    """Each line item lands in the xlsx body."""
    inv = _seed_invoice()
    xlsx_bytes = ar_send.render_invoice_xlsx(inv)
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    # Header row at row 7, lines start row 8.
    descs_in_sheet = []
    for row in ws.iter_rows(min_row=8, max_col=1, values_only=True):
        if row[0]:
            descs_in_sheet.append(row[0])
    for ln in inv.lines:
        assert ln.description in descs_in_sheet


# -----------------------------------------------------------------------------
# render_reminder_html
# -----------------------------------------------------------------------------

@pytest.mark.parametrize("tier", [
    "courtesy_reminder",
    "first_followup",
    "second_followup",
    "third_followup",
    "escalation_to_legal",
])
def test_render_reminder_returns_subject_html_text(fresh_stores, tier):
    """Every tier produces a (subject, html, text) tuple."""
    inv = _seed_invoice()
    subject, html, text = ar_send.render_reminder_html(
        inv, tier, as_of=_dt.date(2026, 5, 1),
    )
    assert subject  # truthy
    assert inv.invoice_number in subject
    assert html.startswith("<html>") or "<html" in html[:50]
    assert "</html>" in html
    assert "<" not in text  # text is text, not html
    # All bodies reference the invoice number.
    assert inv.invoice_number in html
    assert inv.invoice_number in text


def test_render_reminder_unknown_tier_raises(fresh_stores):
    """An unrecognized tier name raises ValueError."""
    inv = _seed_invoice()
    with pytest.raises(ValueError):
        ar_send.render_reminder_html(
            inv, "made_up_tier", as_of=_dt.date(2026, 5, 1),
        )


def test_render_reminder_days_past_due_clamped_at_zero(fresh_stores):
    """If as_of is BEFORE due_date, days_past_due renders as 0."""
    inv = _seed_invoice()
    # Render against a date well before the due date.
    early = inv.due_date - _dt.timedelta(days=10)
    _, html, text = ar_send.render_reminder_html(
        inv, "courtesy_reminder", as_of=early,
    )
    # The template uses {days_past_due} in body — for courtesy_reminder
    # the tone string doesn't include days_past_due, but the table
    # row in html_body and text_body shows it explicitly.
    assert "Days past due" in html
    # Both bodies have the line; should show 0 not negative.
    assert "-" not in text.split("Days past due")[1].split("\n")[0]


def test_render_reminder_outstanding_uses_paid_amount(fresh_stores):
    """Outstanding = total - paid_amount, not just total."""
    inv = _seed_invoice(paid_amount=100.0)
    outstanding = round(inv.total - inv.paid_amount, 2)
    _, html, _ = ar_send.render_reminder_html(
        inv, "courtesy_reminder", as_of=_dt.date(2026, 5, 1),
    )
    assert f"${outstanding:,.2f}" in html
    # Sanity — the full total should NOT appear as the outstanding figure
    # if there's been a partial payment.
    if outstanding != inv.total:
        # Total may still appear elsewhere, but the prominent
        # "Outstanding" row uses the discounted figure.
        assert f"${inv.total:,.2f}" not in html.split("Outstanding")[1].split("</tr>")[0]


# -----------------------------------------------------------------------------
# send_invoice
# -----------------------------------------------------------------------------

def test_send_invoice_unknown_id_returns_error(fresh_stores):
    """Sending a non-existent invoice ID surfaces an error."""
    result = ar_send.send_invoice("does-not-exist")
    assert result["sent"] is False
    assert "not found" in result.get("error", "").lower()


def test_send_invoice_no_recipient_returns_error(fresh_stores):
    """Invoice with no customer_email and no override → error."""
    inv = _seed_invoice(customer_email=None)
    result = ar_send.send_invoice(inv.invoice_id)
    assert result["sent"] is False
    assert "customer_email" in result.get("error", "")


def test_send_invoice_happy_path_marks_sent(fresh_stores, monkeypatch):
    """Successful send fires _send_email_with_attachment + mark_sent."""
    inv = _seed_invoice()
    sent_calls = []
    monkeypatch.setattr(
        ar_send, "_send_email_with_attachment",
        lambda **kw: sent_calls.append(kw) or {"sent": True, "message_id": "m-99"},
    )
    result = ar_send.send_invoice(inv.invoice_id)
    assert result["sent"] is True
    assert result["recipient"] == "ap@google.com"
    assert result["message_id"] == "m-99"
    assert len(sent_calls) == 1
    assert sent_calls[0]["to"] == "ap@google.com"
    assert inv.invoice_number in sent_calls[0]["subject"]
    # State advanced.
    refreshed = ar_invoicing.get(inv.invoice_id)
    assert refreshed.status == "sent"


def test_send_invoice_override_to_wins(fresh_stores, monkeypatch):
    """override_to beats the invoice's customer_email."""
    inv = _seed_invoice(customer_email="default@example.com")
    sent_calls = []
    monkeypatch.setattr(
        ar_send, "_send_email_with_attachment",
        lambda **kw: sent_calls.append(kw) or {"sent": True, "message_id": "m-99"},
    )
    ar_send.send_invoice(inv.invoice_id, override_to="finance@google.com")
    assert sent_calls[0]["to"] == "finance@google.com"


def test_send_invoice_attach_xlsx_false_skips_attachment(fresh_stores, monkeypatch):
    """attach_xlsx=False sends html/text only, no attachment bytes."""
    inv = _seed_invoice()
    sent_calls = []
    monkeypatch.setattr(
        ar_send, "_send_email_with_attachment",
        lambda **kw: sent_calls.append(kw) or {"sent": True, "message_id": "m-99"},
    )
    ar_send.send_invoice(inv.invoice_id, attach_xlsx=False)
    assert sent_calls[0]["attachment_bytes"] is None
    assert sent_calls[0]["attachment_name"] is None


def test_send_invoice_failed_send_does_not_mark_sent(fresh_stores, monkeypatch):
    """If the Gmail send fails, invoice state stays as draft."""
    inv = _seed_invoice()
    monkeypatch.setattr(
        ar_send, "_send_email_with_attachment",
        lambda **kw: {"sent": False, "error": "smtp blew up"},
    )
    result = ar_send.send_invoice(inv.invoice_id)
    assert result["sent"] is False
    assert result.get("error") == "smtp blew up"
    refreshed = ar_invoicing.get(inv.invoice_id)
    assert refreshed.status == "draft"


def test_send_invoice_xlsx_render_failure_falls_back_to_no_attachment(
    fresh_stores, monkeypatch,
):
    """If render_invoice_xlsx blows up, we still send the email (best-effort)."""
    inv = _seed_invoice()

    def _exploding_xlsx(invoice):
        raise RuntimeError("openpyxl dropped its keys")

    monkeypatch.setattr(ar_send, "render_invoice_xlsx", _exploding_xlsx)
    sent_calls = []
    monkeypatch.setattr(
        ar_send, "_send_email_with_attachment",
        lambda **kw: sent_calls.append(kw) or {"sent": True, "message_id": "m-99"},
    )
    result = ar_send.send_invoice(inv.invoice_id, attach_xlsx=True)
    assert result["sent"] is True
    # No attachment shipped.
    assert sent_calls[0]["attachment_bytes"] is None
