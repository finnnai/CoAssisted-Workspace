# © 2026 CoAssisted Workspace. Licensed under MIT.
"""Unit tests for labor_ingest.py — AP-7 StaffWizard ingestion.

Most coverage uses synthetic LaborRow objects; the actual file-parsing
end-to-end test runs only when the gitignored samples/ Overall Report
is present (skipped in CI).
"""

from __future__ import annotations

import datetime as _dt
from pathlib import Path

import pytest

import labor_ingest
import project_registry


# -----------------------------------------------------------------------------
# Filename heuristic
# -----------------------------------------------------------------------------

def test_looks_like_overall_report_matches_real_filename():
    assert labor_ingest.looks_like_overall_report(
        "Overall Report SFOX 1777532406.xls"
    )
    assert labor_ingest.looks_like_overall_report("overall report.xlsx")
    assert labor_ingest.looks_like_overall_report("OVERALL REPORT.xls")


def test_looks_like_overall_report_rejects_other_files():
    assert not labor_ingest.looks_like_overall_report("something.xls")
    assert not labor_ingest.looks_like_overall_report("daily_report.pdf")
    assert not labor_ingest.looks_like_overall_report("invoice.xlsx")


# -----------------------------------------------------------------------------
# Coercion helpers
# -----------------------------------------------------------------------------

def test_coerce_float_handles_currency_strings():
    assert labor_ingest._coerce_float("$1,234.56") == 1234.56
    assert labor_ingest._coerce_float("0") == 0.0
    assert labor_ingest._coerce_float("") == 0.0
    assert labor_ingest._coerce_float(None) == 0.0


def test_coerce_float_passes_through_numbers():
    assert labor_ingest._coerce_float(42) == 42.0
    assert labor_ingest._coerce_float(3.14) == 3.14


def test_coerce_date_handles_multiple_formats():
    assert labor_ingest._coerce_date("2026-04-29") == _dt.date(2026, 4, 29)
    assert labor_ingest._coerce_date("4/29/26") == _dt.date(2026, 4, 29)
    assert labor_ingest._coerce_date("4/29/2026") == _dt.date(2026, 4, 29)


def test_coerce_date_handles_datetime_input():
    dt = _dt.datetime(2026, 4, 29, 14, 30)
    assert labor_ingest._coerce_date(dt) == _dt.date(2026, 4, 29)


def test_coerce_date_returns_none_on_garbage():
    assert labor_ingest._coerce_date("not a date") is None
    assert labor_ingest._coerce_date(None) is None


# -----------------------------------------------------------------------------
# LaborRow + ProjectLabor math
# -----------------------------------------------------------------------------

def _make_row(
    *,
    job_number="Google, LLC",
    job_desc="Golden Eagle 1",
    hours=8.0,
    overtime_hours=0.0,
    doubletime_hours=0.0,
    dollars=200.0,
    overtime_dollars=0.0,
    holiday_dollars=0.0,
    doubletime_dollars=0.0,
    billable_hours=8.0,
    billable_dollars=300.0,
    work_date=None,
):
    return labor_ingest.LaborRow(
        job_number=job_number,
        job_description=job_desc,
        work_date=work_date or _dt.date(2026, 4, 29),
        employee_name="Test, Employee",
        employee_number="100001",
        post_description="Main Post",
        shift_start="00:00",
        shift_end="08:00",
        hours=hours,
        overtime_hours=overtime_hours,
        doubletime_hours=doubletime_hours,
        dollars=dollars,
        holiday_dollars=holiday_dollars,
        overtime_dollars=overtime_dollars,
        doubletime_dollars=doubletime_dollars,
        billable_hours=billable_hours,
        billable_dollars=billable_dollars,
    )


def test_labor_row_total_cost_sums_all_pay_buckets():
    row = _make_row(dollars=200, holiday_dollars=50, overtime_dollars=75, doubletime_dollars=20)
    assert row.total_cost == 345.0


def test_labor_row_total_hours_sums_reg_ot_dbl():
    row = _make_row(hours=8, overtime_hours=2, doubletime_hours=0.5)
    assert row.total_hours == 10.5


def test_labor_row_margin_is_revenue_minus_cost():
    row = _make_row(dollars=200, billable_dollars=350)
    assert row.margin == 150.0


def test_project_labor_aggregates_rows():
    pl = labor_ingest.ProjectLabor(
        project_code="GE1",
        job_number="Google, LLC",
        job_description="Golden Eagle 1",
        work_date=_dt.date(2026, 4, 29),
        rows=[
            _make_row(dollars=100, billable_dollars=200),
            _make_row(dollars=150, billable_dollars=250),
            _make_row(dollars=200, billable_dollars=400, overtime_dollars=50),
        ],
    )
    assert pl.shift_count == 3
    assert pl.total_cost == 500.0  # 100 + 150 + 200 + 50 OT
    assert pl.total_revenue == 850.0
    assert pl.margin == 350.0


# -----------------------------------------------------------------------------
# group_by_project
# -----------------------------------------------------------------------------

@pytest.fixture
def fresh_registry(tmp_path, monkeypatch):
    fake = tmp_path / "projects.json"
    monkeypatch.setattr(project_registry, "_REGISTRY_PATH", fake)
    yield fake


def test_group_routes_known_jobs_to_their_project(fresh_registry):
    project_registry.register(
        "GE1",
        name="Google - Golden Eagle 1",
        staffwizard_job_number="Google, LLC",
        staffwizard_job_desc="Golden Eagle 1",
    )
    parsed = labor_ingest.ParsedReport(rows=[
        _make_row(job_number="Google, LLC", job_desc="Golden Eagle 1"),
        _make_row(job_number="Google, LLC", job_desc="Golden Eagle 1", dollars=100),
    ])
    grouped = labor_ingest.group_by_project(parsed)
    assert len(grouped) == 1
    pl = grouped["GE1"]
    assert pl.project_code == "GE1"
    assert pl.shift_count == 2


def test_group_buckets_unmapped_jobs_separately(fresh_registry):
    """Jobs without a registry entry land under unmapped::* keys."""
    project_registry.register(
        "GE1",
        name="Google - Golden Eagle 1",
        staffwizard_job_number="Google, LLC",
        staffwizard_job_desc="Golden Eagle 1",
    )
    parsed = labor_ingest.ParsedReport(rows=[
        _make_row(job_number="Google, LLC", job_desc="Golden Eagle 1"),
        _make_row(job_number="Mystery", job_desc="Codename X"),
    ])
    grouped = labor_ingest.group_by_project(parsed)
    assert "GE1" in grouped
    unmapped_keys = [k for k in grouped.keys() if k.startswith("unmapped::")]
    assert len(unmapped_keys) == 1
    assert grouped[unmapped_keys[0]].project_code is None


def test_group_separates_two_jobs_at_same_client(fresh_registry):
    """Same client, different sub-job → two registry records → two groups."""
    project_registry.register(
        "GE1",
        name="Google - Golden Eagle 1",
        staffwizard_job_number="Google, LLC",
        staffwizard_job_desc="Golden Eagle 1",
    )
    project_registry.register(
        "GE2",
        name="Google - Golden Eagle 2",
        staffwizard_job_number="Google, LLC",
        staffwizard_job_desc="Golden Eagle 2",
    )
    parsed = labor_ingest.ParsedReport(rows=[
        _make_row(job_number="Google, LLC", job_desc="Golden Eagle 1"),
        _make_row(job_number="Google, LLC", job_desc="Golden Eagle 2"),
    ])
    grouped = labor_ingest.group_by_project(parsed)
    assert {"GE1", "GE2"} <= set(grouped.keys())


# -----------------------------------------------------------------------------
# Workbook writer
# -----------------------------------------------------------------------------

def test_write_project_labor_workbook_round_trip(tmp_path):
    pl = labor_ingest.ProjectLabor(
        project_code="GE1",
        job_number="Google, LLC",
        job_description="Golden Eagle 1",
        work_date=_dt.date(2026, 4, 29),
        rows=[
            _make_row(dollars=200, billable_dollars=300),
            _make_row(dollars=300, billable_dollars=400),
        ],
    )
    out = tmp_path / "labor.xlsx"
    labor_ingest.write_project_labor_workbook(pl, out)
    assert out.exists()

    from openpyxl import load_workbook
    wb = load_workbook(out, data_only=True)
    ws = wb["Daily Labor"]
    assert "GE1" in ws.cell(row=1, column=1).value
    # Two data rows + summary footer separated by gap
    # Row 3 + 4 = data, row 6 = summary (one blank gap)
    summary = ws.cell(row=6, column=1).value
    assert summary == "TOTAL"
    # Total cost in column 13
    assert float(ws.cell(row=6, column=13).value) == 500.0
    # Total revenue in column 15
    assert float(ws.cell(row=6, column=15).value) == 700.0


# -----------------------------------------------------------------------------
# End-to-end against samples/ — runs only when sample is present
# -----------------------------------------------------------------------------

_SAMPLE = Path(__file__).resolve().parent.parent / "samples" / (
    "Overall Report SFOX 1777532406.xls"
)


@pytest.mark.skipif(
    not _SAMPLE.exists(),
    reason="StaffWizard sample missing (samples/ is gitignored)",
)
def test_parse_real_overall_report():
    parsed = labor_ingest.parse_overall_report(_SAMPLE)
    assert len(parsed.rows) > 0
    assert parsed.work_date is not None
    # The April 29 sample has 108 valid shifts.
    assert len(parsed.rows) >= 100
