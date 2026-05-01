# © 2026 CoAssisted Workspace. Licensed under MIT.
"""Unit tests for master_rollup.py — AP-8 baseline math + workbook build."""

from __future__ import annotations

import datetime as _dt

import pytest

import master_rollup
import project_registry


@pytest.fixture
def fresh_stores(tmp_path, monkeypatch):
    """Each test gets clean history + registry stores."""
    monkeypatch.setattr(
        master_rollup, "_HISTORY_PATH", tmp_path / "h.json"
    )
    monkeypatch.setattr(
        project_registry, "_REGISTRY_PATH", tmp_path / "p.json"
    )
    yield tmp_path


# -----------------------------------------------------------------------------
# compute_baseline
# -----------------------------------------------------------------------------

def test_compute_baseline_empty():
    """Zero observations = trivial trustless baseline."""
    bl = master_rollup.compute_baseline([])
    assert bl.days_observed == 0
    assert bl.in_cold_start is True
    assert bl.is_trustworthy is False


def test_compute_baseline_cold_start():
    """Below cold-start threshold (30 days) the baseline isn't trusted."""
    bl = master_rollup.compute_baseline([100.0] * 10)
    assert bl.in_cold_start is True
    assert bl.is_trustworthy is False


def test_compute_baseline_established():
    """Above cold-start, baseline is trustworthy."""
    bl = master_rollup.compute_baseline([100.0] * 31)
    assert bl.is_trustworthy is True
    assert bl.mean_daily_spend == 100.0
    assert bl.stdev_daily_spend == 0.0


def test_compute_baseline_2sigma_envelope():
    """Mean ± 2σ band is computed from input."""
    # Mean=100, stdev≈10
    daily = [90, 100, 110, 90, 100, 110] * 6  # 36 obs
    bl = master_rollup.compute_baseline(daily)
    assert abs(bl.mean_daily_spend - 100) < 1
    # 2σ band roughly mean ± 20.
    assert bl.upper_2sigma > bl.mean_daily_spend
    assert bl.lower_2sigma < bl.mean_daily_spend


def test_baseline_lower_2sigma_clamps_to_zero():
    """Lower 2σ never goes negative — spend can't be < 0."""
    daily = [10.0] * 31
    # Force a high stdev artificially.
    daily += [200.0] * 5  # mostly 10, occasional 200 → stdev > mean
    bl = master_rollup.compute_baseline(daily)
    assert bl.lower_2sigma >= 0.0


# -----------------------------------------------------------------------------
# detect_deviation
# -----------------------------------------------------------------------------

def test_detect_deviation_returns_none_when_in_band():
    bl = master_rollup.compute_baseline([100.0] * 31)
    # No stdev → no band — observed equals mean.
    assert master_rollup.detect_deviation(100.0, bl) is None


def test_detect_deviation_spike():
    daily = [100, 90, 110, 100, 95, 105] * 6  # mean ≈ 100, stdev ≈ 7
    bl = master_rollup.compute_baseline(daily)
    alert = master_rollup.detect_deviation(500.0, bl)
    assert alert is not None
    assert alert["direction"] == "spike"
    assert alert["deviation_pct"] > 0


def test_detect_deviation_drop():
    daily = [100, 90, 110, 100, 95, 105] * 6
    bl = master_rollup.compute_baseline(daily)
    # Sudden zero → drop alert
    alert = master_rollup.detect_deviation(0.0, bl)
    assert alert is not None
    assert alert["direction"] == "drop"
    assert alert["deviation_pct"] < 0


def test_detect_deviation_skips_cold_start():
    """Even huge spikes shouldn't fire alerts during cold start."""
    bl = master_rollup.compute_baseline([100.0] * 5)
    assert master_rollup.detect_deviation(100_000.0, bl) is None


# -----------------------------------------------------------------------------
# run_rate_window
# -----------------------------------------------------------------------------

def test_run_rate_returns_none_without_enough_data():
    assert master_rollup.run_rate_window([100, 200], days=7) is None


def test_run_rate_averages_last_n_days():
    daily = list(range(1, 31))  # 1..30
    rr7 = master_rollup.run_rate_window(daily, days=7)
    # Last 7 = 24..30, mean = 27
    assert rr7 == 27.0


# -----------------------------------------------------------------------------
# DailyFact + record / read round trip
# -----------------------------------------------------------------------------

def test_record_and_get_history(fresh_stores):
    fact = master_rollup.DailyFact(
        project_code="GE1",
        work_date=_dt.date(2026, 5, 1),
        receipts=100,
        invoices=50,
        labor_cost=200,
        labor_revenue=400,
    )
    master_rollup.record_daily_fact(fact)
    rows = master_rollup.get_daily_history("GE1")
    assert len(rows) == 1
    assert rows[0][0] == _dt.date(2026, 5, 1)
    assert rows[0][1]["receipts"] == 100


def test_record_overwrites_same_date(fresh_stores):
    """Re-recording the same (project, date) replaces — used for corrections."""
    master_rollup.record_daily_fact(master_rollup.DailyFact(
        project_code="GE1",
        work_date=_dt.date(2026, 5, 1),
        receipts=100,
    ))
    master_rollup.record_daily_fact(master_rollup.DailyFact(
        project_code="GE1",
        work_date=_dt.date(2026, 5, 1),
        receipts=999,
    ))
    rows = master_rollup.get_daily_history("GE1")
    assert len(rows) == 1
    assert rows[0][1]["receipts"] == 999


def test_total_cost_history_sums_components(fresh_stores):
    master_rollup.record_daily_fact(master_rollup.DailyFact(
        project_code="GE1",
        work_date=_dt.date(2026, 5, 1),
        receipts=100, invoices=50, labor_cost=200, labor_revenue=999,
    ))
    history = master_rollup.total_cost_history("GE1")
    # Note: revenue is NOT in total_cost
    assert history == [350.0]


def test_daily_fact_total_cost_and_margin():
    fact = master_rollup.DailyFact(
        project_code="GE1",
        work_date=_dt.date(2026, 5, 1),
        receipts=100, invoices=50, labor_cost=200, labor_revenue=400,
    )
    assert fact.total_cost == 350.0
    assert fact.margin == 50.0  # 400 - 350


# -----------------------------------------------------------------------------
# build_master_workbook end-to-end
# -----------------------------------------------------------------------------

def test_build_master_workbook_three_tabs(fresh_stores, tmp_path):
    project_registry.register("GE1", name="Google - Golden Eagle 1")
    master_rollup.record_daily_fact(master_rollup.DailyFact(
        project_code="GE1",
        work_date=_dt.date(2026, 5, 1),
        receipts=100, invoices=50, labor_cost=200, labor_revenue=400,
    ))
    out = tmp_path / "rollup.xlsx"
    master_rollup.build_master_workbook(
        output_path=out,
        target_date=_dt.date(2026, 5, 1),
    )
    from openpyxl import load_workbook
    wb = load_workbook(out, data_only=True)
    assert wb.sheetnames == ["All Projects", "PM Dashboard", "Anomalies"]


def test_build_master_workbook_anomaly_appears_in_third_tab(
    fresh_stores, tmp_path
):
    """Spike day after cold start should land in the Anomalies tab."""
    project_registry.register("GE1", name="Google - Golden Eagle 1")
    # 30 days at ~$1000.
    for i in range(30):
        master_rollup.record_daily_fact(master_rollup.DailyFact(
            project_code="GE1",
            work_date=_dt.date(2026, 4, 1) + _dt.timedelta(days=i),
            receipts=300, invoices=200, labor_cost=500, labor_revenue=1400,
        ))
    # Spike day: 5x normal.
    master_rollup.record_daily_fact(master_rollup.DailyFact(
        project_code="GE1",
        work_date=_dt.date(2026, 5, 1),
        receipts=2000, invoices=1500, labor_cost=1500, labor_revenue=400,
    ))
    out = tmp_path / "rollup.xlsx"
    master_rollup.build_master_workbook(
        output_path=out, target_date=_dt.date(2026, 5, 1),
    )
    from openpyxl import load_workbook
    wb = load_workbook(out, data_only=True)
    ws = wb["Anomalies"]
    # Header + at least one alert row.
    assert ws.max_row >= 2
    # Row 2: spike alert for GE1
    assert ws.cell(row=2, column=1).value == "GE1"
    assert ws.cell(row=2, column=3).value == "spike"


def test_build_master_workbook_pm_dashboard_one_row_per_project(
    fresh_stores, tmp_path
):
    """PM Dashboard has exactly one row per project regardless of date count."""
    project_registry.register("GE1", name="Google - Golden Eagle 1")
    project_registry.register("C12", name="Prometheus - Condor 12")
    for code in ("GE1", "C12"):
        for i in range(5):
            master_rollup.record_daily_fact(master_rollup.DailyFact(
                project_code=code,
                work_date=_dt.date(2026, 5, 1) + _dt.timedelta(days=i),
                receipts=100,
            ))
    out = tmp_path / "rollup.xlsx"
    master_rollup.build_master_workbook(
        output_path=out, target_date=_dt.date(2026, 5, 5),
    )
    from openpyxl import load_workbook
    wb = load_workbook(out, data_only=True)
    ws = wb["PM Dashboard"]
    # Header + 2 project rows.
    assert ws.max_row == 3
    project_codes = {ws.cell(row=r, column=1).value for r in (2, 3)}
    assert project_codes == {"GE1", "C12"}
