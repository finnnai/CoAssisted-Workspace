# © 2026 CoAssisted Workspace. Licensed under MIT.
"""Unit tests for workday_journal_eib.py — AP-2 end to end.

Synthetic CSV fixtures inline so tests work in CI without samples/.
"""

from __future__ import annotations

import datetime
import io

import pytest

import workday_journal_eib as wje


# =============================================================================
# AMEX parser
# =============================================================================

# Minimal AMEX CSV with 3 rows: 1 cleared, 1 pending, 1 declined.
# Only the columns the parser reads are populated; the rest are blank.
_AMEX_CSV = (
    "Registered Card Last 4,Funding Source,Type of Account,Transaction Date,"
    "Clear Date,Type,Status,Name On Card,Card User Email,"
    "Authorization Merchant Currency,Authorization Merchant Amount,"
    "Authorization Billing Currency,Authorization Exchange Rate,"
    "Authorization Billing Amount,Settlement Merchant Currency,"
    "Settlement Merchant Amount,Settlement Billing Currency,"
    "Settlement Billing Amount,Settlement Exchange Rate,Merchant Name,"
    "Merchant Address,Merchant City,Merchant State,Merchant Country,"
    "Merchant ZIP/Postal,MCC,MCC Description,VCN Sender,VCN Sender Email,"
    "VCN Nickname,VCN Last 4,Reference Id,Source,External,"
    "Number of Attachments,Attachments,Transaction Amount,Notes,"
    "Transaction Receipt,Receipt,Manager Review,Final Approval,Team\n"
    # row 1: cleared
    "*2375,Surefox,Physical Card,4/30/26,4/30/26,DEBIT,CLEARED,Joshua Szott,"
    "josh@surefox.com,USD,28.27,USD,1,28.27,USD,28.27,USD,28.27,1,"
    "PIRATE SHIP,,,,,,4215,Couriers,,,,,,,FALSE,0,,28.27,,,,,,\n"
    # row 2: pending — should be filtered out by default
    "*2375,Surefox,Physical Card,4/29/26,,DEBIT,PENDING,Joshua Szott,"
    "josh@surefox.com,USD,76.04,USD,1,76.04,,0,,0,,AMAZONCOM,,,,,,5969,"
    "Direct Marketing,,,,,,,FALSE,0,,76.04,,,,,,\n"
    # row 3: declined — should be filtered out
    "*2375,Surefox,Physical Card,4/28/26,,DEBIT,DECLINED,Joshua Szott,"
    "josh@surefox.com,USD,99.99,USD,1,99.99,,0,,0,,SOMEPLACE,,,,,,9999,"
    "Unknown,,,,,,,FALSE,0,,99.99,,,,,,\n"
)


def _amex_path(tmp_path):
    p = tmp_path / "amex.csv"
    p.write_text(_AMEX_CSV)
    return p


def test_amex_parser_keeps_only_cleared(tmp_path):
    """Pending and declined rows are filtered out by default."""
    txns = wje.parse_amex_csv(_amex_path(tmp_path))
    assert len(txns) == 1
    assert txns[0].merchant_name == "PIRATE SHIP"
    assert txns[0].status == "CLEARED"


def test_amex_parser_include_pending(tmp_path):
    """include_pending=True picks up PENDING rows but still drops DECLINED."""
    txns = wje.parse_amex_csv(_amex_path(tmp_path), include_pending=True)
    # CLEARED + PENDING; DECLINED still excluded.
    assert len(txns) == 2
    assert {t.status for t in txns} == {"CLEARED", "PENDING"}


def test_amex_parser_normalizes_cardholder_name(tmp_path):
    """'Joshua Szott' (First Last) → 'Szott, Joshua' (Last, First)."""
    txns = wje.parse_amex_csv(_amex_path(tmp_path))
    assert txns[0].cardholder_name == "Szott, Joshua"


def test_amex_parser_strips_card_last4_asterisk(tmp_path):
    """'*2375' → '2375'."""
    txns = wje.parse_amex_csv(_amex_path(tmp_path))
    assert txns[0].card_last_4 == "2375"


def test_amex_parser_extracts_mcc(tmp_path):
    """MCC is captured as int + description string."""
    txns = wje.parse_amex_csv(_amex_path(tmp_path))
    assert txns[0].mcc_code == 4215
    assert txns[0].mcc_description == "Couriers"


def test_amex_parser_parses_date_correctly(tmp_path):
    """M/D/YY format common in AMEX exports."""
    txns = wje.parse_amex_csv(_amex_path(tmp_path))
    assert txns[0].transaction_date == datetime.date(2026, 4, 30)


# =============================================================================
# WEX parser
# =============================================================================

# Minimal WEX CSV — 2 rows for two different drivers/departments.
_WEX_CSV = (
    "Transaction Date,Transaction Time,Post Date,Account Number,Account Name,"
    "Card Number,Trans ID,Emboss Line 2,Custom Vehicle/Asset ID,Units,"
    "Unit of Measure,Unit Cost,Total Fuel Cost,Service Cost,Other Cost,"
    "Total Non-Fuel Cost,Gross Cost,Exempt Tax,Discount,Net Cost,Reported Tax,"
    "Transaction Fee Type 1,Transaction Fee Amount 1,Transaction Fee Type 2,"
    "Transaction Fee Amount 2,Transaction Fee Type 3,Transaction Fee Amount 3,"
    "Transaction Fee Type 4,Transaction Fee Amount 4,Transaction Fee Type 5,"
    "Transaction Fee Amount 5,Product,Product Description,Transaction Description,"
    "Merchant (Brand),Merchant Name,Merchant Address,Merchant City,"
    "Merchant State / Province,Merchant Postal Code,Merchant Site ID,"
    "Current Odometer,Adjusted Odometer,Previous Odometer,Distance Driven,"
    "Fuel Economy,Cost Per Distance,Vehicle Description,VIN,Tank Capacity,"
    "In Service Date,Start Odometer,Driver Last Name,Driver First Name,"
    "Driver Middle Name,Driver Department,Employee ID,Transaction Ticket Number,"
    "Currency Exchange Rate,Rebate Code,Department\n"
    "4/30/26,0:09:19,5/1/26,0460001739804,Surefox,****21915,1,,ISOC Gas,14.6,"
    "GA,6.49,95,0,0,0,95,0,0,95,0,,0,,0,,0,,0,,0,UNL,Unleaded Regular,OP,"
    "CHEVRON,Chevron 0090562,1265 Lawrence,Sunnyvale,CA,94089,0,18677,,18344,"
    "333,22.78,0.285,,,0,,0,GORDON,BOB,,ISOC,,16,,,ISOC\n"
    "4/29/26,16:55:35,5/1/26,0460001739804,Surefox,****82170,2,GREEN 24,"
    "G-24 / G1 Follow,12.9,GA,7.69,100,0,0,0,100,0,0,100,0,,0,,0,,0,,0,,0,"
    "UNL,Unleaded Regular,OP,CHEVRON,Chevron 0304564,3600 Alameda,Menlo Park,"
    "CA,94025,0,7004,,6892,112,8.62,0.893,G1 Follow,,28,,35506,Team,Hardware,"
    ",GREEN FLEET,,14,,,GREEN FLEET\n"
)


def _wex_path(tmp_path):
    p = tmp_path / "wex.csv"
    p.write_text(_WEX_CSV)
    return p


def test_wex_parser_extracts_driver_and_department(tmp_path):
    txns = wje.parse_wex_csv(_wex_path(tmp_path))
    assert len(txns) == 2
    by_driver = {t.driver_name: t for t in txns}
    assert "GORDON, BOB" in by_driver
    assert by_driver["GORDON, BOB"].department == "ISOC"
    assert by_driver["GORDON, BOB"].merchant_name == "CHEVRON"


def test_wex_parser_uses_net_cost_as_amount(tmp_path):
    """WEX 'Net Cost' column drives the journal amount."""
    txns = wje.parse_wex_csv(_wex_path(tmp_path))
    by_driver = {t.driver_name: t for t in txns}
    assert by_driver["GORDON, BOB"].amount == 95.0


def test_wex_parser_has_no_mcc(tmp_path):
    """WEX exports don't carry MCC — tier 1 routes by memo only."""
    txns = wje.parse_wex_csv(_wex_path(tmp_path))
    for t in txns:
        assert t.mcc_code is None


def test_wex_parser_builds_memo_with_signal(tmp_path):
    """raw_notes carries merchant + product + vehicle for tier-2 matcher."""
    txns = wje.parse_wex_csv(_wex_path(tmp_path))
    by_driver = {t.driver_name: t for t in txns}
    notes = by_driver["GORDON, BOB"].raw_notes
    assert "CHEVRON" in notes
    assert "UNL" in notes  # Product code


# =============================================================================
# EIB writer
# =============================================================================

@pytest.fixture
def synthetic_amex_txns():
    """Two synthetic AMEX transactions: one airline, one Amazon."""
    return [
        wje.CardTransaction(
            card_type="amex",
            card_last_4="2375",
            transaction_date=datetime.date(2026, 4, 15),
            cardholder_name="Szott, Joshua",
            cardholder_email="josh@surefox.com",
            amount=487.20,
            merchant_name="UNITED AIRLINES",
            mcc_code=4511,
            mcc_description="Airlines",
            status="CLEARED",
        ),
        wje.CardTransaction(
            card_type="amex",
            card_last_4="2375",
            transaction_date=datetime.date(2026, 4, 20),
            cardholder_name="Vetre, Michael",
            cardholder_email="mike@surefox.com",
            amount=18.65,
            merchant_name="AMAZON",
            mcc_code=5969,
            mcc_description="Direct Marketing",
            status="CLEARED",
        ),
    ]


def test_build_eib_writes_two_sheets(synthetic_amex_txns, tmp_path):
    """Workday journal EIBs are always two sheets: header + lines."""
    out = tmp_path / "eib.xlsx"
    result = wje.build_journal_eib(
        synthetic_amex_txns, out,
        card_type="amex",
        period_start=datetime.date(2026, 4, 1),
        period_end=datetime.date(2026, 4, 30),
    )
    assert result.output_path == out
    assert out.exists()

    from openpyxl import load_workbook
    wb = load_workbook(out, data_only=True)
    assert wb.sheetnames == [
        "Import Accounting Journal",
        "Journal Entry Line Replacement",
    ]


def test_build_eib_emits_two_lines_per_transaction(synthetic_amex_txns, tmp_path):
    """Every txn produces a debit/credit pair = 2 lines."""
    out = tmp_path / "eib.xlsx"
    result = wje.build_journal_eib(
        synthetic_amex_txns, out,
        card_type="amex",
        period_start=datetime.date(2026, 4, 1),
        period_end=datetime.date(2026, 4, 30),
    )
    assert result.n_transactions == 2
    assert result.n_lines_written == 4


def test_build_eib_credit_side_routes_to_card_payable(
    synthetic_amex_txns, tmp_path
):
    """The credit-side of every AMEX line posts to 22000:Credit Card Payable- AMEX."""
    out = tmp_path / "eib.xlsx"
    wje.build_journal_eib(
        synthetic_amex_txns, out,
        card_type="amex",
        period_start=datetime.date(2026, 4, 1),
        period_end=datetime.date(2026, 4, 30),
    )
    from openpyxl import load_workbook
    ws = load_workbook(out, data_only=True)["Journal Entry Line Replacement"]
    # Credit-side rows have a value in col 11 (Credit Amount).
    credit_accounts = []
    for r in range(6, ws.max_row + 1):
        if ws.cell(row=r, column=11).value:  # has Credit Amount
            credit_accounts.append(ws.cell(row=r, column=6).value)
    assert credit_accounts  # at least one credit row
    for acct in credit_accounts:
        assert acct == "22000:Credit Card Payable- AMEX"


def test_build_eib_wex_credit_routes_to_22030(tmp_path):
    """WEX uses a different card payable (22030)."""
    txns = [wje.CardTransaction(
        card_type="wex",
        card_last_4="21915",
        transaction_date=datetime.date(2026, 4, 30),
        cardholder_name="GORDON, BOB",
        amount=95.0,
        merchant_name="CHEVRON",
        department="ISOC",
        raw_notes="CHEVRON UNL Unleaded Regular ISOC Gas",
    )]
    out = tmp_path / "wex.xlsx"
    wje.build_journal_eib(
        txns, out,
        card_type="wex",
        period_start=datetime.date(2026, 4, 1),
        period_end=datetime.date(2026, 4, 30),
    )
    from openpyxl import load_workbook
    ws = load_workbook(out, data_only=True)["Journal Entry Line Replacement"]
    credit_account = None
    for r in range(6, ws.max_row + 1):
        if ws.cell(row=r, column=11).value:
            credit_account = ws.cell(row=r, column=6).value
            break
    assert credit_account == "22030:Credit Card Payable- WEX"


def test_build_eib_debit_credit_balance(synthetic_amex_txns, tmp_path):
    """Total debits == total credits (every journal must balance)."""
    out = tmp_path / "eib.xlsx"
    wje.build_journal_eib(
        synthetic_amex_txns, out,
        card_type="amex",
        period_start=datetime.date(2026, 4, 1),
        period_end=datetime.date(2026, 4, 30),
    )
    from openpyxl import load_workbook
    ws = load_workbook(out, data_only=True)["Journal Entry Line Replacement"]
    total_debit = sum(
        float(ws.cell(row=r, column=10).value or 0)
        for r in range(6, ws.max_row + 1)
    )
    total_credit = sum(
        float(ws.cell(row=r, column=11).value or 0)
        for r in range(6, ws.max_row + 1)
    )
    assert total_debit == total_credit
    assert total_debit == pytest.approx(487.20 + 18.65)


def test_build_eib_refund_reverses_dr_cr(tmp_path):
    """Negative-amount transactions debit the payable and credit the spend GL."""
    txns = [wje.CardTransaction(
        card_type="amex",
        card_last_4="2375",
        transaction_date=datetime.date(2026, 4, 15),
        cardholder_name="Szott, Joshua",
        cardholder_email="josh@surefox.com",
        amount=-100.0,  # refund
        merchant_name="UNITED AIRLINES",
        mcc_code=4511,
        status="CLEARED",
    )]
    out = tmp_path / "refund.xlsx"
    wje.build_journal_eib(
        txns, out,
        card_type="amex",
        period_start=datetime.date(2026, 4, 1),
        period_end=datetime.date(2026, 4, 30),
    )
    from openpyxl import load_workbook
    ws = load_workbook(out, data_only=True)["Journal Entry Line Replacement"]
    # First row: debit-side. For a refund, debit-side should be the payable.
    assert ws.cell(row=6, column=6).value == "22000:Credit Card Payable- AMEX"
    assert float(ws.cell(row=6, column=10).value) == 100.0
    # Second row: credit-side spend.
    assert ws.cell(row=7, column=6).value == "53000:Travel - COS"
    assert float(ws.cell(row=7, column=11).value) == 100.0


def test_build_eib_uses_cost_center_map(synthetic_amex_txns, tmp_path):
    """Cardholder email → cost center mapping populates the CC column."""
    out = tmp_path / "eib.xlsx"
    wje.build_journal_eib(
        synthetic_amex_txns, out,
        card_type="amex",
        period_start=datetime.date(2026, 4, 1),
        period_end=datetime.date(2026, 4, 30),
        cardholder_cost_center_map={
            "josh@surefox.com": "CC200",
            "mike@surefox.com": "CC300",
        },
    )
    from openpyxl import load_workbook
    ws = load_workbook(out, data_only=True)["Journal Entry Line Replacement"]
    # First two rows are Joshua's (CC200); next two are Michael's (CC300).
    assert ws.cell(row=6, column=23).value == "CC200"
    assert ws.cell(row=7, column=23).value == "CC200"
    assert ws.cell(row=8, column=23).value == "CC300"
    assert ws.cell(row=9, column=23).value == "CC300"


def test_build_eib_falls_back_to_default_cc_when_unmapped(
    synthetic_amex_txns, tmp_path
):
    """Cardholders not in the map get the default_cost_center."""
    out = tmp_path / "eib.xlsx"
    wje.build_journal_eib(
        synthetic_amex_txns, out,
        card_type="amex",
        period_start=datetime.date(2026, 4, 1),
        period_end=datetime.date(2026, 4, 30),
        cardholder_cost_center_map={},
        default_cost_center="CC999",
    )
    from openpyxl import load_workbook
    ws = load_workbook(out, data_only=True)["Journal Entry Line Replacement"]
    for r in range(6, ws.max_row + 1):
        assert ws.cell(row=r, column=23).value == "CC999"


def test_build_eib_header_key_format(synthetic_amex_txns, tmp_path):
    """Header Key follows the SFNA convention: DDMMYY + card-suffix."""
    out = tmp_path / "eib.xlsx"
    wje.build_journal_eib(
        synthetic_amex_txns, out,
        card_type="amex",
        period_start=datetime.date(2026, 4, 1),
        period_end=datetime.date(2026, 4, 30),
    )
    from openpyxl import load_workbook
    ws = load_workbook(out, data_only=True)["Import Accounting Journal"]
    assert ws.cell(row=6, column=2).value == "300426AU"


def test_build_eib_memo_format_matches_sample(synthetic_amex_txns, tmp_path):
    """Memo follows: '{LABEL} Transactions {start}-{end} - {Cardholder} - {Vendor}'."""
    out = tmp_path / "eib.xlsx"
    wje.build_journal_eib(
        synthetic_amex_txns, out,
        card_type="amex",
        period_start=datetime.date(2026, 3, 1),
        period_end=datetime.date(2026, 3, 31),
    )
    from openpyxl import load_workbook
    ws = load_workbook(out, data_only=True)["Journal Entry Line Replacement"]
    memo = ws.cell(row=6, column=20).value
    assert memo == (
        "AMEX Transactions 03.01.26-03.31.26 - "
        "Szott, Joshua - UNITED AIRLINES"
    )


def test_build_eib_unknown_card_type_raises(tmp_path):
    """Card types we don't have a payable account for fail loud."""
    with pytest.raises(ValueError, match="Unknown card_type"):
        wje.build_journal_eib(
            [], tmp_path / "x.xlsx",
            card_type="diners",  # not in the table
            period_start=datetime.date(2026, 4, 1),
            period_end=datetime.date(2026, 4, 30),
        )


def test_build_eib_classifier_tier_counts(synthetic_amex_txns, tmp_path):
    """Result reports which classifier tier handled each transaction."""
    out = tmp_path / "eib.xlsx"
    result = wje.build_journal_eib(
        synthetic_amex_txns, out,
        card_type="amex",
        period_start=datetime.date(2026, 4, 1),
        period_end=datetime.date(2026, 4, 30),
    )
    # Both synthetic txns have MCC matches (4511 + 5969) → tier 1.
    assert result.classifier_tier_counts.get("mcc_table", 0) == 2


# =============================================================================
# Number parsing edge cases
# =============================================================================

def test_parse_float_handles_accounting_negative():
    assert wje._parse_float("(123.45)") == -123.45


def test_parse_float_strips_currency_chars():
    assert wje._parse_float("$1,234.56") == 1234.56


def test_parse_float_blank_returns_zero():
    assert wje._parse_float("") == 0.0
    assert wje._parse_float(None) == 0.0


def test_parse_float_passes_through_numeric():
    assert wje._parse_float(42) == 42.0
    assert wje._parse_float(3.14) == 3.14
