# © 2026 CoAssisted Workspace. Licensed under MIT.
"""AR-9 send wire-up — actually deliver invoices and collection reminders.

`ar_invoicing` carries the deterministic state (record, generate, age,
schedule). This module is the side-effect surface: render an invoice
into an email-ready HTML body + xlsx attachment, dispatch via Gmail,
and log the post-send status back into ar_invoices.json.

Two public entry points:

    send_invoice(invoice_id, ...)
        Renders the invoice as an HTML body (line items, totals, due
        date, terms) plus an Excel attachment, sends from the operator
        to the customer email, and on success calls
        ar_invoicing.mark_sent so the AR-9 lifecycle advances.

    send_collection_reminder(invoice_id, ...)
        Picks the right tier from the cadence ladder (or accepts an
        explicit override), renders the body using a tier-appropriate
        template, sends, and records a collection_event so the next
        scheduled run won't re-send the same tier.

Both call into the existing tools/gmail Gmail send-email path via the
underlying Google API, mirroring how ap_sweep.py wired up its four
stubs. Best-effort: API failures land in the returned status, no
exceptions propagate.
"""

from __future__ import annotations

import datetime as _dt
import io
from pathlib import Path
from typing import Optional

import ar_invoicing


# =============================================================================
# Tier copy — what to say in each collection reminder
# =============================================================================

_TIER_TEMPLATES: dict[str, dict[str, str]] = {
    "courtesy_reminder": {
        "subject": "Friendly reminder: invoice {invoice_number}",
        "tone": (
            "Just a courtesy heads-up — invoice {invoice_number} for "
            "${outstanding:,.2f} was due on {due_date}. If it's already "
            "in your AP queue, no action needed. If not, please let us "
            "know and we'll make sure you have what you need."
        ),
    },
    "first_followup": {
        "subject": "Past due: invoice {invoice_number}",
        "tone": (
            "Following up on invoice {invoice_number} for "
            "${outstanding:,.2f}, which was due on {due_date} ("
            "{days_past_due} days ago). Could you confirm where this "
            "stands in your AP cycle?"
        ),
    },
    "second_followup": {
        "subject": "30+ days past due: invoice {invoice_number}",
        "tone": (
            "Invoice {invoice_number} for ${outstanding:,.2f} is now "
            "{days_past_due} days past due (originally due {due_date}). "
            "We need to escalate internally if we don't hear back this "
            "week. Please send a status update or proposed pay date."
        ),
    },
    "third_followup": {
        "subject": "URGENT: invoice {invoice_number} — escalation pending",
        "tone": (
            "Invoice {invoice_number} for ${outstanding:,.2f} is now "
            "{days_past_due} days past due. Per our terms, we're "
            "preparing to escalate to senior collections. Please respond "
            "within 5 business days with a payment commitment to avoid "
            "this step."
        ),
    },
    "escalation_to_legal": {
        "subject": "Final notice: invoice {invoice_number}",
        "tone": (
            "This is the final notice on invoice {invoice_number} for "
            "${outstanding:,.2f}, now {days_past_due} days past due. "
            "Without a payment commitment in the next 5 business days, "
            "this matter will move to outside collections per the "
            "service agreement. Please contact us immediately."
        ),
    },
}


# =============================================================================
# Internal: Gmail + Drive plumbing (mirrors ap_sweep.py wire-up)
# =============================================================================

def _build_gmail_service():
    """Return an authenticated Gmail service, or None on any failure."""
    try:
        from googleapiclient.discovery import build
        from auth import get_credentials
        return build("gmail", "v1", credentials=get_credentials(),
                     cache_discovery=False)
    except Exception:
        return None


def _send_email_with_attachment(
    *,
    to: str,
    subject: str,
    html_body: str,
    text_body: str,
    attachment_bytes: Optional[bytes] = None,
    attachment_name: Optional[str] = None,
    attachment_mime: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
) -> dict:
    """Send a multipart message via Gmail. Returns {sent: bool, message_id?, error?}."""
    service = _build_gmail_service()
    if not service:
        return {"sent": False, "error": "Gmail service unavailable"}

    import base64
    from email.mime.application import MIMEApplication
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    msg = MIMEMultipart("alternative")
    msg["To"] = to
    msg["Subject"] = subject

    msg.attach(MIMEText(text_body, "plain"))
    msg.attach(MIMEText(html_body, "html"))

    if attachment_bytes and attachment_name:
        part = MIMEApplication(attachment_bytes, Name=attachment_name)
        part["Content-Disposition"] = f'attachment; filename="{attachment_name}"'
        msg.attach(part)

    try:
        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
        sent = service.users().messages().send(
            userId="me", body={"raw": raw},
        ).execute()
        return {"sent": True, "message_id": sent.get("id")}
    except Exception as e:
        return {"sent": False, "error": str(e)}


# =============================================================================
# Renderers
# =============================================================================

def render_invoice_html(invoice: ar_invoicing.InvoiceRecord) -> str:
    """Customer-facing HTML body for the invoice email."""
    lines_html = "".join(
        f"<tr>"
        f"<td style='padding:6px 12px;border-bottom:1px solid #eee'>{ln.description}</td>"
        f"<td style='padding:6px 12px;border-bottom:1px solid #eee;text-align:right'>{ln.quantity:.2f}</td>"
        f"<td style='padding:6px 12px;border-bottom:1px solid #eee;text-align:right'>${ln.rate:,.2f}</td>"
        f"<td style='padding:6px 12px;border-bottom:1px solid #eee;text-align:right'>${ln.amount:,.2f}</td>"
        f"</tr>"
        for ln in invoice.lines
    )
    return f"""\
<html>
<body style="font-family:Helvetica,Arial,sans-serif;color:#222;max-width:680px;margin:24px auto">
<h2 style="margin:0 0 4px 0">Invoice {invoice.invoice_number}</h2>
<p style="margin:0 0 24px 0;color:#666">
  Issued {invoice.invoice_date.isoformat()} · Due {invoice.due_date.isoformat()} · Terms: {invoice.terms}
</p>

<p style="margin:0 0 6px 0"><strong>Bill to:</strong> {invoice.customer_name}</p>
<p style="margin:0 0 24px 0;color:#666">
  Service period: {invoice.period_start.isoformat()} → {invoice.period_end.isoformat()}<br>
  Project: {invoice.project_code}
</p>

<table style="border-collapse:collapse;width:100%;margin-bottom:24px">
  <thead>
    <tr style="background:#f5f5f5">
      <th style="padding:8px 12px;text-align:left">Description</th>
      <th style="padding:8px 12px;text-align:right">Hours</th>
      <th style="padding:8px 12px;text-align:right">Rate</th>
      <th style="padding:8px 12px;text-align:right">Amount</th>
    </tr>
  </thead>
  <tbody>
    {lines_html}
    <tr>
      <td colspan="3" style="padding:12px;text-align:right;font-weight:bold">Total</td>
      <td style="padding:12px;text-align:right;font-weight:bold">${invoice.total:,.2f}</td>
    </tr>
  </tbody>
</table>

<p>Please remit by <strong>{invoice.due_date.isoformat()}</strong>.<br>
Reply to this email with any questions.</p>

<p style="color:#888;font-size:12px;margin-top:32px">
  Surefox North America Inc. · Generated automatically by CoAssisted Workspace.
</p>
</body>
</html>
"""


def render_invoice_text(invoice: ar_invoicing.InvoiceRecord) -> str:
    """Plain-text fallback for the invoice email."""
    lines = "\n".join(
        f"  {ln.description:30s}  {ln.quantity:6.2f}h  @ ${ln.rate:7.2f}  =  ${ln.amount:9.2f}"
        for ln in invoice.lines
    )
    return f"""\
Invoice {invoice.invoice_number}
Issued {invoice.invoice_date.isoformat()} · Due {invoice.due_date.isoformat()} · Terms: {invoice.terms}

Bill to: {invoice.customer_name}
Service period: {invoice.period_start.isoformat()} → {invoice.period_end.isoformat()}
Project: {invoice.project_code}

{lines}

Total: ${invoice.total:,.2f}

Please remit by {invoice.due_date.isoformat()}.

Surefox North America Inc. · Generated by CoAssisted Workspace.
"""


def render_invoice_xlsx(invoice: ar_invoicing.InvoiceRecord) -> bytes:
    """Return the invoice as Excel bytes. Same shape as the email body."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws.cell(row=1, column=1, value=f"Invoice {invoice.invoice_number}")
    ws.cell(row=2, column=1, value=f"Bill to: {invoice.customer_name}")
    ws.cell(row=3, column=1, value=f"Project: {invoice.project_code}")
    ws.cell(row=4, column=1, value=f"Period: {invoice.period_start.isoformat()} → {invoice.period_end.isoformat()}")
    ws.cell(row=5, column=1, value=f"Issued: {invoice.invoice_date.isoformat()} · Due: {invoice.due_date.isoformat()} · Terms: {invoice.terms}")

    # Header row at row 7
    headers = ["Description", "Hours", "Rate", "Amount"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=7, column=c, value=h)
    r = 8
    for ln in invoice.lines:
        ws.cell(row=r, column=1, value=ln.description)
        ws.cell(row=r, column=2, value=ln.quantity)
        ws.cell(row=r, column=3, value=ln.rate)
        ws.cell(row=r, column=4, value=ln.amount)
        r += 1
    ws.cell(row=r + 1, column=3, value="Total")
    ws.cell(row=r + 1, column=4, value=invoice.total)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_reminder_html(
    invoice: ar_invoicing.InvoiceRecord,
    tier: str,
    *,
    as_of: _dt.date,
) -> tuple[str, str, str]:
    """Return (subject, html_body, text_body) for a collection reminder."""
    template = _TIER_TEMPLATES.get(tier)
    if not template:
        raise ValueError(f"Unknown reminder tier: {tier!r}")
    days_past_due = (as_of - invoice.due_date).days
    outstanding = round(invoice.total - invoice.paid_amount, 2)

    fields = {
        "invoice_number": invoice.invoice_number,
        "outstanding": outstanding,
        "due_date": invoice.due_date.isoformat(),
        "days_past_due": max(0, days_past_due),
    }
    subject = template["subject"].format(**fields)
    body_para = template["tone"].format(**fields)

    text_body = f"""\
Hi,

{body_para}

Invoice details:
  Invoice number: {invoice.invoice_number}
  Outstanding:    ${outstanding:,.2f}
  Original due:   {invoice.due_date.isoformat()}
  Days past due:  {max(0, days_past_due)}
  Project:        {invoice.project_code}

If payment has been processed in the last few days, please disregard.
Reply with any questions.

Surefox North America Inc.
"""

    html_body = f"""\
<html>
<body style="font-family:Helvetica,Arial,sans-serif;color:#222;max-width:680px;margin:24px auto">
<p>Hi,</p>
<p>{body_para}</p>

<table style="border-collapse:collapse;margin:16px 0">
  <tr><td style="padding:4px 12px;color:#666">Invoice number</td>
      <td style="padding:4px 12px"><strong>{invoice.invoice_number}</strong></td></tr>
  <tr><td style="padding:4px 12px;color:#666">Outstanding</td>
      <td style="padding:4px 12px"><strong>${outstanding:,.2f}</strong></td></tr>
  <tr><td style="padding:4px 12px;color:#666">Original due</td>
      <td style="padding:4px 12px">{invoice.due_date.isoformat()}</td></tr>
  <tr><td style="padding:4px 12px;color:#666">Days past due</td>
      <td style="padding:4px 12px">{max(0, days_past_due)}</td></tr>
  <tr><td style="padding:4px 12px;color:#666">Project</td>
      <td style="padding:4px 12px">{invoice.project_code}</td></tr>
</table>

<p style="color:#888;font-size:12px">If payment has been processed in
the last few days, please disregard. Reply with any questions.</p>

<p style="color:#888;font-size:12px;margin-top:32px">
  Surefox North America Inc.
</p>
</body>
</html>
"""
    return subject, html_body, text_body


# =============================================================================
# Public entry points
# =============================================================================

def send_invoice(
    invoice_id: str,
    *,
    attach_xlsx: bool = True,
    override_to: Optional[str] = None,
) -> dict:
    """Send an invoice email to the customer.

    On success, calls `ar_invoicing.mark_sent` to advance state.
    Returns {sent, invoice_id, recipient, message_id?, error?}.
    """
    invoice = ar_invoicing.get(invoice_id)
    if not invoice:
        return {"sent": False, "error": f"Invoice {invoice_id!r} not found."}
    recipient = override_to or invoice.customer_email
    if not recipient:
        return {
            "sent": False,
            "error": (
                f"Invoice {invoice.invoice_number} has no customer_email "
                "and no override_to was provided."
            ),
        }

    html = render_invoice_html(invoice)
    text = render_invoice_text(invoice)
    subject = f"Invoice {invoice.invoice_number} from Surefox"

    attachment_bytes = None
    attachment_name = None
    if attach_xlsx:
        try:
            attachment_bytes = render_invoice_xlsx(invoice)
            attachment_name = f"{invoice.invoice_number}.xlsx"
        except Exception:
            attachment_bytes = None  # Best-effort; HTML body is sufficient.

    result = _send_email_with_attachment(
        to=recipient,
        subject=subject,
        html_body=html,
        text_body=text,
        attachment_bytes=attachment_bytes,
        attachment_name=attachment_name,
    )
    if result.get("sent"):
        ar_invoicing.mark_sent(invoice_id)
    return {
        "sent": result.get("sent", False),
        "invoice_id": invoice_id,
        "recipient": recipient,
        "message_id": result.get("message_id"),
        "error": result.get("error"),
    }


# =============================================================================
# Collections-mode resolution — Finnn 2026-05-01 Part F
# =============================================================================

# The three modes — kept as constants so consumers don't sprinkle
# string literals across the codebase.
COLLECTIONS_MODE_SEND = "send"
COLLECTIONS_MODE_DRAFT = "draft"
COLLECTIONS_MODE_DISABLED = "disabled"

_COLLECTIONS_MODES = frozenset(
    [COLLECTIONS_MODE_SEND, COLLECTIONS_MODE_DRAFT, COLLECTIONS_MODE_DISABLED]
)


def resolve_collections_mode(tier: str) -> str:
    """Return the configured mode for a given collection tier.

    Lookup priority:
        1. ``config.ar.collections_mode_per_tier[tier]`` if set
        2. ``config.ar.collections_mode`` (the base default)
        3. Hardcoded fallback: ``"draft"`` (safe-by-default)

    Unknown mode strings fall back to ``"draft"`` to fail safe.
    """
    try:
        import config as _config_mod
        ar_block = _config_mod.get("ar", {}) or {}
    except Exception:
        return COLLECTIONS_MODE_DRAFT

    if not isinstance(ar_block, dict):
        return COLLECTIONS_MODE_DRAFT

    per_tier = ar_block.get("collections_mode_per_tier") or {}
    if isinstance(per_tier, dict) and tier in per_tier:
        candidate = per_tier[tier]
        if candidate in _COLLECTIONS_MODES:
            return candidate

    base = ar_block.get("collections_mode")
    if base in _COLLECTIONS_MODES:
        return base

    return COLLECTIONS_MODE_DRAFT


# =============================================================================
# Post-approval hook — fires when an AR collection draft is approved + sent
# =============================================================================

def _on_ar_collection_approved(rec: dict) -> None:
    """Hook registered with draft_queue. Advances AR state on send.

    Reads ``invoice_id`` and ``tier`` from the draft's ``meta`` dict —
    we put those there in :func:`send_collection_reminder` when
    enqueuing in draft mode.
    """
    meta = (rec or {}).get("meta") or {}
    invoice_id = meta.get("invoice_id")
    tier = meta.get("tier")
    if not invoice_id or not tier:
        return
    note = (
        f"Approved + sent via workflow_approve_draft "
        f"(draft_queue id={rec.get('id')})"
    )
    ar_invoicing.add_collection_event(invoice_id, tier, note=note)


# Register the hook at module import. Idempotent — register_post_approval_hook
# dedups so re-imports during tests don't double-fire.
try:
    import draft_queue as _draft_queue
    _draft_queue.register_post_approval_hook(
        "ar_collection", _on_ar_collection_approved,
    )
except Exception:
    pass


# =============================================================================
# Public — send_collection_reminder (mode-aware)
# =============================================================================

def send_collection_reminder(
    invoice_id: str,
    *,
    tier: Optional[str] = None,
    as_of: Optional[_dt.date] = None,
    override_to: Optional[str] = None,
    mode_override: Optional[str] = None,
) -> dict:
    """Send (or draft, or skip) a collection reminder per the configured mode.

    Per Finnn 2026-05-01 Part F + Joshua's question-3 answer
    (every tier defaults to "draft", escalation_to_legal defaults to
    "disabled"), the v0.8.3+ behavior consults ``config.ar.collections_mode``:

      - ``"disabled"`` — return early with status="skipped".
      - ``"draft"`` (default) — create a Gmail draft + queue in
        workflow_list_drafts. Operator approves via
        workflow_approve_draft. Post-approval hook advances AR state.
      - ``"send"`` — send immediately (the legacy v0.8.1/v0.8.2 path).
        Used only when the operator has explicitly opted in via config.

    Args:
        invoice_id: target invoice in ar_invoices.json
        tier: reminder tier name; defaults to whatever the cadence
            ladder says is due as of `as_of`.
        as_of: cadence anchor date.
        override_to: send to this address instead of customer_email.
        mode_override: bypass the config mode for this one call.
            Useful for tests + admin one-shots. Validated against the
            three known modes; unknown values fall back to "draft".
    """
    invoice = ar_invoicing.get(invoice_id)
    if not invoice:
        return {"sent": False, "error": f"Invoice {invoice_id!r} not found."}
    recipient = override_to or invoice.customer_email
    if not recipient:
        return {
            "sent": False,
            "error": (
                f"Invoice {invoice.invoice_number} has no customer_email "
                "and no override_to was provided."
            ),
        }

    as_of = as_of or _dt.date.today()

    # Pick tier — default to whichever the cadence ladder says is due now.
    if not tier:
        candidates = ar_invoicing.collections_due_today(as_of=as_of)
        match = next(
            (c for c in candidates if c.invoice.invoice_id == invoice_id),
            None,
        )
        if not match:
            return {
                "sent": False,
                "error": (
                    f"No collection reminder is due for invoice "
                    f"{invoice.invoice_number} as of {as_of.isoformat()}. "
                    f"Pass tier explicitly to override the cadence."
                ),
            }
        tier = match.reminder_type

    # Resolve mode: explicit override > config lookup.
    if mode_override and mode_override in _COLLECTIONS_MODES:
        mode = mode_override
    else:
        mode = resolve_collections_mode(tier)

    # MODE: disabled — return early, no draft, no send. due_today still
    # surfaces this invoice via ar_invoicing.collections_due_today.
    if mode == COLLECTIONS_MODE_DISABLED:
        return {
            "sent": False,
            "drafted": False,
            "status": "skipped",
            "reason": "collections_disabled",
            "invoice_id": invoice_id,
            "tier": tier,
            "mode": mode,
        }

    try:
        subject, html, text = render_reminder_html(invoice, tier, as_of=as_of)
    except ValueError as e:
        return {"sent": False, "error": str(e)}

    # MODE: draft — enqueue via draft_queue. Operator approves via
    # workflow_approve_draft (the hook fires add_collection_event on
    # successful send).
    if mode == COLLECTIONS_MODE_DRAFT:
        try:
            import draft_queue
            draft_id = draft_queue.enqueue(
                kind="ar_collection",
                subject=subject,
                body_plain=text,
                body_html=html,
                target=recipient,
                source_ref=f"invoice:{invoice.invoice_number}",
                meta={
                    "invoice_id": invoice_id,
                    "invoice_number": invoice.invoice_number,
                    "tier": tier,
                    "customer_name": invoice.customer_name,
                    "outstanding": round(
                        invoice.total - invoice.paid_amount, 2
                    ),
                },
            )
            return {
                "sent": False,
                "drafted": True,
                "status": "drafted",
                "draft_id": draft_id,
                "invoice_id": invoice_id,
                "tier": tier,
                "recipient": recipient,
                "mode": mode,
            }
        except Exception as e:
            return {
                "sent": False,
                "drafted": False,
                "status": "error",
                "error": f"draft enqueue failed: {e}",
                "invoice_id": invoice_id,
                "tier": tier,
                "mode": mode,
            }

    # MODE: send — legacy immediate-send path. Hook does NOT fire here
    # because we're not going through draft_queue; we call
    # add_collection_event inline on success.
    result = _send_email_with_attachment(
        to=recipient, subject=subject, html_body=html, text_body=text,
    )
    if result.get("sent"):
        ar_invoicing.add_collection_event(
            invoice_id,
            tier,
            note=(
                f"Sent immediately (mode=send) to {recipient}; "
                f"gmail_message_id={result.get('message_id')}"
            ),
        )
    return {
        "sent": result.get("sent", False),
        "drafted": False,
        "status": "sent" if result.get("sent") else "error",
        "invoice_id": invoice_id,
        "recipient": recipient,
        "tier": tier,
        "mode": mode,
        "message_id": result.get("message_id"),
        "error": result.get("error"),
    }
