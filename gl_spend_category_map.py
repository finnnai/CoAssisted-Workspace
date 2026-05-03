# © 2026 CoAssisted Workspace. Licensed under MIT.
# See LICENSE file for terms.
"""GL Account → Workday Spend Category map (auto-derived from JE history).

AP-1 needs every Supplier Invoice line to carry a Spend Category in addition
to the Ledger Account that the GL classifier already produces. Workday's
Supplier Invoice EIB validates Spend Category as a required worktag whenever
the Ledger Account is in the expense range (50000–69999).

Rather than ask the operator to hand-curate ~86 mappings (one per expense
account), we derive the map from `Wolfhound Corp JEs Jan-Mar'26.xlsx` —
the same 17,346-row training base the GL classifier uses. Each posted JE
line carries both the Ledger Account and the Spend Category worktag, so
the mapping is observable.

Algorithm
---------
For every (Ledger Account, Spend Category) pair seen in the JE history:
    1. Count occurrences.
    2. Group by Ledger Account.
    3. Pick the Spend Category with the highest count as the "dominant"
       category for that account.
    4. Compute dominance = top_count / total_count_for_account.

If dominance ≥ 0.80 → mark the mapping CONFIRMED (auto-applies in EIB).
If dominance <  0.80 → mark the mapping AMBIGUOUS (needs operator confirm
                       before EIB lines that hit this account can ship).

Operator overrides land in `gl_spend_category_overrides.json` (gitignored)
and always win, regardless of dominance.

Public surface
--------------
    derive_from_je_workbook(path) -> dict
        One-shot: parse the JE workbook and return the full map. Persists
        to `gl_spend_category_map.json` next to this file.

    lookup(gl_account, *, allow_ambiguous=False) -> SpendCategoryResult
        Runtime lookup the EIB writer calls per line.

    set_override(gl_account, spend_category, note=None) -> None
        Operator confirm / override. Writes to overrides JSON.

    list_ambiguous() -> list[dict]
        Surface every account with dominance < 0.80 plus its candidates,
        for the operator-confirm prompt.

    stats() -> dict
        {confirmed, ambiguous, overrides, total_je_rows_seen, last_built}.
"""

from __future__ import annotations

import datetime as _dt
import json
import os
import tempfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Optional


_PROJECT_ROOT = Path(__file__).resolve().parent
_MAP_PATH = _PROJECT_ROOT / "gl_spend_category_map.json"
_OVERRIDES_PATH = _PROJECT_ROOT / "gl_spend_category_overrides.json"

# Confidence floor below which we won't auto-apply the mapping in an EIB.
DOMINANCE_FLOOR = 0.80

# Workday expense-range bounds. Spend Category is only required for accounts
# in this band; balance-sheet GLs (assets/liabilities/equity) don't need one.
EXPENSE_RANGE = (50000, 69999)


@dataclass
class SpendCategoryResult:
    spend_category: Optional[str]   # None → no mapping; caller must park line
    status: str                     # 'confirmed' | 'ambiguous' | 'override' |
                                    # 'unmapped' | 'not_required'
    dominance: float                # 0.0–1.0; 1.0 for overrides
    sample_count: int               # how many JE rows backed the dominant pick
    candidates: list[tuple[str, int]]  # all (spend_category, count) pairs

    def as_dict(self) -> dict:
        return {
            "spend_category": self.spend_category,
            "status": self.status,
            "dominance": round(self.dominance, 3),
            "sample_count": self.sample_count,
            "candidates": [
                {"spend_category": sc, "count": n} for sc, n in self.candidates
            ],
        }


# --------------------------------------------------------------------------- #
# Storage primitives — atomic-write parity with project_registry.py
# --------------------------------------------------------------------------- #


def _now_iso() -> str:
    return _dt.datetime.now().astimezone().isoformat(timespec="seconds")


def _atomic_write_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(
        prefix=path.stem + ".", suffix=".json.tmp", dir=str(path.parent),
    )
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False, sort_keys=True)
        os.replace(tmp_path, path)
    except (OSError, TypeError, ValueError):
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


def _load_map() -> dict:
    if not _MAP_PATH.exists():
        return {"map": {}, "meta": {}}
    try:
        return json.loads(_MAP_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {"map": {}, "meta": {}}


def _load_overrides() -> dict:
    if not _OVERRIDES_PATH.exists():
        return {}
    try:
        return json.loads(_OVERRIDES_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #


def _normalize_account(account: str) -> str:
    """Workday GLs come in two shapes in the JE export:
        '53000' (5-digit code only)
        '53000:Travel-COS' (code + name)
    We key everything on the 5-digit code.
    """
    if not account:
        return ""
    a = str(account).strip()
    head = a.split(":", 1)[0].strip()
    return head


def _is_expense(account_code: str) -> bool:
    try:
        n = int(account_code)
    except (TypeError, ValueError):
        return False
    lo, hi = EXPENSE_RANGE
    return lo <= n <= hi


def _normalize_category(category: str) -> str:
    if not category:
        return ""
    c = str(category).strip()
    # Workday Spend Category worktags export as either bare name
    # ("Spend Category: Travel") or with the SC- prefix. Normalize to
    # the bare display name for readability.
    for prefix in ("Spend Category:", "SC-", "SC:"):
        if c.startswith(prefix):
            c = c[len(prefix):].strip()
    return c


# --------------------------------------------------------------------------- #
# Derivation from JE workbook
# --------------------------------------------------------------------------- #


def derive_from_je_workbook(path: str | Path) -> dict:
    """Parse the JE training workbook and (re)build the GL→Spend Category map.

    Looks for two columns by header name (case-insensitive contains-match):
        - Ledger Account / GL Account / Account
        - Spend Category / Worktag - Spend Category
    Header row is the first row whose cells contain both kinds of header
    matches.

    Returns a summary dict:
        {confirmed, ambiguous, total_rows, expense_rows, last_built,
         dominance_floor, output_path}
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    ws = wb.active

    # Auto-detect columns by header. Walk the first 5 rows looking for the
    # first row that has both a "ledger account" column and a "spend category"
    # column.
    ledger_col = None
    spend_col = None
    header_row = None
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        for cidx, val in enumerate(cells):
            if ledger_col is None and ("ledger account" in val or val == "gl account" or val == "account"):
                ledger_col = cidx
            if spend_col is None and "spend category" in val:
                spend_col = cidx
        if ledger_col is not None and spend_col is not None:
            header_row = ridx
            break

    if ledger_col is None or spend_col is None:
        raise ValueError(
            f"Could not locate Ledger Account and Spend Category columns "
            f"in {path}. Looked at first 5 rows."
        )

    counts: dict[str, Counter[str]] = defaultdict(Counter)
    total_rows = 0
    expense_rows = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row is None:
            continue
        total_rows += 1
        account = _normalize_account(row[ledger_col] if ledger_col < len(row) else "")
        category = _normalize_category(row[spend_col] if spend_col < len(row) else "")
        if not account or not _is_expense(account):
            continue
        expense_rows += 1
        if not category:
            # Track "blank" so we can see how often Spend Category is missing.
            counts[account]["__BLANK__"] += 1
            continue
        counts[account][category] += 1

    wb.close()

    map_out: dict[str, dict] = {}
    confirmed = 0
    ambiguous = 0
    for account, cat_counts in counts.items():
        # Strip blanks for dominance math — blanks aren't a candidate, they
        # just signal "this account often ships without a worktag in JEs."
        non_blank = Counter({k: v for k, v in cat_counts.items() if k != "__BLANK__"})
        if not non_blank:
            map_out[account] = {
                "spend_category": None,
                "status": "ambiguous",
                "dominance": 0.0,
                "sample_count": 0,
                "blank_count": cat_counts.get("__BLANK__", 0),
                "candidates": [],
            }
            ambiguous += 1
            continue

        top_cat, top_n = non_blank.most_common(1)[0]
        total = sum(non_blank.values())
        dominance = top_n / total if total else 0.0
        status = "confirmed" if dominance >= DOMINANCE_FLOOR else "ambiguous"
        if status == "confirmed":
            confirmed += 1
        else:
            ambiguous += 1

        map_out[account] = {
            "spend_category": top_cat,
            "status": status,
            "dominance": round(dominance, 3),
            "sample_count": top_n,
            "blank_count": cat_counts.get("__BLANK__", 0),
            "candidates": [
                {"spend_category": k, "count": v}
                for k, v in non_blank.most_common(8)
            ],
        }

    payload = {
        "map": map_out,
        "meta": {
            "source": str(path),
            "last_built": _now_iso(),
            "dominance_floor": DOMINANCE_FLOOR,
            "total_rows": total_rows,
            "expense_rows": expense_rows,
            "expense_accounts_seen": len(map_out),
            "confirmed": confirmed,
            "ambiguous": ambiguous,
        },
    }
    _atomic_write_json(_MAP_PATH, payload)

    return {
        "status": "ok",
        "output_path": str(_MAP_PATH),
        "total_rows": total_rows,
        "expense_rows": expense_rows,
        "expense_accounts_seen": len(map_out),
        "confirmed": confirmed,
        "ambiguous": ambiguous,
        "dominance_floor": DOMINANCE_FLOOR,
        "last_built": payload["meta"]["last_built"],
    }


# --------------------------------------------------------------------------- #
# Runtime lookup
# --------------------------------------------------------------------------- #


def lookup(gl_account: str, *, allow_ambiguous: bool = False) -> SpendCategoryResult:
    """Resolve a GL account → Spend Category at EIB-write time.

    Resolution order (first match wins):
        1. Operator override (always confirmed, dominance=1.0)
        2. Auto-derived map, status='confirmed'
        3. Auto-derived map, status='ambiguous' (only if allow_ambiguous=True)
        4. Unmapped — caller must park the line for manual review

    Returns SpendCategoryResult with status='not_required' for non-expense
    GL accounts (assets, liabilities, equity, income).
    """
    code = _normalize_account(gl_account)
    if not code:
        return SpendCategoryResult(
            spend_category=None, status="unmapped",
            dominance=0.0, sample_count=0, candidates=[],
        )
    if not _is_expense(code):
        return SpendCategoryResult(
            spend_category=None, status="not_required",
            dominance=1.0, sample_count=0, candidates=[],
        )

    overrides = _load_overrides()
    if code in overrides:
        rec = overrides[code]
        return SpendCategoryResult(
            spend_category=rec.get("spend_category"),
            status="override",
            dominance=1.0,
            sample_count=0,
            candidates=[(rec.get("spend_category", ""), 0)],
        )

    data = _load_map()
    rec = (data.get("map") or {}).get(code)
    if not rec:
        return SpendCategoryResult(
            spend_category=None, status="unmapped",
            dominance=0.0, sample_count=0, candidates=[],
        )

    cands = [(c.get("spend_category", ""), c.get("count", 0))
             for c in (rec.get("candidates") or [])]
    if rec.get("status") == "confirmed":
        return SpendCategoryResult(
            spend_category=rec.get("spend_category"),
            status="confirmed",
            dominance=float(rec.get("dominance") or 0.0),
            sample_count=int(rec.get("sample_count") or 0),
            candidates=cands,
        )
    # Ambiguous.
    if allow_ambiguous:
        return SpendCategoryResult(
            spend_category=rec.get("spend_category"),
            status="ambiguous",
            dominance=float(rec.get("dominance") or 0.0),
            sample_count=int(rec.get("sample_count") or 0),
            candidates=cands,
        )
    return SpendCategoryResult(
        spend_category=None,
        status="ambiguous",
        dominance=float(rec.get("dominance") or 0.0),
        sample_count=int(rec.get("sample_count") or 0),
        candidates=cands,
    )


# --------------------------------------------------------------------------- #
# Operator confirm / override
# --------------------------------------------------------------------------- #


def set_override(gl_account: str, spend_category: str, note: Optional[str] = None) -> dict:
    """Set or replace an operator-confirmed Spend Category for one GL account.

    Overrides always win, regardless of what the auto-derived map says.
    Use this to clear the ambiguous bucket and to fix any single-account
    mismatch the operator catches in QA.
    """
    code = _normalize_account(gl_account)
    if not code:
        raise ValueError("gl_account is required")
    if not spend_category or not str(spend_category).strip():
        raise ValueError("spend_category is required")

    overrides = _load_overrides()
    overrides[code] = {
        "spend_category": _normalize_category(spend_category),
        "set_at": _now_iso(),
        "note": (note or "").strip() or None,
    }
    _atomic_write_json(_OVERRIDES_PATH, overrides)
    return dict(overrides[code], gl_account=code)


def clear_override(gl_account: str) -> bool:
    code = _normalize_account(gl_account)
    overrides = _load_overrides()
    if code not in overrides:
        return False
    del overrides[code]
    _atomic_write_json(_OVERRIDES_PATH, overrides)
    return True


def list_ambiguous() -> list[dict]:
    """Every expense GL where dominance < 0.80, sorted by sample_count desc.

    Each entry includes the candidates list so the operator can pick the
    right Spend Category in one pass.
    """
    data = _load_map()
    out = []
    overrides = _load_overrides()
    for code, rec in (data.get("map") or {}).items():
        if rec.get("status") != "ambiguous":
            continue
        if code in overrides:
            continue  # already resolved
        out.append({
            "gl_account": code,
            "top_candidate": rec.get("spend_category"),
            "dominance": rec.get("dominance"),
            "sample_count": rec.get("sample_count"),
            "blank_count": rec.get("blank_count"),
            "candidates": rec.get("candidates"),
        })
    out.sort(key=lambda r: r.get("sample_count") or 0, reverse=True)
    return out


def stats() -> dict:
    data = _load_map()
    overrides = _load_overrides()
    meta = data.get("meta") or {}
    return {
        "confirmed": meta.get("confirmed", 0),
        "ambiguous": meta.get("ambiguous", 0),
        "overrides": len(overrides),
        "expense_rows_seen": meta.get("expense_rows", 0),
        "expense_accounts_seen": meta.get("expense_accounts_seen", 0),
        "total_rows": meta.get("total_rows", 0),
        "dominance_floor": meta.get("dominance_floor", DOMINANCE_FLOOR),
        "last_built": meta.get("last_built"),
        "source": meta.get("source"),
        "map_path": str(_MAP_PATH),
        "overrides_path": str(_OVERRIDES_PATH),
    }
