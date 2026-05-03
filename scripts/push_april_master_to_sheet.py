# © 2026 CoAssisted Workspace. Licensed under MIT.
"""Push staffwizard_overall_reports/master_april_2026.xlsx to a Google Sheet.

Reads the 3 tabs from the local xlsx and writes them into the freshly-
created Sheet at SHEET_ID. Adds the two non-default tabs first
(Project Rollup, Daily Totals), then writes all three by name.

Uses the same OAuth credentials the rest of the MCP uses, so no extra
auth setup is required.

Usage:
    python3 scripts/push_april_master_to_sheet.py \
        [--sheet-id 1-k-u7jttAjDpEchSBf61UA3K1SH3wx9K7dH6nE63Axk] \
        [--xlsx staffwizard_overall_reports/master_april_2026.xlsx]
"""

from __future__ import annotations

import argparse
import datetime as _dt
import pathlib
import sys

import openpyxl

PROJECT_ROOT = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import auth  # noqa: E402
from googleapiclient.discovery import build  # noqa: E402

DEFAULT_SHEET_ID = "1-k-u7jttAjDpEchSBf61UA3K1SH3wx9K7dH6nE63Axk"
DEFAULT_XLSX = "staffwizard_overall_reports/master_april_2026.xlsx"


def _coerce(v):
    """Convert openpyxl values to JSON-friendly Sheets values."""
    if v is None:
        return ""
    if isinstance(v, (_dt.date, _dt.datetime)):
        return v.isoformat()
    return v


def _load_tabs(xlsx_path: pathlib.Path) -> dict[str, list[list]]:
    """Read every sheet in the .xlsx into a 2D array."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    out = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([_coerce(c) for c in row])
        out[name] = rows
    return out


def _ensure_tabs(svc, sheet_id: str, want_tabs: list[str]) -> dict[str, int]:
    """Make sure every tab in want_tabs exists. Returns name → sheetId map.

    The default 'Sheet1' is kept (we'll rename it to the first wanted tab
    if it's still around).
    """
    meta = svc.spreadsheets().get(spreadsheetId=sheet_id).execute()
    existing = {s["properties"]["title"]: s["properties"]["sheetId"]
                for s in meta.get("sheets", [])}

    requests = []

    # If 'Sheet1' is the only existing tab and we want renaming,
    # rename it to the first wanted tab.
    if "Sheet1" in existing and want_tabs and want_tabs[0] not in existing:
        requests.append({
            "updateSheetProperties": {
                "properties": {
                    "sheetId": existing["Sheet1"],
                    "title": want_tabs[0],
                },
                "fields": "title",
            }
        })
        existing[want_tabs[0]] = existing.pop("Sheet1")

    # Add any missing tabs.
    for t in want_tabs:
        if t not in existing:
            requests.append({"addSheet": {"properties": {"title": t}}})

    if requests:
        resp = svc.spreadsheets().batchUpdate(
            spreadsheetId=sheet_id, body={"requests": requests},
        ).execute()
        # Refresh existing tabs from the response.
        meta = svc.spreadsheets().get(spreadsheetId=sheet_id).execute()
        existing = {s["properties"]["title"]: s["properties"]["sheetId"]
                    for s in meta.get("sheets", [])}

    return existing


def _write_tab(svc, sheet_id: str, tab_name: str, rows: list[list]) -> int:
    """Overwrite a tab's contents starting at A1. Returns updated cell count."""
    if not rows:
        return 0
    rng = f"'{tab_name}'!A1"
    body = {"values": rows}
    resp = svc.spreadsheets().values().update(
        spreadsheetId=sheet_id,
        range=rng,
        valueInputOption="USER_ENTERED",
        body=body,
    ).execute()
    return resp.get("updatedCells", 0)


def main(argv=None) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--sheet-id", default=DEFAULT_SHEET_ID)
    p.add_argument("--xlsx", default=DEFAULT_XLSX)
    args = p.parse_args(argv)

    xlsx_path = PROJECT_ROOT / args.xlsx
    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} not found", file=sys.stderr)
        return 2

    print(f"Reading: {xlsx_path}")
    tabs = _load_tabs(xlsx_path)
    for name, rows in tabs.items():
        print(f"  {name}: {len(rows)} rows × {len(rows[0]) if rows else 0} cols")

    print(f"\nConnecting to Google Sheets...")
    creds = auth.get_credentials()
    svc = build("sheets", "v4", credentials=creds, cache_discovery=False)

    print(f"Ensuring tabs exist on {args.sheet_id}...")
    _ensure_tabs(svc, args.sheet_id, list(tabs.keys()))

    print(f"\nWriting tabs:")
    total_cells = 0
    for name, rows in tabs.items():
        n = _write_tab(svc, args.sheet_id, name, rows)
        total_cells += n
        print(f"  ✓ {name}: {n:,} cells")

    print(f"\nDone — {total_cells:,} cells written.")
    print(f"Open: https://docs.google.com/spreadsheets/d/{args.sheet_id}/edit")
    return 0


if __name__ == "__main__":
    sys.exit(main())
