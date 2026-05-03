# © 2026 CoAssisted Workspace. Licensed under MIT.
"""Build visual HTML + JSON dashboards from the master xlsx.

Output:
  dashboards/
    data.json                    full structured payload
    index.html                   overview — all projects ranked + flagged items
    projects/<slug>.html         one detail page per project

The dashboards/ folder is uploaded to a 'Surefox Daily Operations Dashboards'
folder in the operator's Drive; the share-friendly folder URL is printed at
the end. Managers bookmark the index.html within that folder.

Three action-item flag rules (Joshua 2026-05-02):
  low_margin    — margin % under 10% over the full window
  high_ot       — OT hours > 15% of total hours
  stalled       — no shifts in the last 5 days

Pages are fully self-contained — Chart.js loads from CDN, all CSS is
inlined. Re-running the script overwrites the existing dashboards in
both local + Drive.

Usage:
    python3 scripts/build_project_dashboards.py
"""

from __future__ import annotations

import argparse
import datetime as _dt
import html as _html
import json
import pathlib
import re
import sys
from collections import defaultdict
from typing import Any

import openpyxl

PROJECT_ROOT = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import auth  # noqa: E402
from googleapiclient.discovery import build  # noqa: E402
from googleapiclient.http import MediaFileUpload  # noqa: E402

REPORTS_DIR = PROJECT_ROOT / "staffwizard_overall_reports"
OUTPUT_DIR = PROJECT_ROOT / "dashboards"
PROJECTS_DIR = OUTPUT_DIR / "projects"
DRIVE_FOLDER_NAME = "Surefox Daily Operations Dashboards"

LOW_MARGIN_THRESHOLD = 10.0   # percent
HIGH_OT_THRESHOLD = 15.0      # percent
STALLED_DAYS = 5


# -----------------------------------------------------------------------------
# Slug + safe HTML helpers
# -----------------------------------------------------------------------------

def slugify(name: str) -> str:
    s = name.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-") or "untitled"


def esc(s: Any) -> str:
    return _html.escape(str(s) if s is not None else "")


def fmt_money(n: float) -> str:
    sign = "-" if n < 0 else ""
    return f"{sign}${abs(n):,.0f}"


def fmt_pct(n: float) -> str:
    return f"{n:+.1f}%"


def fmt_hours(n: float) -> str:
    return f"{n:,.1f}"


# -----------------------------------------------------------------------------
# Read master xlsx → per-project structures
# -----------------------------------------------------------------------------

def _read_detail_rows(xlsx_path: pathlib.Path) -> list[dict]:
    """Pull the Daily Detail tab rows back out of the master xlsx."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Daily Detail" not in wb.sheetnames:
        raise SystemExit(f"'Daily Detail' tab not found in {xlsx_path}")
    ws = wb["Daily Detail"]
    rows = []
    headers: list[str] | None = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h) if h is not None else "" for h in row]
            continue
        rec = {h: v for h, v in zip(headers, row)}
        rows.append(rec)
    return rows


def _build_project_payload(detail_rows: list[dict], window_days: int) -> dict:
    """Compute everything we'll need for both index + per-project pages."""
    today = _dt.date.today()
    cutoff = today - _dt.timedelta(days=window_days)

    # Group rows by project (Job Description, falling back to Job Number).
    projects: dict[str, dict] = defaultdict(
        lambda: {
            "name": "",
            "shifts": 0, "total_hours": 0.0, "reg_h": 0.0, "ot_h": 0.0, "dt_h": 0.0,
            "reg_cost": 0.0, "holiday_cost": 0.0, "ot_cost": 0.0, "dt_cost": 0.0,
            "total_cost": 0.0, "revenue": 0.0, "margin": 0.0,
            "dates": set(), "employees": set(),
            "daily": defaultdict(lambda: {
                "shifts": 0, "reg_h": 0.0, "ot_h": 0.0, "dt_h": 0.0,
                "total_h": 0.0, "cost": 0.0, "revenue": 0.0, "margin": 0.0,
            }),
            "recent_shifts": [],
        }
    )

    overall = {
        "shifts": 0, "total_hours": 0.0, "ot_h": 0.0,
        "revenue": 0.0, "cost": 0.0, "margin": 0.0,
        "dates": set(), "employees": set(),
    }

    for d in detail_rows:
        date_s = d.get("Date") or ""
        if not date_s:
            continue
        try:
            date_d = _dt.date.fromisoformat(date_s)
        except (TypeError, ValueError):
            continue

        proj_name = d.get("Job Description") or d.get("Job Number") or "(unlabeled)"
        p = projects[proj_name]
        p["name"] = proj_name

        reg_h = float(d.get("Reg Hours") or 0)
        ot_h = float(d.get("OT Hours") or 0)
        dt_h = float(d.get("DT Hours") or 0)
        total_h = float(d.get("Total Hours") or 0)
        reg_c = float(d.get("Reg Cost $") or 0)
        hol_c = float(d.get("Holiday Cost $") or 0)
        ot_c = float(d.get("OT Cost $") or 0)
        dt_c = float(d.get("DT Cost $") or 0)
        tot_c = float(d.get("Total Cost $") or 0)
        rev = float(d.get("Billable $") or 0)
        margin = float(d.get("Margin $") or 0)

        p["shifts"] += 1
        p["reg_h"] += reg_h
        p["ot_h"] += ot_h
        p["dt_h"] += dt_h
        p["total_hours"] += total_h
        p["reg_cost"] += reg_c
        p["holiday_cost"] += hol_c
        p["ot_cost"] += ot_c
        p["dt_cost"] += dt_c
        p["total_cost"] += tot_c
        p["revenue"] += rev
        p["margin"] += margin
        p["dates"].add(date_s)
        if d.get("Employee #"):
            p["employees"].add(d["Employee #"])
        # Daily series.
        ds = p["daily"][date_s]
        ds["shifts"] += 1
        ds["reg_h"] += reg_h
        ds["ot_h"] += ot_h
        ds["dt_h"] += dt_h
        ds["total_h"] += total_h
        ds["cost"] += tot_c
        ds["revenue"] += rev
        ds["margin"] += margin

        overall["shifts"] += 1
        overall["total_hours"] += total_h
        overall["ot_h"] += ot_h
        overall["revenue"] += rev
        overall["cost"] += tot_c
        overall["margin"] += margin
        overall["dates"].add(date_s)
        if d.get("Employee #"):
            overall["employees"].add(d["Employee #"])

        # Track shift detail for "recent shifts" section.
        p["recent_shifts"].append({
            "date": date_s,
            "employee": d.get("Employee Name") or "",
            "post": d.get("Post") or "",
            "shift_start": d.get("Shift Start") or "",
            "shift_end": d.get("Shift End") or "",
            "total_hours": total_h,
            "total_cost": tot_c,
            "billable": rev,
            "margin": margin,
        })

    # Compute per-project derived fields + flags + sorted views.
    project_list = []
    for p in projects.values():
        margin_pct = (p["margin"] / p["revenue"] * 100) if p["revenue"] else 0.0
        ot_pct = (p["ot_h"] / p["total_hours"] * 100) if p["total_hours"] else 0.0
        days_active = len(p["dates"])
        last_shift = max(p["dates"]) if p["dates"] else None

        flags: list[dict] = []
        if margin_pct < LOW_MARGIN_THRESHOLD:
            severity = "critical" if margin_pct < 0 else "warning"
            flags.append({
                "code": "low_margin",
                "severity": severity,
                "title": "Low margin" if margin_pct >= 0 else "Losing money",
                "detail": (
                    f"Margin is {margin_pct:.1f}% over the window — "
                    f"under the {LOW_MARGIN_THRESHOLD:.0f}% target."
                ),
            })
        if ot_pct > HIGH_OT_THRESHOLD:
            flags.append({
                "code": "high_ot",
                "severity": "warning",
                "title": "High overtime",
                "detail": (
                    f"OT hours are {ot_pct:.1f}% of total — over the "
                    f"{HIGH_OT_THRESHOLD:.0f}% threshold."
                ),
            })
        if last_shift:
            try:
                last_d = _dt.date.fromisoformat(last_shift)
                days_since = (today - last_d).days
                if days_since >= STALLED_DAYS:
                    flags.append({
                        "code": "stalled",
                        "severity": "info",
                        "title": "Stalled",
                        "detail": (
                            f"No shifts in {days_since} days "
                            f"(last shift {last_shift})."
                        ),
                    })
            except ValueError:
                pass

        # Daily series sorted by date asc, ready for charting.
        daily_series = []
        for date_s in sorted(p["daily"].keys()):
            ds = p["daily"][date_s]
            daily_series.append({
                "date": date_s,
                "shifts": ds["shifts"],
                "reg_h": round(ds["reg_h"], 2),
                "ot_h": round(ds["ot_h"], 2),
                "dt_h": round(ds["dt_h"], 2),
                "total_h": round(ds["total_h"], 2),
                "cost": round(ds["cost"], 2),
                "revenue": round(ds["revenue"], 2),
                "margin": round(ds["margin"], 2),
            })

        # Recent shifts (last 14 days, newest first).
        recent_cutoff = today - _dt.timedelta(days=14)
        recent = sorted(
            (s for s in p["recent_shifts"]
             if _dt.date.fromisoformat(s["date"]) >= recent_cutoff),
            key=lambda s: s["date"], reverse=True,
        )[:50]

        project_list.append({
            "name": p["name"],
            "slug": slugify(p["name"]),
            "shifts": p["shifts"],
            "days_active": days_active,
            "last_shift": last_shift,
            "employees": len(p["employees"]),
            "reg_h": round(p["reg_h"], 2),
            "ot_h": round(p["ot_h"], 2),
            "dt_h": round(p["dt_h"], 2),
            "total_hours": round(p["total_hours"], 2),
            "reg_cost": round(p["reg_cost"], 2),
            "holiday_cost": round(p["holiday_cost"], 2),
            "ot_cost": round(p["ot_cost"], 2),
            "dt_cost": round(p["dt_cost"], 2),
            "total_cost": round(p["total_cost"], 2),
            "revenue": round(p["revenue"], 2),
            "margin": round(p["margin"], 2),
            "margin_pct": round(margin_pct, 1),
            "ot_pct": round(ot_pct, 1),
            "flags": flags,
            "daily_series": daily_series,
            "recent_shifts": recent,
        })

    project_list.sort(key=lambda p: p["revenue"], reverse=True)

    overall_margin_pct = (overall["margin"] / overall["revenue"] * 100) if overall["revenue"] else 0.0
    overall_ot_pct = (overall["ot_h"] / overall["total_hours"] * 100) if overall["total_hours"] else 0.0

    # Build daily totals across all projects — one row per active date.
    # Used by the Summary tab's trend chart (Revenue / Cost / Margin).
    daily_totals_by_date: dict[str, dict] = defaultdict(
        lambda: {"revenue": 0.0, "cost": 0.0, "margin": 0.0,
                 "total_h": 0.0, "ot_h": 0.0, "shifts": 0}
    )
    for p in project_list:
        for d in p["daily_series"]:
            t = daily_totals_by_date[d["date"]]
            t["revenue"] += d["revenue"]
            t["cost"] += d["cost"]
            t["margin"] += d["margin"]
            t["total_h"] += d["total_h"]
            t["ot_h"] += d["ot_h"]
            t["shifts"] += d["shifts"]
    daily_totals_series = [
        {
            "date": date_s,
            "revenue": round(t["revenue"], 2),
            "cost": round(t["cost"], 2),
            "margin": round(t["margin"], 2),
            "total_h": round(t["total_h"], 2),
            "ot_h": round(t["ot_h"], 2),
            "shifts": t["shifts"],
        }
        for date_s, t in sorted(daily_totals_by_date.items())
    ]

    days_active = len(overall["dates"]) or 1
    avg_cost_per_day = overall["cost"] / days_active
    avg_revenue_per_day = overall["revenue"] / days_active

    return {
        "generated_at": _dt.datetime.now().astimezone().isoformat(timespec="seconds"),
        "window": {
            "start": cutoff.isoformat(),
            "end": today.isoformat(),
            "days": window_days,
        },
        "totals": {
            "shifts": overall["shifts"],
            "total_hours": round(overall["total_hours"], 2),
            "ot_h": round(overall["ot_h"], 2),
            "ot_pct": round(overall_ot_pct, 1),
            "revenue": round(overall["revenue"], 2),
            "cost": round(overall["cost"], 2),
            "margin": round(overall["margin"], 2),
            "margin_pct": round(overall_margin_pct, 1),
            "days": len(overall["dates"]),
            "projects": len(project_list),
            "employees": len(overall["employees"]),
            "avg_cost_per_day": round(avg_cost_per_day, 2),
            "avg_revenue_per_day": round(avg_revenue_per_day, 2),
        },
        "daily_totals": daily_totals_series,
        "projects": project_list,
    }


# -----------------------------------------------------------------------------
# HTML emission
# -----------------------------------------------------------------------------

CSS = """
:root {
  --text: #111;
  --text-soft: #555;
  --bg: #ffffff;
  --tile: #f7f7f7;
  --border: #e5e5e5;
  --accent: #2563eb;
  --positive: #047857;
  --warning: #d97706;
  --critical: #dc2626;
  --info: #6b7280;
}
* { box-sizing: border-box; }
body {
  margin: 0; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI',
    Helvetica, Arial, sans-serif;
  color: var(--text); background: var(--bg);
  line-height: 1.45; font-size: 14px;
}
.container { max-width: 1200px; margin: 0 auto; padding: 32px 24px 80px; }
h1, h2, h3 { font-weight: 600; letter-spacing: -0.01em; margin: 0; }
h1 { font-size: 28px; }
h2 { font-size: 18px; margin-top: 32px; }
h3 { font-size: 15px; }
.muted { color: var(--text-soft); font-size: 13px; }
a { color: var(--accent); text-decoration: none; }
a:hover { text-decoration: underline; }
header.hero {
  display: flex; justify-content: space-between; align-items: flex-end;
  border-bottom: 1px solid var(--border); padding-bottom: 16px; margin-bottom: 24px;
}
.kpi-grid {
  display: grid; gap: 12px;
  grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
  margin-bottom: 24px;
}
.kpi {
  background: var(--tile); border: 1px solid var(--border);
  border-radius: 8px; padding: 14px 16px;
}
.kpi .label { font-size: 12px; color: var(--text-soft); text-transform: uppercase;
  letter-spacing: 0.04em; margin-bottom: 4px; }
.kpi .value { font-size: 22px; font-weight: 600; }
.kpi .sub { font-size: 12px; color: var(--text-soft); margin-top: 4px; }
.value.positive { color: var(--positive); }
.value.critical { color: var(--critical); }
.value.warning { color: var(--warning); }
table { width: 100%; border-collapse: collapse; font-size: 13px; margin-top: 8px; }
th, td {
  text-align: left; padding: 8px 10px;
  border-bottom: 1px solid var(--border); white-space: nowrap;
}
th { font-weight: 600; color: var(--text-soft); font-size: 12px;
  text-transform: uppercase; letter-spacing: 0.04em; }
td.num, th.num { text-align: right; font-variant-numeric: tabular-nums; }
tr:hover td { background: #fafafa; }
.flag {
  display: inline-flex; gap: 6px; align-items: center;
  padding: 6px 10px; border-radius: 999px; font-size: 12px;
  font-weight: 600; border: 1px solid;
}
.flag.warning { color: var(--warning); border-color: var(--warning); background: #fef3c7; }
.flag.critical { color: var(--critical); border-color: var(--critical); background: #fee2e2; }
.flag.info { color: var(--info); border-color: var(--info); background: #f3f4f6; }
.flag-list { display: flex; gap: 8px; flex-wrap: wrap; }
.flag-card {
  border: 1px solid; border-radius: 8px; padding: 12px 14px;
  margin-bottom: 10px;
}
.flag-card.warning { border-color: var(--warning); background: #fffbeb; }
.flag-card.critical { border-color: var(--critical); background: #fef2f2; }
.flag-card.info { border-color: var(--info); background: #f9fafb; }
.flag-card strong { display: block; margin-bottom: 2px; }
.charts { display: grid; gap: 24px; grid-template-columns: 1fr; margin-top: 24px; }
@media (min-width: 880px) { .charts { grid-template-columns: 1fr 1fr; } }
.chart-card {
  border: 1px solid var(--border); border-radius: 8px; padding: 16px;
}
.chart-card h3 { margin-bottom: 8px; }
.back-link { font-size: 13px; }
footer { margin-top: 48px; padding-top: 24px; border-top: 1px solid var(--border);
  font-size: 12px; color: var(--text-soft); }
.summary { color: var(--text-soft); font-size: 13px; margin-top: 4px; }
.no-flags { color: var(--positive); font-size: 13px; }
nav.tabs {
  display: flex; gap: 0; border-bottom: 1px solid var(--border);
  margin-bottom: 24px; flex-wrap: wrap;
}
.tab-btn {
  padding: 10px 16px; background: transparent; border: none;
  border-bottom: 2px solid transparent; cursor: pointer;
  color: var(--text-soft); font-weight: 600; font-size: 14px;
  font-family: inherit;
}
.tab-btn:hover { color: var(--text); }
.tab-btn.active { color: var(--accent); border-bottom-color: var(--accent); }
.tab-panel { display: none; }
.tab-panel.active { display: block; }
.chart-card.wide { margin: 24px 0; }
.chart-card.wide canvas { height: 320px !important; }
"""


def _kpi(label: str, value: str, sub: str = "", tone: str = "") -> str:
    tone_cls = f" {tone}" if tone else ""
    return (
        f'<div class="kpi">'
        f'<div class="label">{esc(label)}</div>'
        f'<div class="value{tone_cls}">{esc(value)}</div>'
        f'{f"<div class=sub>{esc(sub)}</div>" if sub else ""}'
        f'</div>'
    )


def _margin_tone(pct: float) -> str:
    if pct < 0:
        return "critical"
    if pct < LOW_MARGIN_THRESHOLD:
        return "warning"
    return "positive"


def _render_index(payload: dict) -> str:
    t = payload["totals"]
    w = payload["window"]
    daily = payload.get("daily_totals", [])
    margin_tone = _margin_tone(t["margin_pct"])

    # KPI grid — Summary tab. Now 10 tiles (added Avg Cost/Day + Avg Billed/Day).
    kpis = "".join([
        _kpi("Revenue", fmt_money(t["revenue"]),
             f"{w['days']}-day window"),
        _kpi("Margin", fmt_money(t["margin"]),
             fmt_pct(t["margin_pct"]),
             tone=margin_tone),
        _kpi("Cost", fmt_money(t["cost"])),
        _kpi("Avg Billed / Day", fmt_money(t.get("avg_revenue_per_day", 0)),
             f"across {t['days']} active days"),
        _kpi("Avg Cost / Day", fmt_money(t.get("avg_cost_per_day", 0)),
             f"across {t['days']} active days"),
        _kpi("Total Hours", fmt_hours(t["total_hours"]),
             f"{t['shifts']:,} shifts"),
        _kpi("OT Hours", fmt_hours(t["ot_h"]),
             f"{t['ot_pct']:.1f}% of total",
             tone="warning" if t["ot_pct"] > HIGH_OT_THRESHOLD else ""),
        _kpi("Projects", str(t["projects"])),
        _kpi("Employees", str(t["employees"])),
        _kpi("Days Active", str(t["days"])),
    ])

    # Daily-totals chart data (Revenue / Cost / Margin lines).
    chart_labels = json.dumps([d["date"] for d in daily])
    chart_revenue = json.dumps([d["revenue"] for d in daily])
    chart_cost = json.dumps([d["cost"] for d in daily])
    chart_margin = json.dumps([d["margin"] for d in daily])

    # Project ranking table.
    rows = []
    for p in payload["projects"]:
        flag_chips = "".join(
            f'<span class="flag {f["severity"]}">{esc(f["title"])}</span>'
            for f in p["flags"]
        )
        m_tone = _margin_tone(p["margin_pct"])
        rows.append(
            f'<tr>'
            f'<td><a href="projects/{esc(p["slug"])}.html">{esc(p["name"])}</a></td>'
            f'<td class="num">{p["days_active"]}</td>'
            f'<td class="num">{p["employees"]}</td>'
            f'<td class="num">{fmt_hours(p["total_hours"])}</td>'
            f'<td class="num">{p["ot_pct"]:.1f}%</td>'
            f'<td class="num">{fmt_money(p["revenue"])}</td>'
            f'<td class="num">{fmt_money(p["total_cost"])}</td>'
            f'<td class="num"><span class="value {m_tone}">{fmt_money(p["margin"])}</span></td>'
            f'<td class="num"><span class="value {m_tone}">{p["margin_pct"]:.1f}%</span></td>'
            f'<td class="num">'
            f'<div class="flag-list">{flag_chips}</div></td>'
            f'</tr>'
        )

    # Flagged-project section (Action Items tab).
    flagged = [p for p in payload["projects"] if p["flags"]]
    if flagged:
        flag_blocks = []
        for p in flagged:
            cards = "".join(
                f'<div class="flag-card {f["severity"]}">'
                f'<strong>{esc(f["title"])}</strong>'
                f'<span class="muted">{esc(f["detail"])}</span>'
                f'</div>'
                for f in p["flags"]
            )
            flag_blocks.append(
                f'<div style="margin-bottom:18px">'
                f'<h3><a href="projects/{esc(p["slug"])}.html">{esc(p["name"])}</a> '
                f'<span class="muted">— {fmt_money(p["revenue"])} revenue · {p["margin_pct"]:.1f}% margin</span></h3>'
                f'{cards}'
                f'</div>'
            )
        flagged_html = (
            f'<div class="summary">{len(flagged)} of {t["projects"]} projects need attention.</div>'
            f'{"".join(flag_blocks)}'
        )
    else:
        flagged_html = '<div class="no-flags">All projects clean — no flags.</div>'

    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Surefox Daily Operations — Overview</title>
  <style>{CSS}</style>
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
</head>
<body>
  <div class="container">
    <header class="hero">
      <div>
        <h1>Surefox Daily Operations</h1>
        <div class="muted">{esc(w['start'])} → {esc(w['end'])} · {w['days']}-day rolling window</div>
      </div>
      <div class="muted">Generated {esc(payload['generated_at'])}</div>
    </header>

    <nav class="tabs" role="tablist">
      <button class="tab-btn active" data-tab="summary" role="tab">Summary</button>
      <button class="tab-btn" data-tab="ranked" role="tab">All Projects Ranked by Revenue</button>
      <button class="tab-btn" data-tab="actions" role="tab">Action Items</button>
    </nav>

    <section class="tab-panel active" id="summary" role="tabpanel">
      <div class="kpi-grid">{kpis}</div>
      <div class="chart-card wide">
        <h3>Daily Revenue · Cost · Margin</h3>
        <canvas id="dailyChart"></canvas>
      </div>
    </section>

    <section class="tab-panel" id="ranked" role="tabpanel">
      <table>
        <thead><tr>
          <th>Project</th>
          <th class="num">Days</th>
          <th class="num">Employees</th>
          <th class="num">Hours</th>
          <th class="num">OT %</th>
          <th class="num">Revenue</th>
          <th class="num">Cost</th>
          <th class="num">Margin</th>
          <th class="num">Margin %</th>
          <th class="num">Flags</th>
        </tr></thead>
        <tbody>{"".join(rows)}</tbody>
      </table>
    </section>

    <section class="tab-panel" id="actions" role="tabpanel">
      {flagged_html}
    </section>

    <footer>
      Source: <code>master_april_2026.xlsx</code> via
      <code>scripts/build_project_dashboards.py</code>.
      Data refreshes when the script is re-run.
    </footer>
  </div>

  <script>
    // Tab switching.
    document.querySelectorAll('.tab-btn').forEach(btn => {{
      btn.addEventListener('click', () => {{
        document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
        document.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
        btn.classList.add('active');
        document.getElementById(btn.dataset.tab).classList.add('active');
      }});
    }});

    // Daily totals chart.
    const labels = {chart_labels};
    new Chart(document.getElementById('dailyChart'), {{
      type: 'line',
      data: {{
        labels,
        datasets: [
          {{ label: 'Revenue', data: {chart_revenue}, borderColor: '#2563eb',
            backgroundColor: 'rgba(37,99,235,0.08)', fill: true, tension: 0.2 }},
          {{ label: 'Cost', data: {chart_cost}, borderColor: '#d97706',
            backgroundColor: 'rgba(217,119,6,0.08)', fill: true, tension: 0.2 }},
          {{ label: 'Margin', data: {chart_margin}, borderColor: '#047857',
            backgroundColor: 'rgba(4,120,87,0.08)', fill: true, tension: 0.2 }},
        ]
      }},
      options: {{
        responsive: true,
        maintainAspectRatio: false,
        plugins: {{ legend: {{ position: 'bottom' }} }},
        scales: {{
          y: {{ ticks: {{ callback: v => '$' + v.toLocaleString() }} }}
        }}
      }}
    }});
  </script>
</body>
</html>
"""


def _render_project_page(payload: dict, p: dict) -> str:
    w = payload["window"]
    margin_tone = _margin_tone(p["margin_pct"])
    daily = p["daily_series"]
    labels = json.dumps([d["date"] for d in daily])
    reg_data = json.dumps([d["reg_h"] for d in daily])
    ot_data = json.dumps([d["ot_h"] for d in daily])
    dt_data = json.dumps([d["dt_h"] for d in daily])
    rev_data = json.dumps([d["revenue"] for d in daily])
    margin_data = json.dumps([d["margin"] for d in daily])

    # KPI grid.
    kpis = "".join([
        _kpi("Revenue", fmt_money(p["revenue"]), f"{p['days_active']} active days"),
        _kpi("Margin", fmt_money(p["margin"]), fmt_pct(p["margin_pct"]), tone=margin_tone),
        _kpi("Cost", fmt_money(p["total_cost"])),
        _kpi("Hours", fmt_hours(p["total_hours"]), f"{p['shifts']:,} shifts"),
        _kpi("OT Hours", fmt_hours(p["ot_h"]), f"{p['ot_pct']:.1f}% of total",
             tone="warning" if p["ot_pct"] > HIGH_OT_THRESHOLD else ""),
        _kpi("Holiday Cost", fmt_money(p["holiday_cost"])),
        _kpi("Employees", str(p["employees"])),
        _kpi("Last Shift", p["last_shift"] or "—"),
    ])

    # Flags.
    if p["flags"]:
        flag_html = "".join(
            f'<div class="flag-card {f["severity"]}">'
            f'<strong>{esc(f["title"])}</strong>'
            f'<span class="muted">{esc(f["detail"])}</span>'
            f'</div>'
            for f in p["flags"]
        )
    else:
        flag_html = '<div class="no-flags">No flags — project on track.</div>'

    # Recent shifts table.
    shift_rows = "".join(
        f'<tr>'
        f'<td>{esc(s["date"])}</td>'
        f'<td>{esc(s["employee"])}</td>'
        f'<td>{esc(s["post"])}</td>'
        f'<td>{esc(s["shift_start"])} – {esc(s["shift_end"])}</td>'
        f'<td class="num">{fmt_hours(s["total_hours"])}</td>'
        f'<td class="num">{fmt_money(s["total_cost"])}</td>'
        f'<td class="num">{fmt_money(s["billable"])}</td>'
        f'<td class="num">{fmt_money(s["margin"])}</td>'
        f'</tr>'
        for s in p["recent_shifts"]
    )
    if not shift_rows:
        shift_rows = '<tr><td colspan="8" class="muted">No shifts in the last 14 days.</td></tr>'

    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{esc(p['name'])} — Surefox Daily Operations</title>
  <style>{CSS}</style>
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
</head>
<body>
  <div class="container">
    <a href="../index.html" class="back-link">← Back to overview</a>
    <header class="hero">
      <div>
        <h1>{esc(p['name'])}</h1>
        <div class="muted">{esc(w['start'])} → {esc(w['end'])} · {w['days']}-day rolling window</div>
      </div>
      <div class="muted">Generated {esc(payload['generated_at'])}</div>
    </header>

    <div class="kpi-grid">{kpis}</div>

    <h2>Action items</h2>
    {flag_html}

    <div class="charts">
      <div class="chart-card">
        <h3>Daily hours · stacked by type</h3>
        <canvas id="hoursChart" height="220"></canvas>
      </div>
      <div class="chart-card">
        <h3>Daily revenue & margin</h3>
        <canvas id="revChart" height="220"></canvas>
      </div>
    </div>

    <h2>Recent shifts <span class="muted">— last 14 days</span></h2>
    <table>
      <thead><tr>
        <th>Date</th>
        <th>Employee</th>
        <th>Post</th>
        <th>Shift</th>
        <th class="num">Hours</th>
        <th class="num">Cost</th>
        <th class="num">Billable</th>
        <th class="num">Margin</th>
      </tr></thead>
      <tbody>{shift_rows}</tbody>
    </table>

    <footer>
      Source: <code>master_april_2026.xlsx</code> via
      <code>scripts/build_project_dashboards.py</code>.
    </footer>
  </div>

  <script>
    const labels = {labels};
    new Chart(document.getElementById('hoursChart'), {{
      type: 'bar',
      data: {{
        labels,
        datasets: [
          {{ label: 'Reg', data: {reg_data}, backgroundColor: '#2563eb' }},
          {{ label: 'OT',  data: {ot_data},  backgroundColor: '#d97706' }},
          {{ label: 'DT',  data: {dt_data},  backgroundColor: '#dc2626' }},
        ]
      }},
      options: {{
        responsive: true,
        scales: {{ x: {{ stacked: true }}, y: {{ stacked: true }} }},
        plugins: {{ legend: {{ position: 'bottom' }} }},
      }}
    }});
    new Chart(document.getElementById('revChart'), {{
      type: 'line',
      data: {{
        labels,
        datasets: [
          {{ label: 'Revenue', data: {rev_data}, borderColor: '#2563eb',
            backgroundColor: 'rgba(37,99,235,0.1)', fill: true, tension: 0.2 }},
          {{ label: 'Margin', data: {margin_data}, borderColor: '#047857',
            backgroundColor: 'rgba(4,120,87,0.1)', fill: true, tension: 0.2 }},
        ]
      }},
      options: {{
        responsive: true,
        plugins: {{ legend: {{ position: 'bottom' }} }},
      }}
    }});
  </script>
</body>
</html>
"""


# -----------------------------------------------------------------------------
# Drive upload
# -----------------------------------------------------------------------------

def _ensure_drive_folder(svc, folder_name: str) -> str:
    """Find or create a top-level Drive folder. Returns its ID."""
    q = (
        f"mimeType='application/vnd.google-apps.folder' "
        f"and name='{folder_name}' and trashed=false"
    )
    resp = svc.files().list(q=q, fields="files(id,name)", pageSize=1).execute()
    files = resp.get("files", [])
    if files:
        return files[0]["id"]
    folder = svc.files().create(
        body={
            "name": folder_name,
            "mimeType": "application/vnd.google-apps.folder",
        },
        fields="id",
    ).execute()
    return folder["id"]


def _ensure_subfolder(svc, parent_id: str, name: str) -> str:
    q = (
        f"mimeType='application/vnd.google-apps.folder' "
        f"and name='{name}' and '{parent_id}' in parents and trashed=false"
    )
    resp = svc.files().list(q=q, fields="files(id,name)", pageSize=1).execute()
    files = resp.get("files", [])
    if files:
        return files[0]["id"]
    folder = svc.files().create(
        body={
            "name": name,
            "mimeType": "application/vnd.google-apps.folder",
            "parents": [parent_id],
        },
        fields="id",
    ).execute()
    return folder["id"]


def _upload_or_replace(
    svc, parent_id: str, local_path: pathlib.Path, mime: str,
) -> str:
    """Upload a file to Drive, replacing any existing file with the same name."""
    name = local_path.name
    q = f"name='{name}' and '{parent_id}' in parents and trashed=false"
    resp = svc.files().list(q=q, fields="files(id,name)", pageSize=1).execute()
    media = MediaFileUpload(str(local_path), mimetype=mime)
    if resp.get("files"):
        fid = resp["files"][0]["id"]
        svc.files().update(fileId=fid, media_body=media).execute()
        return fid
    new = svc.files().create(
        body={"name": name, "parents": [parent_id]},
        media_body=media,
        fields="id",
    ).execute()
    return new["id"]


# -----------------------------------------------------------------------------
# main
# -----------------------------------------------------------------------------

def main(argv=None) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--xlsx", default=str(REPORTS_DIR / "master_april_2026.xlsx"))
    p.add_argument("--window-days", type=int, default=90)
    p.add_argument("--no-upload", action="store_true",
                   help="Skip Drive upload, just emit local files.")
    args = p.parse_args(argv)

    xlsx_path = pathlib.Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} not found.")
        print("Run scripts/refresh_staffwizard_master.py first to build the master xlsx.")
        return 2

    print(f"Reading {xlsx_path}...")
    detail_rows = _read_detail_rows(xlsx_path)
    print(f"  {len(detail_rows):,} detail rows")

    print(f"Computing payload (window: {args.window_days} days)...")
    payload = _build_project_payload(detail_rows, args.window_days)
    print(f"  {payload['totals']['projects']} projects, "
          f"{sum(1 for p in payload['projects'] if p['flags'])} flagged")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    PROJECTS_DIR.mkdir(parents=True, exist_ok=True)

    # data.json
    json_path = OUTPUT_DIR / "data.json"
    json_path.write_text(json.dumps(payload, indent=2, default=str))
    print(f"Wrote {json_path}")

    # index.html
    idx_path = OUTPUT_DIR / "index.html"
    idx_path.write_text(_render_index(payload))
    print(f"Wrote {idx_path}")

    # per-project pages
    for proj in payload["projects"]:
        out = PROJECTS_DIR / f"{proj['slug']}.html"
        out.write_text(_render_project_page(payload, proj))
    print(f"Wrote {len(payload['projects'])} per-project pages to {PROJECTS_DIR}")

    if args.no_upload:
        print("\n(--no-upload set; skipping Drive upload.)")
        return 0

    print(f"\nUploading to Drive...")
    creds = auth.get_credentials()
    svc = build("drive", "v3", credentials=creds, cache_discovery=False)

    folder_id = _ensure_drive_folder(svc, DRIVE_FOLDER_NAME)
    proj_folder_id = _ensure_subfolder(svc, folder_id, "projects")

    _upload_or_replace(svc, folder_id, json_path, "application/json")
    _upload_or_replace(svc, folder_id, idx_path, "text/html")
    print(f"  uploaded data.json + index.html")

    for fp in sorted(PROJECTS_DIR.glob("*.html")):
        _upload_or_replace(svc, proj_folder_id, fp, "text/html")
    print(f"  uploaded {len(list(PROJECTS_DIR.glob('*.html')))} project pages")

    folder_url = f"https://drive.google.com/drive/folders/{folder_id}"
    print(f"\nDrive folder: {folder_url}")
    print(f"Local index:  {idx_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
