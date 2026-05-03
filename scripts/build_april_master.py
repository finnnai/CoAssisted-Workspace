# © 2026 CoAssisted Workspace. Licensed under MIT.
"""Build a master labor view from the April Overall Reports.

Runs through every .xls in staffwizard_overall_reports/, parses with
labor_ingest.parse_overall_report (which already handles the legacy
.xls conversion via libreoffice + the 66-column StaffWizard layout),
and writes:

  staffwizard_overall_reports/master_april_2026.xlsx
      Three tabs:
        Daily Detail        — every row, every day, every employee.
        Project Rollup      — by JobNumber × Date totals (hours,
                              cost, revenue, margin).
        Daily Totals        — one row per day (shifts, hours,
                              cost, revenue, margin).

The .xlsx is the source of truth for the eventual Google Sheet
mirror — that step happens via the sheets MCP tools after this
script finishes.

Usage:
    python3 scripts/build_april_master.py
"""

from __future__ import annotations

import datetime as _dt
import pathlib
import sys
from collections import defaultdict

import openpyxl

PROJECT_ROOT = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import labor_ingest  # noqa: E402

REPORTS_DIR = PROJECT_ROOT / "staffwizard_overall_reports"
OUTPUT = REPORTS_DIR / "master_april_2026.xlsx"


def main() -> int:
    files = sorted(REPORTS_DIR.glob("overall_report_2026-04-*.xls"))
    if not files:
        print(f"No .xls files found in {REPORTS_DIR}")
        return 2

    print(f"Found {len(files)} files to process.")

    # Detail rows + per-(project, date) aggregates + per-date totals.
    detail_rows: list[dict] = []
    project_date_agg: dict[tuple[str, _dt.date], dict] = defaultdict(
        lambda: {"shifts": 0, "hours": 0.0, "cost": 0.0,
                 "revenue": 0.0, "margin": 0.0}
    )
    daily_totals: dict[_dt.date, dict] = defaultdict(
        lambda: {"shifts": 0, "hours": 0.0, "cost": 0.0,
                 "revenue": 0.0, "margin": 0.0,
                 "unique_projects": set(), "unique_employees": set()}
    )

    for f in files:
        try:
            parsed = labor_ingest.parse_overall_report(f)
        except Exception as e:
            print(f"  FAIL {f.name}: {e}")
            continue
        d = parsed.work_date
        print(f"  {f.name}: {len(parsed.rows)} shifts, "
              f"work_date={d}, skipped_no_job={parsed.skipped_no_job}")

        for r in parsed.rows:
            # Detail row.
            detail_rows.append({
                "Date": r.work_date.isoformat() if r.work_date else "",
                "Job Number": r.job_number,
                "Job Description": r.job_description,
                "Employee #": r.employee_number,
                "Employee Name": r.employee_name,
                "Post": r.post_description,
                "Shift Start": r.shift_start,
                "Shift End": r.shift_end,
                "Reg Hours": r.hours,
                "OT Hours": r.overtime_hours,
                "DT Hours": r.doubletime_hours,
                "Total Hours": r.total_hours,
                "Reg Cost $": r.dollars,
                "Holiday Cost $": r.holiday_dollars,
                "OT Cost $": r.overtime_dollars,
                "DT Cost $": r.doubletime_dollars,
                "Total Cost $": r.total_cost,
                "Billable Hours": r.billable_hours,
                "Billable $": r.billable_dollars,
                "Margin $": r.margin,
            })

            # Per-(project, date) aggregate.
            if r.work_date:
                key = (r.job_number, r.work_date)
                a = project_date_agg[key]
                a["shifts"] += 1
                a["hours"] += r.total_hours
                a["cost"] += r.total_cost
                a["revenue"] += r.billable_dollars
                a["margin"] += r.margin
                a["job_description"] = r.job_description

                # Daily totals.
                t = daily_totals[r.work_date]
                t["shifts"] += 1
                t["hours"] += r.total_hours
                t["cost"] += r.total_cost
                t["revenue"] += r.billable_dollars
                t["margin"] += r.margin
                t["unique_projects"].add(r.job_number)
                if r.employee_number:
                    t["unique_employees"].add(r.employee_number)

    print(f"\nTotal detail rows: {len(detail_rows)}")
    print(f"Unique (project, date) pairs: {len(project_date_agg)}")
    print(f"Days covered: {len(daily_totals)}")

    # ----------------------------------------------------------------------
    # Build the workbook.
    # ----------------------------------------------------------------------

    wb = openpyxl.Workbook()
    # Remove the default sheet so our three named tabs are the only ones.
    wb.remove(wb.active)

    # Tab 1: Daily Detail.
    detail = wb.create_sheet("Daily Detail")
    if detail_rows:
        cols = list(detail_rows[0].keys())
        detail.append(cols)
        for row in detail_rows:
            detail.append([row[c] for c in cols])
        # Freeze header row + autosize-ish.
        detail.freeze_panes = "A2"

    # Tab 2: Project Rollup (by JobNumber × Date).
    rollup = wb.create_sheet("Project Rollup")
    rollup.append([
        "Date", "Job Number", "Job Description",
        "Shifts", "Hours", "Cost $", "Revenue $", "Margin $",
        "Margin %",
    ])
    for (job, date), a in sorted(project_date_agg.items(), key=lambda x: (x[0][1], x[0][0])):
        margin_pct = (a["margin"] / a["revenue"] * 100) if a["revenue"] else 0
        rollup.append([
            date.isoformat(), job, a.get("job_description", ""),
            a["shifts"], round(a["hours"], 2),
            round(a["cost"], 2), round(a["revenue"], 2),
            round(a["margin"], 2), round(margin_pct, 1),
        ])
    rollup.freeze_panes = "A2"

    # Tab 3: Daily Totals.
    totals = wb.create_sheet("Daily Totals")
    totals.append([
        "Date", "Shifts", "Unique Projects", "Unique Employees",
        "Total Hours", "Total Cost $", "Total Revenue $",
        "Total Margin $", "Margin %",
    ])
    for date, t in sorted(daily_totals.items()):
        margin_pct = (t["margin"] / t["revenue"] * 100) if t["revenue"] else 0
        totals.append([
            date.isoformat(), t["shifts"],
            len(t["unique_projects"]), len(t["unique_employees"]),
            round(t["hours"], 2), round(t["cost"], 2),
            round(t["revenue"], 2), round(t["margin"], 2),
            round(margin_pct, 1),
        ])
    totals.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"\nWrote: {OUTPUT}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
