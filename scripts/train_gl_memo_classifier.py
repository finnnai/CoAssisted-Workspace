#!/usr/bin/env python3
# © 2026 CoAssisted Workspace. Licensed under MIT.
"""One-shot trainer for the GL memo classifier (Tier 2 of AP-3).

Reads the labeled journal-entry corpus from
`samples/Wolfhound Corp JEs Jan-Mar'26.xlsx`, tokenizes each Line Memo,
counts (token × GL account) co-occurrences, and writes the index to
`gl_memo_index.json` for the runtime classifier in
`gl_memo_classifier.py` to load on first lookup.

Usage:
    python3 scripts/train_gl_memo_classifier.py
        --input  samples/Wolfhound\\ Corp\\ JEs\\ Jan-Mar\\'26.xlsx
        --output gl_memo_index.json

Both flags have sensible defaults so a bare `python3 scripts/train_gl_memo_classifier.py`
works from the project root. Re-run whenever the JE corpus refreshes
(quarterly is plenty — memo patterns don't drift fast).

Index file shape:

    {
      "trained_on":      "2026-05-01T22:30:00-07:00",
      "input_path":      "samples/Wolfhound Corp JEs Jan-Mar'26.xlsx",
      "n_documents":     17346,
      "vocab_size":      4823,
      "gl_priors":       {"62300:IT Expenses": 1234, ...},
      "gl_token_totals": {"62300:IT Expenses": 9876, ...},
      "tokens_by_gl":    {
        "62300:IT Expenses": {"amazon": 245, "knack": 12, ...},
        ...
      }
    }

Memo extraction rule: prefer column 13 (Line Memo). When that's empty,
fall back to column 1 (Journal) which carries the journal-level
description (e.g., "JE-3935 - 320 Chinquapa LLC - 03/10/2026 - To
record taxes/utility prorations from Tahoe property purchase"). Strip
the JE-id and date prefixes so we train on the actual rationale text.
"""

from __future__ import annotations

import argparse
import datetime as _dt
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

# Local imports.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from gl_memo_classifier import tokenize  # reuse exact tokenization


# =============================================================================
# Constants — tuned to the Wolfhound JE Excel layout
# =============================================================================

# Sheet 1 layout (1-indexed columns from the openpyxl read):
#   Col 1  : Journal description ("JE-3935 - ... - To record ...")
#   Col 10 : Ledger Account ("63700:Taxes")
#   Col 11 : Ledger Debit Amount
#   Col 12 : Ledger Credit Amount
#   Col 13 : Line Memo ("To record taxes/utility prorations ...")
#   Col 15 : Worktags ("Cost Center: Logistics\n\nSpend Category: ...")
COL_JOURNAL = 1
COL_LEDGER_ACCOUNT = 10
COL_LINE_MEMO = 13
COL_WORKTAGS = 15

HEADER_ROW = 27   # column headers; data starts at row 28
DATA_START = 28


# Strip "JE-#### - Company - MM/DD/YYYY - " prefix from journal-level
# descriptions when we have to fall back from Line Memo.
_JE_PREFIX = re.compile(
    r"^JE-\d+\s*-\s*[^-]+-\s*\d{1,2}/\d{1,2}/\d{2,4}\s*-\s*",
    re.IGNORECASE,
)


# GL accounts are formatted "{number}:{name}" — e.g. "62300:IT Expenses".
_GL_NUMBER_PREFIX = re.compile(r"^(\d+):")


def _gl_account_number(gl: str) -> int | None:
    """Extract the leading numeric account code from a GL account string."""
    if not gl:
        return None
    m = _GL_NUMBER_PREFIX.match(gl.strip())
    if not m:
        return None
    try:
        return int(m.group(1))
    except ValueError:
        return None


def _extract_memo(journal_desc: str | None, line_memo: str | None) -> str:
    """Pick the best-available memo text and strip JE-id boilerplate."""
    if line_memo and line_memo.strip():
        return line_memo.strip()
    if journal_desc and journal_desc.strip():
        return _JE_PREFIX.sub("", journal_desc.strip()).strip()
    return ""


# =============================================================================
# Training
# =============================================================================

def train(input_path: Path, output_path: Path) -> None:
    """Read the JE Excel, build the index, write JSON."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(
            "openpyxl is required. Install with: "
            "pip install openpyxl --break-system-packages",
            file=sys.stderr,
        )
        sys.exit(1)

    if not input_path.exists():
        print(f"Input file not found: {input_path}", file=sys.stderr)
        sys.exit(2)

    print(f"Loading {input_path} ...", flush=True)
    # Note: read_only=False is intentional. The Wolfhound JE export
    # doesn't carry the row-count metadata read_only mode relies on, so
    # iter_rows returns nothing under read_only=True. The full-load cost
    # is fine — this script runs once per quarter, not per request.
    wb = load_workbook(input_path, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    print(f"  Sheet: {ws.title}, rows: {ws.max_row}, cols: {ws.max_column}",
          flush=True)

    # Counters built up row-by-row.
    gl_priors: dict[str, int] = defaultdict(int)
    gl_token_totals: dict[str, int] = defaultdict(int)
    tokens_by_gl: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    vocab: set[str] = set()
    n_documents = 0
    n_skipped_no_gl = 0
    n_skipped_no_memo = 0
    n_skipped_credit_side = 0
    n_skipped_non_expense = 0

    print(f"Scanning rows {DATA_START}..{ws.max_row} ...", flush=True)
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=DATA_START, values_only=True), start=DATA_START
    ):
        # ws.iter_rows returns tuples; pad to expected width if needed.
        if len(row) < COL_WORKTAGS:
            row = row + (None,) * (COL_WORKTAGS - len(row))

        ledger_account = row[COL_LEDGER_ACCOUNT - 1]
        line_memo = row[COL_LINE_MEMO - 1]
        journal_desc = row[COL_JOURNAL - 1]
        debit_amount = row[COL_LEDGER_ACCOUNT]      # col 11 — debit
        credit_amount = row[COL_LEDGER_ACCOUNT + 1] # col 12 — credit

        if not ledger_account or not isinstance(ledger_account, str):
            n_skipped_no_gl += 1
            continue
        gl = ledger_account.strip()

        # AP classifier only learns from DEBIT-side rows. Credit-side
        # rows post to the offsetting payable / cash / liability — those
        # are determined by the card type or bank, not by the memo, so
        # training on them would teach the model to predict the wrong
        # side of the journal. Skip them.
        debit_val = float(debit_amount or 0)
        credit_val = float(credit_amount or 0)
        if debit_val <= 0 and credit_val > 0:
            n_skipped_credit_side += 1
            continue

        # Further filter to EXPENSE accounts only (50000-69999 per the
        # Surefox chart of accounts: 50000s = COS, 60000s = SG&A). AP
        # card spend always lands in this range — debit-side cash, AR,
        # fixed-asset, and payable-settlement rows would teach the model
        # the wrong patterns. Parse the GL prefix.
        gl_num = _gl_account_number(gl)
        if gl_num is None or not (50000 <= gl_num <= 69999):
            n_skipped_non_expense += 1
            continue

        memo = _extract_memo(journal_desc, line_memo)
        if not memo:
            n_skipped_no_memo += 1
            continue

        tokens = tokenize(memo)
        if not tokens:
            n_skipped_no_memo += 1
            continue

        gl_priors[gl] += 1
        gl_token_totals[gl] += len(tokens)
        for tok in tokens:
            tokens_by_gl[gl][tok] += 1
            vocab.add(tok)
        n_documents += 1

        if n_documents % 2000 == 0:
            print(f"  ... {n_documents} rows ingested", flush=True)

    if n_documents == 0:
        print("No usable rows found — index would be empty. Aborting.", file=sys.stderr)
        sys.exit(3)

    # Convert defaultdict-of-defaultdicts to plain dicts for JSON.
    tokens_by_gl_plain = {
        gl: dict(counts) for gl, counts in tokens_by_gl.items()
    }

    index = {
        "trained_on": _dt.datetime.now().astimezone().isoformat(timespec="seconds"),
        "input_path": str(input_path),
        "n_documents": n_documents,
        "vocab_size": len(vocab),
        "gl_priors": dict(gl_priors),
        "gl_token_totals": dict(gl_token_totals),
        "tokens_by_gl": tokens_by_gl_plain,
    }

    print(
        f"\nTraining done.\n"
        f"  Documents ingested:        {n_documents}\n"
        f"  GL accounts seen:          {len(gl_priors)}\n"
        f"  Vocabulary size:           {len(vocab)}\n"
        f"  Skipped (no GL):           {n_skipped_no_gl}\n"
        f"  Skipped (no memo):         {n_skipped_no_memo}\n"
        f"  Skipped (credit-side row): {n_skipped_credit_side}\n"
        f"  Skipped (non-expense GL):  {n_skipped_non_expense}\n",
        flush=True,
    )

    print(f"Writing index → {output_path}", flush=True)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(index, f, indent=2, ensure_ascii=False, sort_keys=True)
    size_kb = output_path.stat().st_size / 1024
    print(f"  index size: {size_kb:.1f} KB", flush=True)

    # Top-5 GL by training rows — quick sanity check.
    print("\nTop 5 GL accounts by training-row count:")
    top = sorted(gl_priors.items(), key=lambda kv: kv[1], reverse=True)[:5]
    for gl, count in top:
        print(f"  {count:5d}  {gl}")


# =============================================================================
# CLI
# =============================================================================

def main() -> None:
    project_root = Path(__file__).resolve().parent.parent
    default_input = (
        project_root / "samples" / "Wolfhound Corp JEs Jan-Mar'26.xlsx"
    )
    default_output = project_root / "gl_memo_index.json"

    parser = argparse.ArgumentParser(
        description="Train the GL memo classifier on labeled JE data.",
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=default_input,
        help=f"Path to the JE Excel (default: {default_input.name}).",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=default_output,
        help=f"Where to write the index (default: {default_output.name}).",
    )
    args = parser.parse_args()
    train(args.input, args.output)


if __name__ == "__main__":
    main()
