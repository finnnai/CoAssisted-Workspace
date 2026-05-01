# © 2026 CoAssisted Workspace. Licensed under MIT.
"""Persistent cardholder/department → Cost Center mapping store.

Used by AP-2's Workday Journal EIB writer to attach the right Cost Center
worktag to every journal line. Without this map, every line falls back
to the system default (CC100), which is fine for go-live but makes
post-period analytics noisier than they should be.

Storage:
    ~/Developer/google_workspace_mcp/cost_center_map.json (gitignored)

Two key types — both store-relative, looked up in this order:

    1. cardholder_email   "michael.vetre@xenture.com" → "CC200"
    2. department          "OXBLOOD"                    → "CC400"

When AP-2 fires `lookup(cardholder_email=..., department=...)`, we check
email first (more specific), then department, then None (the caller falls
back to its `default_cost_center`).

Mirrors the architectural pattern of gl_merchant_map.py:
    - atomic JSON writes (tempfile + os.replace)
    - operator > import > training source precedence
    - history events capped at 5
    - hit_count + last_seen advance on lookup
"""

from __future__ import annotations

import datetime as _dt
import json
import os
import tempfile
from pathlib import Path
from typing import Optional


# =============================================================================
# Storage paths + constants
# =============================================================================

_PROJECT_ROOT = Path(__file__).resolve().parent
_MAP_PATH = _PROJECT_ROOT / "cost_center_map.json"

_SOURCE_PRECEDENCE = {
    "operator": 3,   # human approval / explicit set
    "import":   2,   # bulk import (CSV upload, etc.)
    "training": 1,   # auto-derived from JE history
}

# Two distinct key namespaces in one store.
_KIND_EMAIL = "email"
_KIND_DEPT = "dept"


# =============================================================================
# Internals
# =============================================================================

def _key(kind: str, value: str) -> str:
    return f"{kind}::{value.strip().lower()}"


def _now_iso() -> str:
    return _dt.datetime.now().astimezone().isoformat(timespec="seconds")


def _load() -> dict[str, dict]:
    if not _MAP_PATH.exists():
        return {}
    try:
        with _MAP_PATH.open("r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (json.JSONDecodeError, OSError):
        return {}


def _save(data: dict[str, dict]) -> None:
    _MAP_PATH.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(
        prefix="cost_center_map.", suffix=".json.tmp",
        dir=str(_MAP_PATH.parent),
    )
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False, sort_keys=True)
        os.replace(tmp_path, _MAP_PATH)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


# =============================================================================
# Public API
# =============================================================================

def lookup(
    *,
    cardholder_email: Optional[str] = None,
    department: Optional[str] = None,
) -> Optional[str]:
    """Return the cost-center for a cardholder or department.

    Lookup precedence: cardholder_email > department > None.
    On a hit, hit_count + last_seen advance. The first hit by either
    key wins — we don't merge per-email and per-dept entries.
    """
    data = _load()

    if cardholder_email:
        k = _key(_KIND_EMAIL, cardholder_email)
        if k in data:
            data[k]["hit_count"] = int(data[k].get("hit_count", 0)) + 1
            data[k]["last_seen"] = _now_iso()
            _save(data)
            return data[k].get("cost_center") or None

    if department:
        k = _key(_KIND_DEPT, department)
        if k in data:
            data[k]["hit_count"] = int(data[k].get("hit_count", 0)) + 1
            data[k]["last_seen"] = _now_iso()
            _save(data)
            return data[k].get("cost_center") or None

    return None


def set_for_email(
    cardholder_email: str,
    cost_center: str,
    *,
    source: str = "operator",
    user: Optional[str] = None,
) -> None:
    """Map a cardholder email to a cost center."""
    _set(_KIND_EMAIL, cardholder_email, cost_center, source=source, user=user)


def set_for_department(
    department: str,
    cost_center: str,
    *,
    source: str = "operator",
    user: Optional[str] = None,
) -> None:
    """Map a department label (OXBLOOD / GREEN FLEET / etc.) to a cost center."""
    _set(_KIND_DEPT, department, cost_center, source=source, user=user)


def _set(
    kind: str,
    value: str,
    cost_center: str,
    *,
    source: str,
    user: Optional[str],
) -> None:
    if not value or not cost_center:
        return
    if source not in _SOURCE_PRECEDENCE:
        raise ValueError(
            f"Unknown source {source!r}; expected one of "
            f"{sorted(_SOURCE_PRECEDENCE)}"
        )
    k = _key(kind, value)
    data = _load()
    now = _now_iso()
    existing = data.get(k)

    # Precedence guard.
    if existing:
        existing_rank = _SOURCE_PRECEDENCE.get(
            existing.get("source", "training"), 0
        )
        if _SOURCE_PRECEDENCE[source] < existing_rank:
            return  # lower precedence; don't clobber

    event: dict = {"source": source, "iso": now}
    if user:
        event["user"] = user

    if existing:
        history = list(existing.get("history") or [])
        history.append(event)
        history = history[-5:]
        record = {
            **existing,
            "cost_center": cost_center,
            "source": source,
            "last_seen": now,
            "history": history,
        }
    else:
        record = {
            "kind": kind,
            "value": value,
            "cost_center": cost_center,
            "source": source,
            "first_seen": now,
            "last_seen": now,
            "hit_count": 0,
            "history": [event],
        }
    data[k] = record
    _save(data)


def forget(
    *,
    cardholder_email: Optional[str] = None,
    department: Optional[str] = None,
) -> bool:
    """Drop one entry by email or department. Returns True if removed."""
    if cardholder_email:
        k = _key(_KIND_EMAIL, cardholder_email)
    elif department:
        k = _key(_KIND_DEPT, department)
    else:
        return False
    data = _load()
    if k not in data:
        return False
    del data[k]
    _save(data)
    return True


def list_all(*, kind: Optional[str] = None) -> list[dict]:
    """Inventory of all entries, optionally filtered to email or dept only."""
    data = _load()
    out = []
    for record in data.values():
        if kind and record.get("kind") != kind:
            continue
        out.append(dict(record))
    out.sort(key=lambda r: int(r.get("hit_count", 0)), reverse=True)
    return out


def export_lookup_dict() -> dict[str, str]:
    """Flatten the store to a {email_or_dept: cost_center} dict.

    AP-2's `build_journal_eib(cardholder_cost_center_map=...)` accepts
    exactly this shape. Email keys win when both an email and a dept
    map to the same cardholder; the AP-2 writer also tries dept as a
    fallback, so dept entries are still useful when email is unknown.
    """
    flat: dict[str, str] = {}
    for record in _load().values():
        cc = record.get("cost_center")
        value = record.get("value")
        if cc and value:
            flat[value] = cc
    return flat


def stats() -> dict:
    data = _load()
    by_kind: dict[str, int] = {}
    by_source: dict[str, int] = {}
    for record in data.values():
        by_kind[record.get("kind", "unknown")] = (
            by_kind.get(record.get("kind", "unknown"), 0) + 1
        )
        by_source[record.get("source", "unknown")] = (
            by_source.get(record.get("source", "unknown"), 0) + 1
        )
    return {
        "total_entries": len(data),
        "by_kind": by_kind,
        "by_source": by_source,
    }


def clear() -> int:
    data = _load()
    n = len(data)
    if n:
        _save({})
    return n


# =============================================================================
# Auto-derivation from JE training data
# =============================================================================

def derive_from_je_corpus(je_xlsx_path: str) -> dict[str, str]:
    """Scan the Wolfhound JE corpus to suggest department → CC mappings.

    Parses the Worktags column for `Cost Center: X` markers and groups
    by department-like signals in the line memo. Returns a draft map
    that the operator should review via the workflow_cost_center_map_init
    flow before committing as `source="operator"`.

    Returns {department_or_email: cost_center} as a STRAWMAN — does not
    write to the persistent store. Caller decides which entries to keep.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}

    p = Path(je_xlsx_path)
    if not p.exists():
        return {}

    wb = load_workbook(p, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]

    # Worktags are in col 15 (per train_gl_memo_classifier.py layout).
    # We extract the "Cost Center: X" marker per row.
    cc_observations: dict[str, dict[str, int]] = {}  # department → {cc → count}
    DATA_START = 28
    COL_WORKTAGS = 15
    COL_LINE_MEMO = 13

    import re
    cc_re = re.compile(r"Cost Center:\s*([^\n]+)", re.IGNORECASE)

    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        if len(row) < COL_WORKTAGS:
            continue
        worktags = row[COL_WORKTAGS - 1]
        memo = row[COL_LINE_MEMO - 1] or ""
        if not worktags or not isinstance(worktags, str):
            continue
        m = cc_re.search(worktags)
        if not m:
            continue
        cc = m.group(1).strip()
        # Department-like signals show up in memos as "ISOC", "OXBLOOD",
        # "GREEN FLEET" etc. We don't try to parse them out here — the
        # AP-2 caller passes the WEX Department field directly to
        # set_for_department, which is more reliable than memo extraction.
        # This function returns a CC-distribution dict per memo for the
        # operator to eyeball.
        memo_lower = (memo or "").lower()
        for dept_token in ("isoc", "oxblood", "green fleet", "yellow fleet", "violet fleet"):
            if dept_token in memo_lower:
                bucket = cc_observations.setdefault(dept_token.upper(), {})
                bucket[cc] = bucket.get(cc, 0) + 1

    # Pick the modal CC for each department.
    suggestions: dict[str, str] = {}
    for dept, ccs in cc_observations.items():
        if not ccs:
            continue
        winner = max(ccs.items(), key=lambda kv: kv[1])
        suggestions[dept] = winner[0]
    return suggestions
