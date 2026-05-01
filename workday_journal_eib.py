# © 2026 CoAssisted Workspace. Licensed under MIT.
"""AP-2: Card statement → Workday Accounting Journal EIB.

Parses AMEX and WEX card statements, classifies each transaction's spend
GL via gl_classifier (Tier 0-3 ladder), and emits a two-sheet Workday
EIB matching the `Import Accounting Journal` + `Journal Entry Line
Replacement` template that ops uploads on close.

Each transaction expands to two journal lines:
    DEBIT  → spend GL (from classifier), Cost Center (from cardholder map)
    CREDIT → card payable (22000 AMEX / 22030 WEX), same Cost Center

Refunds (negative-amount transactions) reverse the dr/cr direction —
the merchant gets credited (refund posts as a credit to spend) and
the card payable gets debited (reducing what we owe).

Anchor reference: samples/SFNA AMEX EIB MARCH 26.xlsx is the exact
shape this module reproduces. Ops loads the output via Workday's
Submit Accounting Journals task.

Public surface:
    parse_amex_csv(path)  -> list[CardTransaction]
    parse_wex_csv(path)   -> list[CardTransaction]
    build_journal_eib(transactions, output_path, ...) -> EIBResult
"""

from __future__ import annotations

import csv
import datetime as _dt
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional


# =============================================================================
# Data model
# =============================================================================

@dataclass
class CardTransaction:
    """Normalized representation of a card-statement line."""

    # --- Identification ---
    card_type: str                     # "amex" | "wex"
    card_last_4: str
    transaction_date: _dt.date
    merchant_name: str

    # --- Amount (positive = debit / charge; negative = credit / refund) ---
    amount: float

    # --- Cardholder / cost center attribution ---
    cardholder_name: str = ""          # "Last, First" preferred
    cardholder_email: Optional[str] = None

    # --- AMEX-specific ---
    mcc_code: Optional[int] = None
    mcc_description: Optional[str] = None
    status: Optional[str] = None       # CLEARED / PENDING / DECLINED

    # --- WEX-specific ---
    department: Optional[str] = None   # OXBLOOD, GREEN FLEET, ISOC, etc.
    vehicle_id: Optional[str] = None
    driver_name: Optional[str] = None
    fuel_product: Optional[str] = None # UNL, SUP, WASH

    # --- Reconciliation hooks (set by AP-2 caller, not parsers) ---
    receipt_path: Optional[Path] = None
    receipt_matched: bool = False

    # --- Free-form ---
    raw_notes: str = ""                # AMEX 'Notes' column or similar


@dataclass
class EIBResult:
    """Returned by build_journal_eib so callers know what landed."""

    output_path: Path
    n_transactions: int
    n_lines_written: int               # always 2 × n_transactions
    n_orphan_no_gl: int                # txns where every classifier tier missed
    classifier_tier_counts: dict[str, int] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)


# =============================================================================
# Workday card-payable accounts (per the COA)
# =============================================================================

_CARD_PAYABLE_ACCOUNT: dict[str, str] = {
    "amex":  "22000:Credit Card Payable- AMEX",
    "wex":   "22030:Credit Card Payable- WEX",
    # Future: "chase": "22010:Credit Card Payable- Chase", etc.
}


# =============================================================================
# AMEX parser
# =============================================================================

# Column names in the AMEX CSV header (samples/Amex Transactions - April.csv).
# Tied to the export Surefox uses today; if AMEX changes the export shape this
# is where it'll show up.
_AMEX_COL_LAST4 = "Registered Card Last 4"
_AMEX_COL_TXN_DATE = "Transaction Date"
_AMEX_COL_STATUS = "Status"
_AMEX_COL_NAME = "Name On Card"
_AMEX_COL_EMAIL = "Card User Email"
_AMEX_COL_AMOUNT = "Authorization Billing Amount"
_AMEX_COL_MERCHANT = "Merchant Name"
_AMEX_COL_MCC = "MCC"
_AMEX_COL_MCC_DESC = "MCC Description"
_AMEX_COL_NOTES = "Notes"


def _parse_amex_date(raw: str) -> Optional[_dt.date]:
    """AMEX exports use M/D/YY format (e.g. '4/30/26')."""
    if not raw:
        return None
    raw = raw.strip()
    for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return _dt.datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _parse_float(raw) -> float:
    """Tolerant float parse — handles strings, blanks, and accounting parens."""
    if raw is None or raw == "":
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().replace(",", "").replace("$", "")
    # Accounting-style negative: "(123.45)"
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return 0.0


def _normalize_cardholder_name(raw: str) -> str:
    """Convert 'First Last' → 'Last, First' (Workday memo convention)."""
    if not raw:
        return ""
    parts = raw.strip().split()
    if len(parts) >= 2:
        first = " ".join(parts[:-1])
        last = parts[-1]
        return f"{last}, {first}"
    return raw.strip()


def parse_amex_csv(
    path: Path,
    *,
    include_pending: bool = False,
) -> list[CardTransaction]:
    """Parse an AMEX export CSV into normalized CardTransaction records.

    Args:
        path: Path to the CSV.
        include_pending: When False (default), only CLEARED transactions
            are emitted — PENDING and DECLINED are skipped because they
            don't belong on a posted journal. Flip this to True to ingest
            mid-cycle for cash-forecasting.
    """
    out: list[CardTransaction] = []
    with Path(path).open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            status = (row.get(_AMEX_COL_STATUS) or "").strip().upper()
            if not include_pending and status != "CLEARED":
                continue
            txn_date = _parse_amex_date(row.get(_AMEX_COL_TXN_DATE) or "")
            if not txn_date:
                # Without a date the line can't anchor to a posting period.
                continue
            amount = _parse_float(row.get(_AMEX_COL_AMOUNT))
            if amount == 0.0:
                continue
            try:
                mcc_int = int(row.get(_AMEX_COL_MCC) or "0") or None
            except ValueError:
                mcc_int = None
            out.append(CardTransaction(
                card_type="amex",
                card_last_4=(row.get(_AMEX_COL_LAST4) or "").lstrip("*").strip(),
                transaction_date=txn_date,
                cardholder_name=_normalize_cardholder_name(
                    row.get(_AMEX_COL_NAME) or ""
                ),
                cardholder_email=(row.get(_AMEX_COL_EMAIL) or "").strip() or None,
                amount=amount,
                merchant_name=(row.get(_AMEX_COL_MERCHANT) or "").strip(),
                mcc_code=mcc_int,
                mcc_description=(row.get(_AMEX_COL_MCC_DESC) or "").strip() or None,
                status=status,
                raw_notes=(row.get(_AMEX_COL_NOTES) or "").strip(),
            ))
    return out


# =============================================================================
# WEX parser
# =============================================================================

_WEX_COL_DATE = "Transaction Date"
_WEX_COL_CARD = "Card Number"
_WEX_COL_AMOUNT = "Net Cost"
_WEX_COL_MERCHANT_BRAND = "Merchant (Brand)"
_WEX_COL_MERCHANT_NAME = "Merchant Name"
_WEX_COL_DRIVER_LAST = "Driver Last Name"
_WEX_COL_DRIVER_FIRST = "Driver First Name"
_WEX_COL_DEPT = "Department"
_WEX_COL_VEHICLE = "Vehicle Description"
_WEX_COL_PRODUCT = "Product"


def parse_wex_csv(path: Path) -> list[CardTransaction]:
    """Parse a WEX Fuel export CSV into CardTransaction records.

    WEX has no MCC; the AP-3 classifier falls through to tier 2
    (JE-trained memo matcher) on these. Memo gets built from
    {Merchant Brand} + {Product Description} + {Vehicle ID} so the
    matcher has signal to work with.
    """
    out: list[CardTransaction] = []
    with Path(path).open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            txn_date = _parse_amex_date(row.get(_WEX_COL_DATE) or "")
            if not txn_date:
                continue
            amount = _parse_float(row.get(_WEX_COL_AMOUNT))
            if amount == 0.0:
                continue
            driver_last = (row.get(_WEX_COL_DRIVER_LAST) or "").strip()
            driver_first = (row.get(_WEX_COL_DRIVER_FIRST) or "").strip()
            cardholder = (
                f"{driver_last}, {driver_first}".strip(", ")
                if driver_last or driver_first
                else ""
            )
            merchant_brand = (row.get(_WEX_COL_MERCHANT_BRAND) or "").strip()
            merchant_name = (row.get(_WEX_COL_MERCHANT_NAME) or "").strip()
            merchant = merchant_brand or merchant_name
            out.append(CardTransaction(
                card_type="wex",
                card_last_4=(row.get(_WEX_COL_CARD) or "").lstrip("*").strip(),
                transaction_date=txn_date,
                cardholder_name=cardholder,
                amount=amount,
                merchant_name=merchant,
                department=(row.get(_WEX_COL_DEPT) or "").strip() or None,
                vehicle_id=(row.get(_WEX_COL_VEHICLE) or "").strip() or None,
                driver_name=cardholder or None,
                fuel_product=(row.get(_WEX_COL_PRODUCT) or "").strip() or None,
                # Composite memo so tier-2 matcher gets useful tokens.
                raw_notes=(
                    f"{merchant} {row.get(_WEX_COL_PRODUCT) or ''} "
                    f"{row.get(_WEX_COL_VEHICLE) or ''}"
                ).strip(),
            ))
    return out


# =============================================================================
# EIB writer
# =============================================================================

# Sheet 1 layout — Import Accounting Journal (header).
# Column index → field name. Empty entries are reserved for Workday.
_HEADER_COLUMNS: list[tuple[int, str]] = [
    (1,  "Fields"),  # column-A label cell, holds nothing per row
    (2,  "Header Key"),
    (3,  "Add Only"),
    (4,  "Create Journal with Errors"),
    (5,  "Accounting Journal"),
    (6,  "Auto Complete"),
    (7,  "Comment"),
    (8,  "Worker"),
    (9,  "Accounting Journal ID"),
    (10, "Submit"),
    (11, "Disable Optional Worktag Balancing"),
    (12, "Locked in Workday"),
    (13, "Round Ledger Amounts"),
    (14, "Journal for All Ledgers"),
    (15, "Journal Number"),
    (16, "Company*"),
    (17, "Currency*"),
    (18, "Ledger Type*"),
    (19, "Book Code"),
    (20, "Accounting Date*"),
    (21, "Journal Source*"),
    (22, "Balancing Worktag"),
    (23, "Optional Balancing Worktags+"),
    (24, "External Supplier Invoice Source"),
    (25, "Record Quantity"),
]

# Sheet 2 layout — Journal Entry Line Replacement.
_LINE_COLUMNS: list[tuple[int, str]] = [
    (1,  "Fields"),
    (2,  "Header Key"),
    (3,  "Line Key"),
    (4,  "Line Order"),
    (5,  "Line Company"),
    (6,  "Ledger Account"),
    (7,  "Account Set"),
    (8,  "Alternate Ledger Account"),
    (9,  "Account Set"),
    (10, "Debit Amount"),
    (11, "Credit Amount"),
    (12, "Currency"),
    (13, "Currency Rate"),
    (14, "Ledger Debit Amount"),
    (15, "Ledger Credit Amount"),
    (16, "Quantity"),
    (17, "Unit of Measure"),
    (18, "Quantity 2"),
    (19, "Unit of Measure 2"),
    (20, "Memo"),
    (21, "External Reference ID"),
    (22, "Budget Date"),
    (23, "Cost Center"),
    (24, "External Supplier Invoice Source"),
    (25, "Billable"),
]


def _format_header_key(card_type: str, period_end: _dt.date) -> str:
    """Mirror the SFNA AMEX EIB convention: '150326U' = period end + suffix.

    Pattern observed in samples/SFNA AMEX EIB MARCH 26.xlsx is `DDMMYY` + 'U'.
    We retain that exact shape so existing Workday integrations don't have
    to change. Card-type letter is prepended for AMEX vs WEX disambiguation
    so two periods can load same-day without colliding header keys.
    """
    suffix = {"amex": "AU", "wex": "WU"}.get(card_type, "U")
    return f"{period_end:%d%m%y}{suffix}"


def _format_line_memo(
    txn: CardTransaction, period_start: _dt.date, period_end: _dt.date
) -> str:
    """Match the AMEX sample format precisely so review reads identically.

    Example: 'AMEX Transactions 03.01.26-03.31.26 - Szott, Joshua - PIRATE SHIP'
    """
    label = {"amex": "AMEX", "wex": "WEX"}.get(txn.card_type, txn.card_type.upper())
    return (
        f"{label} Transactions "
        f"{period_start:%m.%d.%y}-{period_end:%m.%d.%y} - "
        f"{txn.cardholder_name or 'Unknown'} - "
        f"{txn.merchant_name or 'Unknown Merchant'}"
    )


def build_journal_eib(
    transactions: list[CardTransaction],
    output_path: Path,
    *,
    card_type: str,
    period_start: _dt.date,
    period_end: _dt.date,
    cardholder_cost_center_map: Optional[dict[str, str]] = None,
    company_id: str = "CO-100",
    currency: str = "USD",
    ledger_type: str = "ACTUALS",
    journal_source: str = "Spreadsheet_Upload",
    default_cost_center: str = "CC100",
) -> EIBResult:
    """Write a two-sheet Workday Journal EIB from card transactions.

    Each transaction emits two journal lines (debit + credit pair) linked
    by Header Key. The DEBIT line posts to the spend GL (from gl_classifier);
    the CREDIT line posts to the card payable.

    Args:
        transactions: parsed card-statement records.
        output_path: where to write the .xlsx.
        card_type: "amex" or "wex".
        period_start, period_end: statement period bounds (inclusive).
        cardholder_cost_center_map: optional {email: cost_center_id} map.
            Falls back to default_cost_center when a cardholder isn't in
            the map.
        company_id, currency, ledger_type, journal_source: Workday EIB
            header values; defaults match Surefox's typical Wolfhound load.
        default_cost_center: used when no cardholder map is provided OR
            the cardholder isn't in the map.

    Returns:
        EIBResult summarizing what was written.
    """
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to build Workday EIBs. "
            "Install with: pip install openpyxl --break-system-packages"
        ) from exc

    if card_type not in _CARD_PAYABLE_ACCOUNT:
        raise ValueError(
            f"Unknown card_type {card_type!r}; expected one of "
            f"{sorted(_CARD_PAYABLE_ACCOUNT)}"
        )

    # Lazy import to keep test runs that don't exercise classification fast.
    import gl_classifier

    cc_map = cardholder_cost_center_map or {}
    payable_account = _CARD_PAYABLE_ACCOUNT[card_type]
    header_key = _format_header_key(card_type, period_end)

    wb = Workbook()
    # Sheet 1 — Import Accounting Journal (header)
    ws_h = wb.active
    ws_h.title = "Import Accounting Journal"
    _write_header_sheet(
        ws_h,
        header_key=header_key,
        company_id=company_id,
        currency=currency,
        ledger_type=ledger_type,
        accounting_date=period_end,
        journal_source=journal_source,
    )

    # Sheet 2 — Journal Entry Line Replacement (lines)
    ws_l = wb.create_sheet(title="Journal Entry Line Replacement")
    _write_line_sheet_header(ws_l)

    # Counters for the result.
    n_orphan = 0
    tier_counts: dict[str, int] = {}

    line_key = 1
    line_row = 6  # data starts at row 6 (rows 1-5 are the EIB header band)
    for txn in transactions:
        memo = _format_line_memo(txn, period_start, period_end)
        cost_center = (
            cc_map.get(txn.cardholder_email or "")
            or cc_map.get(txn.department or "")
            or default_cost_center
        )

        # Classify the spend GL via the AP-3 ladder.
        result = gl_classifier.classify_transaction(
            merchant_name=txn.merchant_name,
            mcc_code=txn.mcc_code,
            memo=f"{memo} {txn.raw_notes}".strip(),
            amount=abs(txn.amount),
            cardholder_email=txn.cardholder_email,
            department_hint=txn.department,
        )
        spend_gl = result.gl_account
        tier_key = result.tier_used.value
        tier_counts[tier_key] = tier_counts.get(tier_key, 0) + 1
        if spend_gl == "22040:Credit Card Clearing":
            n_orphan += 1

        is_refund = txn.amount < 0
        amt = abs(txn.amount)

        # DEBIT row (or CREDIT-then-DEBIT for refunds — Workday journal
        # entry replacement still expects two rows per linked event).
        if not is_refund:
            debit_gl, debit_amt = spend_gl, amt
            credit_gl, credit_amt = payable_account, amt
        else:
            debit_gl, debit_amt = payable_account, amt
            credit_gl, credit_amt = spend_gl, amt

        # Write the two rows.
        _write_journal_line(
            ws_l, line_row, header_key, line_key,
            ledger_account=debit_gl,
            debit=debit_amt, credit=0,
            currency=currency, memo=memo, cost_center=cost_center,
        )
        line_row += 1
        line_key += 1

        _write_journal_line(
            ws_l, line_row, header_key, line_key,
            ledger_account=credit_gl,
            debit=0, credit=credit_amt,
            currency=currency, memo=memo, cost_center=cost_center,
        )
        line_row += 1
        line_key += 1

    # Save.
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    return EIBResult(
        output_path=output_path,
        n_transactions=len(transactions),
        n_lines_written=(line_key - 1),
        n_orphan_no_gl=n_orphan,
        classifier_tier_counts=tier_counts,
        warnings=(
            [f"{n_orphan} transactions held in 22040:Credit Card Clearing — "
             f"operator review required."] if n_orphan else []
        ),
    )


# =============================================================================
# Sheet writers
# =============================================================================

def _write_header_sheet(
    ws,
    *,
    header_key: str,
    company_id: str,
    currency: str,
    ledger_type: str,
    accounting_date: _dt.date,
    journal_source: str,
) -> None:
    """Populate Sheet 1 with the 5-row EIB header band + one data row."""
    # Row 1: title
    ws.cell(row=1, column=1, value="Import Accounting Journal")
    # Row 2: 'Area' / sections (mirror of the Workday template)
    ws.cell(row=2, column=1, value="Area")
    ws.cell(row=2, column=2, value="All")
    ws.cell(row=2, column=7, value="Comment Data")
    ws.cell(row=2, column=9, value="Accounting Journal Data")
    # Row 3: Required/Optional restrictions
    ws.cell(row=3, column=1, value="Restrictions")
    for col in (2,):
        ws.cell(row=3, column=col, value="Required")
    for col in (16, 17, 18, 20, 21):
        ws.cell(row=3, column=col, value="Required")
    # Row 4: Format hints
    ws.cell(row=4, column=1, value="Format")
    ws.cell(row=4, column=2, value="Text")
    ws.cell(row=4, column=16, value="Company_Reference_ID")
    ws.cell(row=4, column=17, value="Currency_ID")
    ws.cell(row=4, column=18, value="Ledger_Type_ID")
    ws.cell(row=4, column=20, value="YYYY-MM-DD")
    ws.cell(row=4, column=21, value="Journal_Source_ID")
    # Row 5: Field names
    for col, name in _HEADER_COLUMNS:
        ws.cell(row=5, column=col, value=name)
    # Row 6: data
    ws.cell(row=6, column=2, value=header_key)
    ws.cell(row=6, column=6, value="Y")          # Auto Complete
    ws.cell(row=6, column=9, value=header_key)   # Accounting Journal ID
    ws.cell(row=6, column=10, value="Y")         # Submit
    ws.cell(row=6, column=16, value=company_id)
    ws.cell(row=6, column=17, value=currency)
    ws.cell(row=6, column=18, value=ledger_type)
    ws.cell(row=6, column=20, value=accounting_date.isoformat())
    ws.cell(row=6, column=21, value=journal_source)


def _write_line_sheet_header(ws) -> None:
    """Populate the 5-row band on Sheet 2."""
    ws.cell(row=1, column=1, value="Journal Entry Line Replacement Data")
    ws.cell(row=2, column=1, value="Area")
    ws.cell(row=2, column=2, value="All")
    ws.cell(row=3, column=1, value="Restrictions")
    ws.cell(row=3, column=2, value="Required")
    ws.cell(row=3, column=3, value="Required")
    ws.cell(row=4, column=1, value="Format")
    ws.cell(row=4, column=2, value="Text")
    ws.cell(row=4, column=3, value="Text")
    ws.cell(row=4, column=10, value="Number (26,6)")
    ws.cell(row=4, column=11, value="Number (26,6)")
    ws.cell(row=4, column=12, value="Currency_ID")
    for col, name in _LINE_COLUMNS:
        ws.cell(row=5, column=col, value=name)


def _write_journal_line(
    ws,
    row: int,
    header_key: str,
    line_key: int,
    *,
    ledger_account: str,
    debit: float,
    credit: float,
    currency: str,
    memo: str,
    cost_center: str,
) -> None:
    """Write one Journal Entry Line Replacement row."""
    ws.cell(row=row, column=2, value=header_key)
    ws.cell(row=row, column=3, value=line_key)
    ws.cell(row=row, column=6, value=ledger_account)
    ws.cell(row=row, column=7, value="Master_Child")
    if debit > 0:
        ws.cell(row=row, column=10, value=round(debit, 2))
    if credit > 0:
        ws.cell(row=row, column=11, value=round(credit, 2))
    ws.cell(row=row, column=12, value=currency)
    ws.cell(row=row, column=20, value=memo)
    ws.cell(row=row, column=23, value=cost_center)
