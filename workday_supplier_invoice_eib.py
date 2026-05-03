# © 2026 CoAssisted Workspace. Licensed under MIT.
# See LICENSE file for terms.
"""Workday Supplier Invoice EIB writer (Submit_Supplier_Invoice_v39.1).

AP-1. Generates an upload-ready workbook in the v39.1 shape from a list of
canonical invoice rows. Pure-Python — no MCP, no Google API, no Drive.

The MCP tool wrapper in `tools/ap_wave1.py` pulls rows from the per-project
AP Sheets and hands them here.

Workbook shape (v39.1)
----------------------
The full template is 204 columns wide across several sections. We populate
only what's required for a clean upload:

    Header (per-invoice)
      - Spreadsheet Key                — deterministic from invoice number
      - Submit                         — 1 to auto-submit, 0 to leave as draft
      - Company                        — from project mapping
      - Supplier                       — vendor record reference (Name)
      - Currency                       — invoice currency
      - Invoice Date                   — ISO date
      - Invoice Received Date          — ISO date
      - Supplier's Invoice Number      — vendor's number
      - Control Total Amount           — sum of line amounts
      - Memo                           — free text

    Invoice Line Replacement (per-line)
      - Spreadsheet Key                — same as header
      - Line Order                     — 1, 2, 3...
      - Item Description               — line memo
      - Spend Category                 — from gl_spend_category_map
      - Quantity                       — 1
      - Unit Cost                      — line amount
      - Extended Amount                — line amount (qty × unit cost)
      - Ledger Account                 — from gl_classifier
      - Cost Center                    — from cost_center_map
      - Memo                           — line memo

Out-of-scope worktags (withholding tax, retention release, prepaid
amortization) stay blank — Workday accepts that for a first cut.

Public surface
--------------
    InvoiceLine, Invoice                — typed dicts
    build_workbook(invoices, *, output_path) -> dict
    classify_and_build(invoices, *, output_path, allow_ambiguous=False) -> dict
        Same as build_workbook but resolves Spend Category and Ledger
        Account on the fly via gl_classifier + gl_spend_category_map.

Either function returns a summary:
    {status, output_path, header_rows, line_rows, total_amount, parked,
     ambiguous_lines, missing_classification}

Anything that can't ship (no GL classification, ambiguous Spend Category
without override) goes into the parked list with the reason. The caller
surfaces those for operator action before re-running.
"""

from __future__ import annotations

import datetime as _dt
import hashlib
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Optional


# Header-section column order for sheet 1 ("Submit Supplier Invoice").
HEADER_COLUMNS = [
    "Spreadsheet Key",
    "Submit",
    "Company",
    "Supplier",
    "Currency",
    "Invoice Date",
    "Invoice Received Date",
    "Supplier's Invoice Number",
    "Control Total Amount",
    "Memo",
]

# Line-section column order for sheet 2 ("Invoice Line Replacement Data").
LINE_COLUMNS = [
    "Spreadsheet Key",
    "Line Order",
    "Item Description",
    "Spend Category",
    "Quantity",
    "Unit Cost",
    "Extended Amount",
    "Ledger Account",
    "Cost Center",
    "Memo",
]

# Default Workday company code. Overridable per-invoice; AP-1 v0.9.0 ships
# with this single-tenant default to match the Wave 1 scope.
DEFAULT_COMPANY = "CO-100"


# --------------------------------------------------------------------------- #
# Models
# --------------------------------------------------------------------------- #


@dataclass
class InvoiceLine:
    amount: float
    memo: str = ""
    gl_account: Optional[str] = None         # falls back to classifier
    spend_category: Optional[str] = None     # falls back to spend-cat map
    cost_center: Optional[str] = None
    quantity: float = 1.0


@dataclass
class Invoice:
    invoice_number: str
    vendor: str
    invoice_date: str                        # ISO date
    due_date: Optional[str] = None
    received_date: Optional[str] = None      # falls back to invoice_date
    project_code: Optional[str] = None
    company: str = DEFAULT_COMPANY
    currency: str = "USD"
    memo: str = ""
    lines: list[InvoiceLine] = field(default_factory=list)
    submit: bool = False                     # 0/1 in EIB

    @property
    def control_total(self) -> float:
        return round(sum(ln.amount for ln in self.lines), 2)


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #


_KEY_INVALID = re.compile(r"[^A-Za-z0-9_\-]")


def _spreadsheet_key(invoice_number: str, vendor: str) -> str:
    """Deterministic key. Workday rejects duplicate keys across uploads,
    so we hash vendor + invoice number for stability across re-runs.
    """
    raw = f"{(vendor or '').strip()}|{(invoice_number or '').strip()}"
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:10].upper()
    safe_inv = _KEY_INVALID.sub("-", invoice_number or "")[:24]
    return f"INV-{safe_inv}-{digest}" if safe_inv else f"INV-{digest}"


def _iso_date(s: Optional[str]) -> str:
    if not s:
        return ""
    s = str(s).strip()
    # Pass through ISO; tolerate MM/DD/YYYY by converting.
    try:
        if "/" in s:
            mm, dd, yyyy = s.split("/")
            return _dt.date(int(yyyy), int(mm), int(dd)).isoformat()
        return _dt.date.fromisoformat(s).isoformat()
    except (ValueError, TypeError):
        return s


# --------------------------------------------------------------------------- #
# Workbook builder
# --------------------------------------------------------------------------- #


def build_workbook(
    invoices: Iterable[Invoice],
    *,
    output_path: str | Path,
    skip_lines_with_no_gl: bool = True,
    skip_lines_with_no_spend_cat: bool = True,
) -> dict:
    """Write a v39.1-shape EIB workbook from a list of fully-populated invoices.

    Each invoice must have:
        - invoice_number, vendor, invoice_date
        - at least one InvoiceLine
        - every line.gl_account populated (else the line is parked unless
          skip_lines_with_no_gl=False, in which case it ships with a blank
          Ledger Account and the EIB will reject — useful for QA dry-runs)
        - every line.spend_category populated for expense GLs (same gating
          via skip_lines_with_no_spend_cat)

    Returns a summary dict including parked/dropped lines and totals.
    """
    import openpyxl

    inv_list = list(invoices)

    parked_lines: list[dict] = []  # {invoice_number, line_idx, reason}
    parked_invoices: list[dict] = []  # {invoice_number, reason}
    header_rows: list[list] = [HEADER_COLUMNS]
    line_rows: list[list] = [LINE_COLUMNS]
    total_amount = 0.0

    for inv in inv_list:
        if not inv.invoice_number or not inv.vendor or not inv.invoice_date:
            parked_invoices.append({
                "invoice_number": inv.invoice_number or "",
                "vendor": inv.vendor or "",
                "reason": "missing required header field "
                          "(invoice_number / vendor / invoice_date)",
            })
            continue

        # Build line rows first so we know if anything ships for this header.
        invoice_key = _spreadsheet_key(inv.invoice_number, inv.vendor)
        kept_lines: list[list] = []
        for idx, ln in enumerate(inv.lines, start=1):
            if not ln.gl_account and skip_lines_with_no_gl:
                parked_lines.append({
                    "invoice_number": inv.invoice_number,
                    "line": idx,
                    "amount": ln.amount,
                    "reason": "missing Ledger Account (GL classifier did not resolve)",
                })
                continue
            if not ln.spend_category and skip_lines_with_no_spend_cat:
                parked_lines.append({
                    "invoice_number": inv.invoice_number,
                    "line": idx,
                    "amount": ln.amount,
                    "reason": "missing Spend Category (ambiguous or unmapped GL)",
                })
                continue
            kept_lines.append([
                invoice_key,
                idx,
                (ln.memo or inv.memo or "")[:120],
                ln.spend_category or "",
                round(ln.quantity or 1.0, 4),
                round(ln.amount, 2),
                round((ln.quantity or 1.0) * ln.amount, 2),
                ln.gl_account or "",
                ln.cost_center or "",
                (ln.memo or "")[:120],
            ])
        if not kept_lines:
            parked_invoices.append({
                "invoice_number": inv.invoice_number,
                "vendor": inv.vendor,
                "reason": "all lines parked (no shippable lines)",
            })
            continue

        line_rows.extend(kept_lines)
        invoice_total = sum(row[6] for row in kept_lines)  # extended amount
        total_amount += invoice_total

        header_rows.append([
            invoice_key,
            1 if inv.submit else 0,
            inv.company or DEFAULT_COMPANY,
            inv.vendor,
            inv.currency or "USD",
            _iso_date(inv.invoice_date),
            _iso_date(inv.received_date or inv.invoice_date),
            inv.invoice_number,
            round(invoice_total, 2),
            (inv.memo or "")[:240],
        ])

    # Emit workbook.
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheet_h = wb.create_sheet("Submit Supplier Invoice")
    for r in header_rows:
        sheet_h.append(r)
    sheet_h.freeze_panes = "A2"

    sheet_l = wb.create_sheet("Invoice Line Replacement Data")
    for r in line_rows:
        sheet_l.append(r)
    sheet_l.freeze_panes = "A2"

    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    return {
        "status": "ok",
        "output_path": str(out_path),
        "headers_written": len(header_rows) - 1,
        "lines_written": len(line_rows) - 1,
        "total_amount": round(total_amount, 2),
        "parked_invoices": parked_invoices,
        "parked_lines": parked_lines,
        "company": DEFAULT_COMPANY,
    }


# --------------------------------------------------------------------------- #
# Classifier-aware variant
# --------------------------------------------------------------------------- #


def classify_and_build(
    invoices: Iterable[Invoice],
    *,
    output_path: str | Path,
    allow_ambiguous_spend_cat: bool = False,
) -> dict:
    """Resolve Ledger Account + Spend Category per line via the in-tree
    classifiers, then emit the workbook.

    For each line:
        - gl_account: if line.gl_account is None, call gl_classifier on
          (vendor, line.memo, amount). MCC is unknown for invoice lines
          (vendor invoice, not a card transaction), so the classifier
          falls back to memo-pattern → LLM tiers.
        - spend_category: gl_spend_category_map.lookup(gl_account).
          If status='confirmed' OR 'override' → ship.
          If status='ambiguous' → ship only if allow_ambiguous_spend_cat=True.
          If status='unmapped' → park the line.

    Lazy imports so this module stays importable in test contexts that
    don't have the classifier deps installed.
    """
    import gl_classifier  # type: ignore
    import gl_spend_category_map as _scm

    inv_list = list(invoices)
    classifier_misses: list[dict] = []
    ambiguous_lines: list[dict] = []

    for inv in inv_list:
        for idx, ln in enumerate(inv.lines, start=1):
            # 1) GL account.
            if not ln.gl_account:
                try:
                    res = gl_classifier.classify_transaction(
                        merchant_name=inv.vendor,
                        mcc_code=None,
                        memo=ln.memo or inv.memo,
                        amount=ln.amount,
                    )
                    ln.gl_account = getattr(res, "gl_account", None) or None
                except Exception as e:
                    classifier_misses.append({
                        "invoice_number": inv.invoice_number,
                        "line": idx,
                        "reason": f"gl_classifier raised: {e}",
                    })

            # 2) Spend Category.
            if ln.gl_account and not ln.spend_category:
                lookup = _scm.lookup(
                    ln.gl_account, allow_ambiguous=allow_ambiguous_spend_cat,
                )
                if lookup.status in ("confirmed", "override"):
                    ln.spend_category = lookup.spend_category
                elif lookup.status == "ambiguous" and allow_ambiguous_spend_cat:
                    ln.spend_category = lookup.spend_category
                    ambiguous_lines.append({
                        "invoice_number": inv.invoice_number,
                        "line": idx,
                        "gl_account": ln.gl_account,
                        "spend_category": ln.spend_category,
                        "dominance": lookup.dominance,
                    })
                # 'not_required' (non-expense GL) → leave blank, EIB allows.
                # 'unmapped' → leave blank; build_workbook will park it.

    summary = build_workbook(inv_list, output_path=output_path)
    summary["classifier_misses"] = classifier_misses
    summary["ambiguous_lines"] = ambiguous_lines
    return summary
