# © 2026 CoAssisted Workspace. Licensed under MIT.
"""AP-7: StaffWizard daily Overall Report ingestion.

The Overall Report ships as a 66-column .xls with one row per shift.
This module:

    1. Detects the report (filename pattern + sender heuristic).
    2. Converts .xls → .xlsx via libreoffice headless (StaffWizard's
       legacy export format isn't openpyxl-readable directly).
    3. Parses the rows into per-shift `LaborRow` records.
    4. Groups by (JobNumber, JobDescription) → resolves to project_code
       via `project_registry.resolve_by_staffwizard_job`.
    5. Writes per-project daily labor rows to Drive
       `Projects/{name}/Labor/Daily/{YYYY-MM-DD}_labor.xlsx`.
    6. Returns a structured ingest report so AP-8 can pull the
       data into the master roll-up.

The Wolfhound JE training data tells us labor lines post via the
Salaries & Wages tier (50000s for COS, 60000s for SG&A) at the
journal layer, so labor doesn't go through gl_classifier — it has
its own canonical posting rule per row's job classification. AP-7
just captures + groups + reports; AP-8 handles the master view.

Public surface:
    parse_overall_report(path)  -> ParsedReport
    group_by_project(parsed)    -> dict[project_code, ProjectLabor]
"""

from __future__ import annotations

import datetime as _dt
import re
import subprocess
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Optional

import project_registry


# =============================================================================
# Schema — column positions in the StaffWizard Overall Report
# =============================================================================

# 1-indexed column positions (matches openpyxl convention).
# Header row is R2; data starts at R3 in the official format.
_COL_JOB_NUMBER = 1
_COL_JOB_DESC = 2
_COL_TIER1_DESC = 9
_COL_TIER2_DESC = 10
_COL_POST_DESC = 15
_COL_SHIFT_START = 18
_COL_SHIFT_END = 19
_COL_WORK_DATE = 21
_COL_EMP_NUMBER = 22
_COL_EMP_NAME = 27
_COL_HOURS = 32
_COL_PAY_RATE = 33
_COL_OT_HOURS = 37
_COL_OT_RATE = 38
_COL_DBL_HOURS = 39
_COL_DBL_RATE = 40
_COL_DOLLARS = 42
_COL_HOL_DOLLARS = 43
_COL_OT_DOLLARS = 44
_COL_DBL_DOLLARS = 45
_COL_BILLABLE_HOURS = 49
_COL_BILL_RATE = 52
_COL_BILLABLE_DOLLARS = 59

HEADER_ROW = 2
DATA_START = 3


# Filename signature StaffWizard uses ("Overall Report SFOX 1777532406.xls").
_REPORT_FILENAME_RE = re.compile(
    r"^overall\s*report.*\.xls[xm]?$", re.IGNORECASE,
)


def looks_like_overall_report(path: str | Path) -> bool:
    """Filename heuristic — used by capture pipeline to triage attachments."""
    name = Path(path).name
    return bool(_REPORT_FILENAME_RE.match(name))


# =============================================================================
# Data model
# =============================================================================

@dataclass
class LaborRow:
    """One shift / one employee on one job. Mirrors a single Overall Report row."""

    job_number: str
    job_description: str
    work_date: Optional[_dt.date]
    employee_name: str
    employee_number: str
    post_description: str
    shift_start: str
    shift_end: str
    hours: float
    overtime_hours: float
    doubletime_hours: float
    dollars: float            # Regular labor cost
    holiday_dollars: float
    overtime_dollars: float
    doubletime_dollars: float
    billable_hours: float
    billable_dollars: float

    @property
    def total_cost(self) -> float:
        return (
            self.dollars
            + self.holiday_dollars
            + self.overtime_dollars
            + self.doubletime_dollars
        )

    @property
    def total_hours(self) -> float:
        return self.hours + self.overtime_hours + self.doubletime_hours

    @property
    def margin(self) -> float:
        return self.billable_dollars - self.total_cost


@dataclass
class ParsedReport:
    """All shifts in one Overall Report file."""
    rows: list[LaborRow] = field(default_factory=list)
    work_date: Optional[_dt.date] = None
    source_path: Optional[Path] = None
    skipped_no_job: int = 0
    skipped_no_date: int = 0


@dataclass
class ProjectLabor:
    """Per-project rollup of one day's labor from a single Overall Report."""

    project_code: Optional[str]      # None when project_registry doesn't have it
    job_number: str
    job_description: str
    work_date: Optional[_dt.date]
    rows: list[LaborRow] = field(default_factory=list)

    @property
    def total_cost(self) -> float:
        return sum(r.total_cost for r in self.rows)

    @property
    def total_hours(self) -> float:
        return sum(r.total_hours for r in self.rows)

    @property
    def total_revenue(self) -> float:
        return sum(r.billable_dollars for r in self.rows)

    @property
    def total_billable_hours(self) -> float:
        return sum(r.billable_hours for r in self.rows)

    @property
    def margin(self) -> float:
        return self.total_revenue - self.total_cost

    @property
    def shift_count(self) -> int:
        return len(self.rows)


# =============================================================================
# .xls → .xlsx conversion
# =============================================================================

def _convert_xls_to_xlsx(xls_path: Path) -> Path:
    """Convert a legacy .xls to .xlsx via libreoffice headless.

    StaffWizard's Overall Report uses the legacy CDF V2 format that
    openpyxl can't read directly (xlrd chokes on it too — verified
    earlier). libreoffice converts cleanly. Output lives next to the
    input with a .xlsx suffix.

    Returns the converted path, or the input unchanged if it's already
    .xlsx.
    """
    if xls_path.suffix.lower() in {".xlsx", ".xlsm"}:
        return xls_path
    out_dir = xls_path.parent
    try:
        subprocess.run(
            [
                "libreoffice", "--headless",
                "--convert-to", "xlsx",
                str(xls_path),
                "--outdir", str(out_dir),
            ],
            check=True,
            capture_output=True,
            timeout=60,
        )
    except (subprocess.SubprocessError, FileNotFoundError) as exc:
        raise RuntimeError(
            f"Failed to convert {xls_path} via libreoffice: {exc}. "
            "Install LibreOffice or pre-convert the file."
        ) from exc
    converted = out_dir / (xls_path.stem + ".xlsx")
    if not converted.exists():
        raise RuntimeError(
            f"libreoffice ran but {converted} wasn't produced."
        )
    return converted


# =============================================================================
# Parser
# =============================================================================

def _coerce_float(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _coerce_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _coerce_date(v) -> Optional[_dt.date]:
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return _dt.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def parse_overall_report(path: str | Path) -> ParsedReport:
    """Parse a StaffWizard Overall Report into LaborRow records.

    Handles both .xls (auto-converts via libreoffice) and .xlsx.
    Skips rows missing JobNumber + JobDescription (footer / header
    artifacts) or WorkDate (one-day report shouldn't have any, but
    multi-day exports can have blanks).
    """
    src = Path(path).expanduser().resolve()
    if not src.exists():
        raise FileNotFoundError(f"Overall Report not found: {src}")

    workbook_path = _convert_xls_to_xlsx(src) if src.suffix.lower() == ".xls" else src

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl required. Install: pip install openpyxl --break-system-packages"
        ) from exc

    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Overall Report"] if "Overall Report" in wb.sheetnames else wb[wb.sheetnames[0]]

    report = ParsedReport(source_path=src)
    work_dates_seen: set[_dt.date] = set()

    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        if len(row) < _COL_BILLABLE_DOLLARS:
            continue
        job_number = _coerce_str(row[_COL_JOB_NUMBER - 1])
        job_desc = _coerce_str(row[_COL_JOB_DESC - 1])
        if not job_number and not job_desc:
            report.skipped_no_job += 1
            continue
        work_date = _coerce_date(row[_COL_WORK_DATE - 1])
        if not work_date:
            report.skipped_no_date += 1
            continue
        work_dates_seen.add(work_date)

        report.rows.append(LaborRow(
            job_number=job_number,
            job_description=job_desc,
            work_date=work_date,
            employee_name=_coerce_str(row[_COL_EMP_NAME - 1]),
            employee_number=_coerce_str(row[_COL_EMP_NUMBER - 1]),
            post_description=_coerce_str(row[_COL_POST_DESC - 1]),
            shift_start=_coerce_str(row[_COL_SHIFT_START - 1]),
            shift_end=_coerce_str(row[_COL_SHIFT_END - 1]),
            hours=_coerce_float(row[_COL_HOURS - 1]),
            overtime_hours=_coerce_float(row[_COL_OT_HOURS - 1]),
            doubletime_hours=_coerce_float(row[_COL_DBL_HOURS - 1]),
            dollars=_coerce_float(row[_COL_DOLLARS - 1]),
            holiday_dollars=_coerce_float(row[_COL_HOL_DOLLARS - 1]),
            overtime_dollars=_coerce_float(row[_COL_OT_DOLLARS - 1]),
            doubletime_dollars=_coerce_float(row[_COL_DBL_DOLLARS - 1]),
            billable_hours=_coerce_float(row[_COL_BILLABLE_HOURS - 1]),
            billable_dollars=_coerce_float(row[_COL_BILLABLE_DOLLARS - 1]),
        ))

    if len(work_dates_seen) == 1:
        report.work_date = next(iter(work_dates_seen))
    return report


# =============================================================================
# Per-project grouping
# =============================================================================

def group_by_project(parsed: ParsedReport) -> dict[str, ProjectLabor]:
    """Group LaborRows by project, resolving via project_registry.

    Returns dict keyed on either the registered project_code or, when
    no registry match exists, a synthetic key `unmapped::{job_number}::
    {job_desc}` so unmapped jobs surface in the report instead of
    silently merging.
    """
    out: dict[str, ProjectLabor] = {}
    for row in parsed.rows:
        registry_match = project_registry.resolve_by_staffwizard_job(
            row.job_number, row.job_description,
        )
        if registry_match:
            key = registry_match["code"]
            project_code: Optional[str] = registry_match["code"]
        else:
            key = f"unmapped::{row.job_number}::{row.job_description}"
            project_code = None

        if key not in out:
            out[key] = ProjectLabor(
                project_code=project_code,
                job_number=row.job_number,
                job_description=row.job_description,
                work_date=row.work_date,
                rows=[],
            )
        out[key].rows.append(row)
    return out


# =============================================================================
# Per-project Drive write — one row per shift, summary footer
# =============================================================================

def write_project_labor_workbook(
    project_labor: ProjectLabor,
    output_path: str | Path,
) -> Path:
    """Write a one-day labor workbook for a single project.

    Layout: header band (row 1 title, row 2 column headers) + one row
    per shift, then a summary footer row totaling cost, revenue, hours,
    margin.
    """
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl required") from exc

    out = Path(output_path).expanduser().resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Labor"

    title = (
        f"{project_labor.project_code or '(unmapped)'} · "
        f"{project_labor.job_description} · "
        f"{project_labor.work_date.isoformat() if project_labor.work_date else 'unknown date'}"
    )
    ws.cell(row=1, column=1, value=title)
    headers = [
        "Employee", "Post", "Shift Start", "Shift End",
        "Reg Hrs", "OT Hrs", "DBL Hrs", "Total Hrs",
        "Pay Cost", "Holiday $", "OT $", "DBL $", "Total Cost",
        "Billable Hrs", "Billable $", "Margin",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)

    r = 3
    for row in project_labor.rows:
        ws.cell(row=r, column=1, value=row.employee_name)
        ws.cell(row=r, column=2, value=row.post_description)
        ws.cell(row=r, column=3, value=row.shift_start)
        ws.cell(row=r, column=4, value=row.shift_end)
        ws.cell(row=r, column=5, value=row.hours)
        ws.cell(row=r, column=6, value=row.overtime_hours)
        ws.cell(row=r, column=7, value=row.doubletime_hours)
        ws.cell(row=r, column=8, value=row.total_hours)
        ws.cell(row=r, column=9, value=row.dollars)
        ws.cell(row=r, column=10, value=row.holiday_dollars)
        ws.cell(row=r, column=11, value=row.overtime_dollars)
        ws.cell(row=r, column=12, value=row.doubletime_dollars)
        ws.cell(row=r, column=13, value=row.total_cost)
        ws.cell(row=r, column=14, value=row.billable_hours)
        ws.cell(row=r, column=15, value=row.billable_dollars)
        ws.cell(row=r, column=16, value=row.margin)
        r += 1

    # Summary footer.
    ws.cell(row=r + 1, column=1, value="TOTAL")
    ws.cell(row=r + 1, column=8, value=project_labor.total_hours)
    ws.cell(row=r + 1, column=13, value=round(project_labor.total_cost, 2))
    ws.cell(row=r + 1, column=14, value=project_labor.total_billable_hours)
    ws.cell(row=r + 1, column=15, value=round(project_labor.total_revenue, 2))
    ws.cell(row=r + 1, column=16, value=round(project_labor.margin, 2))

    wb.save(str(out))
    return out


def ingest_report(
    report_path: str | Path,
    *,
    output_dir: Optional[str | Path] = None,
) -> dict:
    """End-to-end: parse + group + write per-project workbooks.

    Returns a structured report:
        {
            "source_path":   "...",
            "work_date":     "2026-04-29",
            "shifts":        110,
            "skipped_no_job":  3,
            "skipped_no_date": 0,
            "projects": [
                {
                    "project_code": "GE1",
                    "job_number":   "Google, LLC",
                    "job_description": "Golden Eagle 1",
                    "shifts":       18,
                    "total_cost":   3247.50,
                    "total_revenue": 4980.00,
                    "margin":       1732.50,
                    "output_path":  "/.../Labor/Daily/2026-04-29_labor.xlsx",
                },
                ...
            ],
            "unmapped": [...]  # jobs without a project_registry entry
        }
    """
    parsed = parse_overall_report(report_path)
    grouped = group_by_project(parsed)

    output_root = Path(output_dir).expanduser().resolve() if output_dir else (
        Path(report_path).expanduser().resolve().parent
    )

    projects_out: list[dict] = []
    unmapped_out: list[dict] = []
    for key, pl in grouped.items():
        date_str = pl.work_date.isoformat() if pl.work_date else "unknown"
        if pl.project_code:
            slug = re.sub(r"[^\w]+", "_", pl.project_code).strip("_")
            file_dir = output_root / slug / "Labor" / "Daily"
            file_path = file_dir / f"{date_str}_labor.xlsx"
            write_project_labor_workbook(pl, file_path)
            projects_out.append({
                "project_code": pl.project_code,
                "job_number": pl.job_number,
                "job_description": pl.job_description,
                "shifts": pl.shift_count,
                "total_hours": round(pl.total_hours, 2),
                "total_cost": round(pl.total_cost, 2),
                "total_revenue": round(pl.total_revenue, 2),
                "margin": round(pl.margin, 2),
                "output_path": str(file_path),
            })
        else:
            unmapped_out.append({
                "job_number": pl.job_number,
                "job_description": pl.job_description,
                "shifts": pl.shift_count,
                "total_hours": round(pl.total_hours, 2),
                "total_cost": round(pl.total_cost, 2),
                "total_revenue": round(pl.total_revenue, 2),
                "margin": round(pl.margin, 2),
            })

    return {
        "source_path": str(parsed.source_path) if parsed.source_path else None,
        "work_date": parsed.work_date.isoformat() if parsed.work_date else None,
        "shifts": len(parsed.rows),
        "skipped_no_job": parsed.skipped_no_job,
        "skipped_no_date": parsed.skipped_no_date,
        "projects": projects_out,
        "unmapped": unmapped_out,
    }
