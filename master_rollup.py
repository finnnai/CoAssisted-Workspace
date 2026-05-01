# © 2026 CoAssisted Workspace. Licensed under MIT.
"""AP-8: Master roll-up + run-rate dashboard.

Aggregates per-project spend across three sources:
    1. Captured receipts (Surefox AP/Projects/{name}/Receipts/{YYYY-MM}/)
    2. Captured invoices (Surefox AP/Projects/{name}/Invoices/{YYYY-MM}/)
    3. StaffWizard labor (per AP-7 — Labor/Daily/{YYYY-MM-DD}_labor.xlsx)

Produces a single workbook with three tabs:

    Tab 1 — All Projects (master row per project per day)
        Columns: project_code, project_name, work_date, receipts$,
        invoices$, labor_cost$, labor_revenue$, total_cost,
        margin, run_rate_7d, run_rate_30d, baseline_status

    Tab 2 — PM Dashboard (one row per project, summary view)
        Columns: project_code, project_name, today_spend, week_spend,
        month_spend, run_rate_7d, run_rate_30d, baseline_band,
        deviation_flag, projected_30d_burn

    Tab 3 — Anomalies (rows that triggered a >2σ deviation alert)
        Columns: project_code, work_date, observed_spend,
        baseline_mean, baseline_2sigma, deviation_pct, action

Baseline-deviation model (replaces budget burn since registered budgets
don't exist yet):
    - N=30-day cold start per project — collect daily spend
      observations.
    - After cold start, compute mean + std dev of daily spend.
    - Alert when actual day's spend > baseline_mean + 2 * std_dev OR
      < baseline_mean - 2 * std_dev. Both directions matter — sudden
      drops can mean a project pause or a missing report.
    - 7-day rolling sum gets the same treatment against weekly
      baseline.

This module reads the per-project artifacts AP-7 wrote. It doesn't
re-parse statements or receipts directly — they're already
structured by Wave 1 + AP-7. Master rollup is pure aggregation.

Public surface:
    build_master_workbook(project_data, output_path) -> Path
    compute_baseline(daily_history, *, min_days=14) -> BaselineStats
    run_daily_rollup(...) -> dict   # full pipeline including PM tab
"""

from __future__ import annotations

import datetime as _dt
import json
import math
import os
import statistics
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import project_registry


# =============================================================================
# Storage — per-project daily history persists between rollups
# =============================================================================

_PROJECT_ROOT = Path(__file__).resolve().parent
_HISTORY_PATH = _PROJECT_ROOT / "master_rollup_history.json"


def _load_history() -> dict[str, dict]:
    """Layout: {project_code: {YYYY-MM-DD: {receipts, invoices, labor_cost, labor_revenue}}}."""
    if not _HISTORY_PATH.exists():
        return {}
    try:
        with _HISTORY_PATH.open("r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (json.JSONDecodeError, OSError):
        return {}


def _save_history(data: dict[str, dict]) -> None:
    _HISTORY_PATH.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(
        prefix="master_rollup_history.", suffix=".json.tmp",
        dir=str(_HISTORY_PATH.parent),
    )
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, sort_keys=True)
        os.replace(tmp, _HISTORY_PATH)
    except Exception:
        try:
            os.unlink(tmp)
        except OSError:
            pass
        raise


# =============================================================================
# Baseline statistics
# =============================================================================

# Minimum days of history required before we trust the baseline.
# Below this, every alert says "still in cold start, no signal yet."
MIN_DAYS_FOR_BASELINE = 14
COLD_START_DAYS = 30
SIGMA_ALERT_THRESHOLD = 2.0


@dataclass
class BaselineStats:
    days_observed: int
    mean_daily_spend: float
    stdev_daily_spend: float
    upper_2sigma: float
    lower_2sigma: float
    in_cold_start: bool

    @property
    def is_trustworthy(self) -> bool:
        """True only when the baseline has enough observations."""
        return not self.in_cold_start and self.days_observed >= MIN_DAYS_FOR_BASELINE


def compute_baseline(
    daily_spend_history: list[float],
    *,
    min_days: int = MIN_DAYS_FOR_BASELINE,
    cold_start: int = COLD_START_DAYS,
) -> BaselineStats:
    """Compute mean + 2σ envelope from a daily-spend list.

    `daily_spend_history` is one float per day, oldest first. We use the
    full list for now; in a later iteration we may window to last N days
    to follow seasonal drift.
    """
    n = len(daily_spend_history)
    if n == 0:
        return BaselineStats(0, 0.0, 0.0, 0.0, 0.0, in_cold_start=True)
    mean = statistics.fmean(daily_spend_history)
    if n < 2:
        sd = 0.0
    else:
        sd = statistics.stdev(daily_spend_history)
    return BaselineStats(
        days_observed=n,
        mean_daily_spend=mean,
        stdev_daily_spend=sd,
        upper_2sigma=mean + SIGMA_ALERT_THRESHOLD * sd,
        lower_2sigma=max(0.0, mean - SIGMA_ALERT_THRESHOLD * sd),
        in_cold_start=(n < cold_start),
    )


def detect_deviation(
    today_spend: float,
    baseline: BaselineStats,
) -> Optional[dict]:
    """Return an alert dict if today_spend deviates >2σ, else None."""
    if not baseline.is_trustworthy:
        return None
    if today_spend > baseline.upper_2sigma:
        return {
            "direction": "spike",
            "observed": today_spend,
            "baseline_mean": baseline.mean_daily_spend,
            "upper_2sigma": baseline.upper_2sigma,
            "deviation_pct": (
                (today_spend - baseline.mean_daily_spend)
                / max(baseline.mean_daily_spend, 0.01)
                * 100
            ),
        }
    if today_spend < baseline.lower_2sigma:
        return {
            "direction": "drop",
            "observed": today_spend,
            "baseline_mean": baseline.mean_daily_spend,
            "lower_2sigma": baseline.lower_2sigma,
            "deviation_pct": (
                (today_spend - baseline.mean_daily_spend)
                / max(baseline.mean_daily_spend, 0.01)
                * 100
            ),
        }
    return None


# =============================================================================
# Run-rate computation
# =============================================================================

def run_rate_window(
    daily_spend: list[float],
    *,
    days: int,
) -> Optional[float]:
    """Average daily spend over the last `days` observations.

    Returns None when fewer than `days` observations are available.
    """
    if len(daily_spend) < days:
        return None
    return statistics.fmean(daily_spend[-days:])


# =============================================================================
# Daily fact recording
# =============================================================================

@dataclass
class DailyFact:
    """One day's spend across all sources for one project."""
    project_code: str
    work_date: _dt.date
    receipts: float = 0.0
    invoices: float = 0.0
    labor_cost: float = 0.0
    labor_revenue: float = 0.0

    @property
    def total_cost(self) -> float:
        return self.receipts + self.invoices + self.labor_cost

    @property
    def margin(self) -> float:
        return self.labor_revenue - self.total_cost


def record_daily_fact(fact: DailyFact) -> None:
    """Persist one day's spend to the history store. Idempotent — same
    (project, date) overwrites prior value (use for corrections)."""
    history = _load_history()
    project_book = history.setdefault(fact.project_code, {})
    project_book[fact.work_date.isoformat()] = {
        "receipts": round(fact.receipts, 2),
        "invoices": round(fact.invoices, 2),
        "labor_cost": round(fact.labor_cost, 2),
        "labor_revenue": round(fact.labor_revenue, 2),
    }
    _save_history(history)


def get_daily_history(
    project_code: str,
    *,
    sort_chronological: bool = True,
) -> list[tuple[_dt.date, dict]]:
    """Return [(date, {receipts, invoices, ...}), ...] for a project."""
    history = _load_history()
    project_book = history.get(project_code) or {}
    items: list[tuple[_dt.date, dict]] = []
    for date_str, payload in project_book.items():
        try:
            d = _dt.date.fromisoformat(date_str)
        except ValueError:
            continue
        items.append((d, payload))
    if sort_chronological:
        items.sort(key=lambda kv: kv[0])
    return items


def total_cost_history(project_code: str) -> list[float]:
    """Daily total cost (receipts + invoices + labor) as a chronological list.

    Convenience for compute_baseline + run_rate_window callers.
    """
    return [
        float(p.get("receipts", 0))
        + float(p.get("invoices", 0))
        + float(p.get("labor_cost", 0))
        for _date, p in get_daily_history(project_code)
    ]


# =============================================================================
# Master workbook builder
# =============================================================================

def build_master_workbook(
    *,
    output_path: str | Path,
    target_date: Optional[_dt.date] = None,
) -> Path:
    """Build the three-tab master rollup workbook.

    target_date defaults to yesterday (typical morning refresh runs at
    6am for the prior day's data). Tabs:
        - All Projects: every (project, date) row with run-rates
        - PM Dashboard: today/week/month summary per project
        - Anomalies: rows that tripped baseline deviation
    """
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl required") from exc

    target = target_date or (_dt.date.today() - _dt.timedelta(days=1))
    out = Path(output_path).expanduser().resolve()
    out.parent.mkdir(parents=True, exist_ok=True)

    history = _load_history()
    wb = Workbook()

    # ---------- Tab 1: All Projects ----------
    ws_all = wb.active
    ws_all.title = "All Projects"
    headers_all = [
        "Project Code", "Project Name", "Work Date",
        "Receipts $", "Invoices $", "Labor Cost $", "Labor Revenue $",
        "Total Cost", "Margin",
        "Run-Rate 7d", "Run-Rate 30d",
        "Baseline Status",
    ]
    for c, h in enumerate(headers_all, start=1):
        ws_all.cell(row=1, column=c, value=h)

    # ---------- Tab 2: PM Dashboard ----------
    ws_pm = wb.create_sheet(title="PM Dashboard")
    headers_pm = [
        "Project Code", "Project Name",
        "Today Spend", "Week Spend (7d)", "Month Spend (30d)",
        "Run-Rate 7d", "Run-Rate 30d",
        "Baseline Mean ± 2σ", "Deviation Flag",
        "Projected 30d Burn",
    ]
    for c, h in enumerate(headers_pm, start=1):
        ws_pm.cell(row=1, column=c, value=h)

    # ---------- Tab 3: Anomalies ----------
    ws_anom = wb.create_sheet(title="Anomalies")
    headers_anom = [
        "Project Code", "Work Date",
        "Direction", "Observed Spend",
        "Baseline Mean", "2σ Threshold",
        "Deviation %",
        "Suggested Action",
    ]
    for c, h in enumerate(headers_anom, start=1):
        ws_anom.cell(row=1, column=c, value=h)

    row_all = 2
    row_pm = 2
    row_anom = 2

    for project_code, project_book in sorted(history.items()):
        registry = project_registry.get(project_code) or {}
        project_name = registry.get("name", project_code)

        # Chronological daily totals.
        daily_totals = total_cost_history(project_code)
        baseline = compute_baseline(daily_totals)

        # Tab 1 rows.
        for date_str, payload in sorted(project_book.items()):
            try:
                d = _dt.date.fromisoformat(date_str)
            except ValueError:
                continue
            receipts = float(payload.get("receipts", 0))
            invoices = float(payload.get("invoices", 0))
            labor_cost = float(payload.get("labor_cost", 0))
            labor_rev = float(payload.get("labor_revenue", 0))
            total_cost = receipts + invoices + labor_cost
            margin = labor_rev - total_cost

            # Slice history up to this date for trailing run-rates.
            historical_up_to = [
                float(p.get("receipts", 0))
                + float(p.get("invoices", 0))
                + float(p.get("labor_cost", 0))
                for ds, p in sorted(project_book.items())
                if ds <= date_str
            ]
            rr7 = run_rate_window(historical_up_to, days=7) or 0.0
            rr30 = run_rate_window(historical_up_to, days=30) or 0.0

            ws_all.cell(row=row_all, column=1, value=project_code)
            ws_all.cell(row=row_all, column=2, value=project_name)
            ws_all.cell(row=row_all, column=3, value=d.isoformat())
            ws_all.cell(row=row_all, column=4, value=round(receipts, 2))
            ws_all.cell(row=row_all, column=5, value=round(invoices, 2))
            ws_all.cell(row=row_all, column=6, value=round(labor_cost, 2))
            ws_all.cell(row=row_all, column=7, value=round(labor_rev, 2))
            ws_all.cell(row=row_all, column=8, value=round(total_cost, 2))
            ws_all.cell(row=row_all, column=9, value=round(margin, 2))
            ws_all.cell(row=row_all, column=10, value=round(rr7, 2))
            ws_all.cell(row=row_all, column=11, value=round(rr30, 2))
            ws_all.cell(
                row=row_all, column=12,
                value=("cold start" if baseline.in_cold_start else "established"),
            )
            row_all += 1

            # Anomaly check (only against baselines we trust).
            alert = detect_deviation(total_cost, baseline)
            if alert:
                ws_anom.cell(row=row_anom, column=1, value=project_code)
                ws_anom.cell(row=row_anom, column=2, value=d.isoformat())
                ws_anom.cell(row=row_anom, column=3, value=alert["direction"])
                ws_anom.cell(row=row_anom, column=4, value=round(alert["observed"], 2))
                ws_anom.cell(
                    row=row_anom, column=5,
                    value=round(alert["baseline_mean"], 2),
                )
                threshold = alert.get("upper_2sigma") or alert.get("lower_2sigma") or 0
                ws_anom.cell(row=row_anom, column=6, value=round(threshold, 2))
                ws_anom.cell(
                    row=row_anom, column=7,
                    value=round(alert["deviation_pct"], 1),
                )
                ws_anom.cell(
                    row=row_anom, column=8,
                    value=(
                        "Investigate spike — confirm receipts are real and "
                        "no duplicate posting."
                        if alert["direction"] == "spike"
                        else "Confirm labor / receipt feeds — sudden drop "
                        "may mean a missing source."
                    ),
                )
                row_anom += 1

        # Tab 2 PM Dashboard summary.
        target_iso = target.isoformat()
        today_payload = project_book.get(target_iso) or {}
        today_spend = (
            float(today_payload.get("receipts", 0))
            + float(today_payload.get("invoices", 0))
            + float(today_payload.get("labor_cost", 0))
        )
        rr7 = run_rate_window(daily_totals, days=7) or 0.0
        rr30 = run_rate_window(daily_totals, days=30) or 0.0
        week_total = sum(daily_totals[-7:])
        month_total = sum(daily_totals[-30:])
        baseline_str = (
            f"${baseline.mean_daily_spend:.2f} ± ${baseline.stdev_daily_spend * 2:.2f}"
            if baseline.is_trustworthy
            else f"cold start ({baseline.days_observed}/{COLD_START_DAYS} days)"
        )
        deviation = detect_deviation(today_spend, baseline)
        deviation_flag = (
            f"{deviation['direction']} ({deviation['deviation_pct']:.0f}%)"
            if deviation else "in band"
        )
        projected_30d = rr30 * 30 if rr30 else 0.0

        ws_pm.cell(row=row_pm, column=1, value=project_code)
        ws_pm.cell(row=row_pm, column=2, value=project_name)
        ws_pm.cell(row=row_pm, column=3, value=round(today_spend, 2))
        ws_pm.cell(row=row_pm, column=4, value=round(week_total, 2))
        ws_pm.cell(row=row_pm, column=5, value=round(month_total, 2))
        ws_pm.cell(row=row_pm, column=6, value=round(rr7, 2))
        ws_pm.cell(row=row_pm, column=7, value=round(rr30, 2))
        ws_pm.cell(row=row_pm, column=8, value=baseline_str)
        ws_pm.cell(row=row_pm, column=9, value=deviation_flag)
        ws_pm.cell(row=row_pm, column=10, value=round(projected_30d, 2))
        row_pm += 1

    wb.save(str(out))
    return out
